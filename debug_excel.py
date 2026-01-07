import openpyxl
import sys
import os

filepath = 'Eindafrekening/input_master.xlsx'
if not os.path.exists(filepath):
    print(f"File not found: {filepath}")
    sys.exit(1)

wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb['Input'] if 'Input' in wb.sheetnames else wb.active

# Print header row (row 1)
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print("Headers:")
for i, h in enumerate(headers):
    print(f"{i}: {h}")

print("\nFirst Data Row (Basis):")
# Find first Basis row
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1] == 'Basis':
        for i, val in enumerate(row):
            print(f"{i}: {val} ({headers[i] if i < len(headers) else '?'})")
        break
