#!/usr/bin/env python3
"""
Create Sample Filled Excel Templates

Creates realistic example templates for testing and demonstration.
"""

import openpyxl
from datetime import date, timedelta
import shutil
from openpyxl.cell import MergedCell


def safe_write_cell(ws, cell_ref, value):
    """Safely write to a cell, unmerging if necessary"""
    try:
        cell = ws[cell_ref]
        # Check if it's a merged cell
        if isinstance(cell, MergedCell):
            # Find and unmerge the range containing this cell
            for merged_range in list(ws.merged_cells.ranges):
                if cell.coordinate in merged_range:
                    ws.unmerge_cells(str(merged_range))
                    break
        # Now write the value
        ws[cell_ref] = value
    except Exception as e:
        print(f"Warning: Could not write to {cell_ref}: {e}")


def create_sample_1_normal_case():
    """Sample 1: Normal case - small refund, no extras"""

    # Copy template
    shutil.copy('input_template.xlsx', 'sample_1_normale_afrekening.xlsx')
    wb = openpyxl.load_workbook('sample_1_normale_afrekening.xlsx')

    # Fill using named ranges for safety
    def set_named_value(name, value):
        try:
            cells = wb.defined_names[name].destinations
            for sheet_name, cell_range in cells:
                ws = wb[sheet_name]
                # Get first cell from range
                cell_coord = cell_range.split(':')[0]
                ws[cell_coord] = value
        except:
            pass  # Skip if named range doesn't exist

    # Client info
    set_named_value('Klantnaam', 'Familie Janssen')
    set_named_value('Contactpersoon', 'Dhr. P. Janssen')
    set_named_value('Email', 'p.janssen@email.nl')
    set_named_value('Telefoonnummer', '+31 6 12345678')

    # Object info
    set_named_value('Object_adres', 'Strandweg 42')
    set_named_value('Unit_nr', 'A3')
    set_named_value('Postcode', '1234 AB')
    set_named_value('Plaats', 'Zandvoort')
    set_named_value('Object_ID', 'OBJ-2024-042')

    # Period
    set_named_value('Incheck_datum', date(2024, 7, 1))
    set_named_value('Uitcheck_datum', date(2024, 7, 15))

    # Voorschotten
    set_named_value('Voorschot_borg', 800.00)
    set_named_value('Voorschot_GWE', 350.00)
    set_named_value('Voorschot_schoonmaak', 250.00)
    set_named_value('Overige_voorschotten', 0.00)

    # Contract
    set_named_value('Schoonmaak_pakket', '5_uur')
    set_named_value('Uurtarief_schoonmaak', 50.00)
    set_named_value('Meterbeheerder', 'Stedin')
    set_named_value('Energie_leverancier', 'Vattenfall')
    set_named_value('Contractnummer', 'VATT-123456')

    # Internal (optional)
    set_named_value('RR_Klantnummer', 'KL-2024-089')
    set_named_value('RR_Folder_link', 'https://onedrive.com/janssen_juli2024')
    set_named_value('RR_Projectleider', 'Anna de Vries')
    set_named_value('RR_Inspecteur', 'Mark Peters')
    set_named_value('RR_Factuurnummer', 'FACT-2024-0421')

    # GWE_Detail sheet - Meter readings
    set_named_value('KWh_begin', 10000)
    set_named_value('KWh_eind', 11200)
    set_named_value('Gas_begin', 5000.00)
    set_named_value('Gas_eind', 5145.00)

    ws_gwe = wb['GWE_Detail']

    # Add real cost rows (overwriting examples)
    # Find row after instructions (should be around row 11-12)
    start_row = 12

    # Add real cost rows using safe_write_cell
    safe_write_cell(ws_gwe, f'A{start_row}', 'Elektra verbruik')
    safe_write_cell(ws_gwe, f'B{start_row}', '=KWh_verbruik')
    safe_write_cell(ws_gwe, f'C{start_row}', 0.28)
    safe_write_cell(ws_gwe, f'D{start_row}', f'=B{start_row}*C{start_row}')

    safe_write_cell(ws_gwe, f'A{start_row+1}', 'Gas verbruik')
    safe_write_cell(ws_gwe, f'B{start_row+1}', '=Gas_verbruik')
    safe_write_cell(ws_gwe, f'C{start_row+1}', 1.15)
    safe_write_cell(ws_gwe, f'D{start_row+1}', f'=B{start_row+1}*C{start_row+1}')

    safe_write_cell(ws_gwe, f'A{start_row+2}', 'Vaste levering elektra')
    safe_write_cell(ws_gwe, f'B{start_row+2}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+2}', 0.48)
    safe_write_cell(ws_gwe, f'D{start_row+2}', f'=B{start_row+2}*C{start_row+2}')

    safe_write_cell(ws_gwe, f'A{start_row+3}', 'Vaste levering gas')
    safe_write_cell(ws_gwe, f'B{start_row+3}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+3}', 0.42)
    safe_write_cell(ws_gwe, f'D{start_row+3}', f'=B{start_row+3}*C{start_row+3}')

    safe_write_cell(ws_gwe, f'A{start_row+4}', 'Waterverbruik')
    safe_write_cell(ws_gwe, f'B{start_row+4}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+4}', 3.20)
    safe_write_cell(ws_gwe, f'D{start_row+4}', f'=B{start_row+4}*C{start_row+4}')

    # Schoonmaak sheet
    set_named_value('Totaal_uren_gew', 5.5)

    # Schade sheet - NO DAMAGE (leave empty)

    wb.save('sample_1_normale_afrekening.xlsx')
    print("✅ Sample 1 created: sample_1_normale_afrekening.xlsx")
    print("   Scenario: Normale afrekening, klein GWE meerverbruik, weinig extra schoonmaak, geen schade")


def create_sample_2_extra_kosten():
    """Sample 2: Extra costs - overage, extra cleaning, damage"""

    # Copy template
    shutil.copy('input_template.xlsx', 'sample_2_extra_kosten.xlsx')
    wb = openpyxl.load_workbook('sample_2_extra_kosten.xlsx')

    # Fill using named ranges
    def set_named_value(name, value):
        try:
            cells = wb.defined_names[name].destinations
            for sheet_name, cell_range in cells:
                ws = wb[sheet_name]
                cell_coord = cell_range.split(':')[0]
                ws[cell_coord] = value
        except:
            pass

    # Client info
    set_named_value('Klantnaam', 'Familie de Groot')
    set_named_value('Contactpersoon', 'Mevr. S. de Groot')
    set_named_value('Email', 's.degroot@provider.com')
    set_named_value('Telefoonnummer', '+31 6 98765432')

    # Object info
    set_named_value('Object_adres', 'Duinweg 15')
    set_named_value('Postcode', '2042 LM')
    set_named_value('Plaats', 'Bergen aan Zee')
    set_named_value('Object_ID', 'OBJ-2024-018')

    # Period
    set_named_value('Incheck_datum', date(2024, 8, 10))
    set_named_value('Uitcheck_datum', date(2024, 8, 24))

    # Voorschotten
    set_named_value('Voorschot_borg', 1000.00)
    set_named_value('Voorschot_GWE', 400.00)
    set_named_value('Voorschot_schoonmaak', 300.00)

    # Contract
    set_named_value('Schoonmaak_pakket', '7_uur')
    set_named_value('Uurtarief_schoonmaak', 55.00)
    set_named_value('Meterbeheerder', 'Liander')
    set_named_value('Energie_leverancier', 'Eneco')
    set_named_value('Contractnummer', 'ENEC-789012')

    # Internal
    set_named_value('RR_Klantnummer', 'KL-2024-156')
    set_named_value('RR_Projectleider', 'Anna de Vries')
    set_named_value('RR_Inspecteur', 'Lisa van Dam')

    # GWE - HIGH USAGE
    set_named_value('KWh_begin', 8500)
    set_named_value('KWh_eind', 10200)  # 1700 kWh!
    set_named_value('Gas_begin', 4200.00)
    set_named_value('Gas_eind', 4450.00)  # 250 m³!

    ws_gwe = wb['GWE_Detail']

    # Clear examples and add real data
    start_row = 12
    safe_write_cell(ws_gwe, f'A{start_row}', 'Elektra verbruik')
    safe_write_cell(ws_gwe, f'B{start_row}', '=KWh_verbruik')
    safe_write_cell(ws_gwe, f'C{start_row}', 0.32)
    safe_write_cell(ws_gwe, f'D{start_row}', f'=B{start_row}*C{start_row}')

    safe_write_cell(ws_gwe, f'A{start_row+1}', 'Gas verbruik')
    safe_write_cell(ws_gwe, f'B{start_row+1}', '=Gas_verbruik')
    safe_write_cell(ws_gwe, f'C{start_row+1}', 1.25)
    safe_write_cell(ws_gwe, f'D{start_row+1}', f'=B{start_row+1}*C{start_row+1}')

    safe_write_cell(ws_gwe, f'A{start_row+2}', 'Vaste levering elektra')
    safe_write_cell(ws_gwe, f'B{start_row+2}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+2}', 0.52)
    safe_write_cell(ws_gwe, f'D{start_row+2}', f'=B{start_row+2}*C{start_row+2}')

    safe_write_cell(ws_gwe, f'A{start_row+3}', 'Vaste levering gas')
    safe_write_cell(ws_gwe, f'B{start_row+3}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+3}', 0.45)
    safe_write_cell(ws_gwe, f'D{start_row+3}', f'=B{start_row+3}*C{start_row+3}')

    safe_write_cell(ws_gwe, f'A{start_row+4}', 'Waterverbruik')
    safe_write_cell(ws_gwe, f'B{start_row+4}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+4}', 3.50)
    safe_write_cell(ws_gwe, f'D{start_row+4}', f'=B{start_row+4}*C{start_row+4}')

    # Schoonmaak - LOTS OF EXTRA HOURS
    set_named_value('Totaal_uren_gew', 11.5)  # 7 inbegrepen, 4.5 extra!

    # Schade sheet - MULTIPLE ITEMS
    ws_schade = wb['Schade']
    damages = [
        ('Gebroken raam keuken', 1, 175.00),
        ('Beschadigde vloerbedekking woonkamer', 3, 55.00),
        ('Vervangen handdoeken set', 2, 28.00),
        ('Schoonmaak verwijderen vlekken muur', 1, 85.00),
        ('Reparatie deurklink slaapkamer', 1, 42.00),
    ]

    for idx, (desc, aantal, tarief) in enumerate(damages, start=6):
        ws_schade[f'A{idx}'] = desc
        ws_schade[f'B{idx}'] = aantal
        ws_schade[f'C{idx}'] = tarief

    wb.save('sample_2_extra_kosten.xlsx')
    print("✅ Sample 2 created: sample_2_extra_kosten.xlsx")
    print("   Scenario: Veel meerverbruik GWE, veel extra schoonmaak uren, meerdere schadeposten")


def create_sample_3_perfect_stay():
    """Sample 3: Perfect stay - everything within budget, refund"""

    # Copy template
    shutil.copy('input_template.xlsx', 'sample_3_perfecte_afrekening.xlsx')
    wb = openpyxl.load_workbook('sample_3_perfecte_afrekening.xlsx')

    # Fill using named ranges
    def set_named_value(name, value):
        try:
            cells = wb.defined_names[name].destinations
            for sheet_name, cell_range in cells:
                ws = wb[sheet_name]
                cell_coord = cell_range.split(':')[0]
                ws[cell_coord] = value
        except:
            pass

    # Client info
    set_named_value('Klantnaam', 'Familie Vermeulen')
    set_named_value('Contactpersoon', 'Dhr. J. Vermeulen')
    set_named_value('Email', 'j.vermeulen@mail.com')
    set_named_value('Telefoonnummer', '+31 6 55544433')

    # Object info
    set_named_value('Object_adres', 'Zeezichtlaan 88')
    set_named_value('Unit_nr', 'B2')
    set_named_value('Postcode', '1971 NK')
    set_named_value('Plaats', 'IJmuiden')
    set_named_value('Object_ID', 'OBJ-2024-031')

    # Period - SHORT STAY
    set_named_value('Incheck_datum', date(2024, 9, 5))
    set_named_value('Uitcheck_datum', date(2024, 9, 12))

    # Voorschotten
    set_named_value('Voorschot_borg', 600.00)
    set_named_value('Voorschot_GWE', 300.00)
    set_named_value('Voorschot_schoonmaak', 200.00)

    # Contract
    set_named_value('Schoonmaak_pakket', '5_uur')
    set_named_value('Uurtarief_schoonmaak', 50.00)
    set_named_value('Meterbeheerder', 'Stedin')
    set_named_value('Energie_leverancier', 'Essent')
    set_named_value('Contractnummer', 'ESS-456789')

    # Internal
    set_named_value('RR_Klantnummer', 'KL-2024-201')
    set_named_value('RR_Projectleider', 'Anna de Vries')
    set_named_value('RR_Inspecteur', 'Mark Peters')

    # GWE - LOW USAGE
    set_named_value('KWh_begin', 12000)
    set_named_value('KWh_eind', 12580)  # Only 580 kWh
    set_named_value('Gas_begin', 6000.00)
    set_named_value('Gas_eind', 6065.00)  # Only 65 m³

    ws_gwe = wb['GWE_Detail']

    # Clear examples and add real data
    start_row = 12
    safe_write_cell(ws_gwe, f'A{start_row}', 'Elektra verbruik')
    safe_write_cell(ws_gwe, f'B{start_row}', '=KWh_verbruik')
    safe_write_cell(ws_gwe, f'C{start_row}', 0.29)
    safe_write_cell(ws_gwe, f'D{start_row}', f'=B{start_row}*C{start_row}')

    safe_write_cell(ws_gwe, f'A{start_row+1}', 'Gas verbruik')
    safe_write_cell(ws_gwe, f'B{start_row+1}', '=Gas_verbruik')
    safe_write_cell(ws_gwe, f'C{start_row+1}', 1.18)
    safe_write_cell(ws_gwe, f'D{start_row+1}', f'=B{start_row+1}*C{start_row+1}')

    safe_write_cell(ws_gwe, f'A{start_row+2}', 'Vaste levering elektra')
    safe_write_cell(ws_gwe, f'B{start_row+2}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+2}', 0.47)
    safe_write_cell(ws_gwe, f'D{start_row+2}', f'=B{start_row+2}*C{start_row+2}')

    safe_write_cell(ws_gwe, f'A{start_row+3}', 'Vaste levering gas')
    safe_write_cell(ws_gwe, f'B{start_row+3}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+3}', 0.40)
    safe_write_cell(ws_gwe, f'D{start_row+3}', f'=B{start_row+3}*C{start_row+3}')

    safe_write_cell(ws_gwe, f'A{start_row+4}', 'Waterverbruik')
    safe_write_cell(ws_gwe, f'B{start_row+4}', '=Algemeen!Aantal_dagen')
    safe_write_cell(ws_gwe, f'C{start_row+4}', 3.00)
    safe_write_cell(ws_gwe, f'D{start_row+4}', f'=B{start_row+4}*C{start_row+4}')

    # Schoonmaak sheet - EXACTLY ON TIME
    ws_clean = wb['Schoonmaak']
    ws_clean['B7'] = 5.0  # Exactly 5 hours, no extras!

    # Schade sheet - NO DAMAGE
    # Leave empty

    wb.save('sample_3_perfecte_afrekening.xlsx')
    print("✅ Sample 3 created: sample_3_perfecte_afrekening.xlsx")
    print("   Scenario: Perfecte afrekening, weinig verbruik, geen extra kosten, geen schade - klant krijgt geld terug!")


if __name__ == "__main__":
    print("🏗️  Creating Sample Filled Excel Templates...")
    print("=" * 60)

    create_sample_1_normal_case()
    create_sample_2_extra_kosten()
    create_sample_3_perfect_stay()

    print("\n" + "=" * 60)
    print("✅ All samples created!")
    print("\n📋 Samples:")
    print("   1. sample_1_normale_afrekening.xlsx - Normale case")
    print("   2. sample_2_extra_kosten.xlsx - Veel extra kosten")
    print("   3. sample_3_perfecte_afrekening.xlsx - Perfect verblijf")
    print("\n💡 Open deze bestanden om te zien hoe het template werkt!")
