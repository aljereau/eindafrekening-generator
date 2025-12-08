import openpyxl
from openpyxl.utils import get_column_letter
import sqlite3
import os

# Paths
db_path = "database/ryanrent_core.db"
template_path = "Eindafrekening/src/input_template.xlsx"
output_path = "Eindafrekening/src/input_template_v2.xlsx" # Temporary output

print(f"Updating template: {template_path}")

# 1. Fetch House Data
conn = sqlite3.connect(db_path)
cursor = conn.cursor()
cursor.execute("SELECT adres, postcode, plaats FROM huizen WHERE status='active'")
houses = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
conn.close()
print(f"Loaded {len(houses)} houses from DB.")

# 2. Load Workbook
wb = openpyxl.load_workbook(template_path)

# --- Update Lists Sheet ---
if 'Lists' in wb.sheetnames:
    ws_lists = wb['Lists']
    # Assuming Col A is Address. We will add Postcode to Col G (7) and Plaats to Col H (8)
    # We need to iterate rows and match address
    print("Updating Lists sheet...")
    
    # Add Headers
    ws_lists.cell(row=1, column=7).value = "Postcode"
    ws_lists.cell(row=1, column=8).value = "Plaats"
    
    for row in ws_lists.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        address = str(cell.value).strip() if cell.value else ""
        if address in houses:
            postcode, plaats = houses[address]
            ws_lists.cell(row=cell.row, column=7).value = postcode
            ws_lists.cell(row=cell.row, column=8).value = plaats
            
# --- Update Algemeen Sheet ---
ws_alg = wb['Algemeen']
print("Updating Algemeen sheet...")

# 1. Add VLOOKUPs for Postcode/Plaats
# Postcode is B13, Plaats is B14. Address is B10.
ws_alg['B13'] = '=IFERROR(VLOOKUP(B10,Lists!$A:$H,7,FALSE),"")'
ws_alg['B14'] = '=IFERROR(VLOOKUP(B10,Lists!$A:$H,8,FALSE),"")'

# 2. Add GWE Maandbedrag
# Insert row at 24
ws_alg.insert_rows(24)
ws_alg['A24'] = "GWE Maandbedrag"
ws_alg['B24'] = 0 # Default

# Update Voorschot GWE formula (now at B25)
# Aantal dagen is at B20
ws_alg['B25'] = '=IF(ISNUMBER(B24), (B24 * 12 / 365) * B20, 0)'

# Save
wb.save(output_path)
print(f"Saved updated template to: {output_path}")

# Replace original?
# os.replace(output_path, template_path) 
