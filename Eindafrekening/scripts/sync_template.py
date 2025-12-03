#!/usr/bin/env python3
"""
Upgrade Excel Template Script

This script upgrades the 'input_template.xlsx' with:
1. Client & Object Dropdowns (Data Validation) backed by a hidden 'Data' sheet.
2. VLOOKUP formulas to prefill client/object details.
3. 'Water' meter reading columns in GWE_Detail.
4. 'BTW %' column in cost tables.
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import sqlite3
import os
import sys

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(BASE_DIR, "src", "input_template.xlsx")
DB_PATH = os.path.join(os.path.dirname(BASE_DIR), "Shared", "database", "ryanrent_core.db")

def get_db_data():
    """Fetch Clients and Houses from SQLite DB"""
    if not os.path.exists(DB_PATH):
        print(f"⚠️ Database not found at {DB_PATH}")
        return [], []
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Fetch Clients
    cursor.execute("SELECT naam, contactpersoon, email, telefoonnummer FROM klanten ORDER BY naam")
    clients = cursor.fetchall()
    
    # Fetch Houses
    cursor.execute("SELECT adres, object_id, postcode, plaats, woning_type FROM huizen WHERE status='active' ORDER BY adres")
    houses = cursor.fetchall()
    
    conn.close()
    return clients, houses

def upgrade_template():
    print(f"📂 Opening template: {TEMPLATE_PATH}")
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        print("❌ Template file not found!")
        return

    # 1. Setup DATA Sheet
    if "Data" not in wb.sheetnames:
        ws_data = wb.create_sheet("Data")
        ws_data.sheet_state = "hidden" # Hide it
    else:
        ws_data = wb["Data"]
        # Clear existing data
        ws_data.delete_rows(1, ws_data.max_row)
    
    clients, houses = get_db_data()
    
    # Write Clients (Cols A-D)
    ws_data["A1"] = "Klantnaam"
    ws_data["B1"] = "Contactpersoon"
    ws_data["C1"] = "Email"
    ws_data["D1"] = "Telefoon"
    
    for i, c in enumerate(clients, 2):
        ws_data[f"A{i}"] = c[0]
        ws_data[f"B{i}"] = c[1]
        ws_data[f"C{i}"] = c[2]
        ws_data[f"D{i}"] = c[3]
        
    # Write Houses (Cols G-K)
    ws_data["G1"] = "Adres"
    ws_data["H1"] = "Unit" # We don't have unit in DB explicitly usually, mapped to object_id or part of address? 
                           # Actually DB has object_id, adres, postcode, plaats. 
                           # Let's map: Adres -> Adres, Unit -> ?, Postcode -> Postcode, Plaats -> Plaats, ID -> Object_ID
    ws_data["I1"] = "Postcode"
    ws_data["J1"] = "Plaats"
    ws_data["K1"] = "Object ID"
    
    for i, h in enumerate(houses, 2):
        ws_data[f"G{i}"] = h[0] # Adres
        ws_data[f"H{i}"] = ""   # Unit (Empty for now)
        ws_data[f"I{i}"] = h[2] # Postcode
        ws_data[f"J{i}"] = h[3] # Plaats
        ws_data[f"K{i}"] = h[1] # Object ID

    print(f"✅ Populated Data sheet with {len(clients)} clients and {len(houses)} houses.")

    # 2. Add Data Validation & VLOOKUPs to 'Algemeen'
    ws_alg = wb["Algemeen"]
    
    # Client Dropdown (B4)
    dv_client = DataValidation(type="list", formula1=f"=Data!$A$2:$A${len(clients)+1}", allow_blank=True)
    ws_alg.add_data_validation(dv_client)
    dv_client.add(ws_alg["B4"])
    
    # Client VLOOKUPs
    # B5 (Contact): =IFERROR(VLOOKUP(B4,Data!A:D,2,FALSE),"")
    ws_alg["B5"] = '=IFERROR(VLOOKUP(B4,Data!A:D,2,FALSE),"")'
    ws_alg["B6"] = '=IFERROR(VLOOKUP(B4,Data!A:D,3,FALSE),"")'
    ws_alg["B7"] = '=IFERROR(VLOOKUP(B4,Data!A:D,4,FALSE),"")'
    
    # Object Dropdown (B11)
    dv_house = DataValidation(type="list", formula1=f"=Data!$G$2:$G${len(houses)+1}", allow_blank=True)
    ws_alg.add_data_validation(dv_house)
    dv_house.add(ws_alg["B11"])
    
    # Object VLOOKUPs
    # B12 (Unit): =IFERROR(VLOOKUP(B11,Data!G:K,2,FALSE),"")
    ws_alg["B12"] = '=IFERROR(VLOOKUP(B11,Data!G:K,2,FALSE),"")'
    ws_alg["B13"] = '=IFERROR(VLOOKUP(B11,Data!G:K,3,FALSE),"")'
    ws_alg["B14"] = '=IFERROR(VLOOKUP(B11,Data!G:K,4,FALSE),"")'
    ws_alg["B15"] = '=IFERROR(VLOOKUP(B11,Data!G:K,5,FALSE),"")'
    
    print("✅ Added Dropdowns and VLOOKUPs to 'Algemeen'.")

    # 3. Add Water Columns to 'GWE_Detail'
    ws_gwe = wb["GWE_Detail"]
    
    # Define styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") # Blue
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_font = Font(bold=True)
    
    # Check if Water headers exist, but apply styling regardless
    # Merge Header
    ws_gwe.merge_cells("J2:L2")
    cell = ws_gwe["J2"]
    cell.value = "Water"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    
    # Subheaders
    headers = ["Beginstand", "Eindstand", "Verbruik (m³)"]
    for col, title in enumerate(headers, 10): # J=10
        cell = ws_gwe.cell(row=3, column=col)
        cell.value = title
        cell.font = subheader_font
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
        # Data cell (row 4)
        data_cell = ws_gwe.cell(row=4, column=col)
        data_cell.border = thin_border
        # Yellow background for input cells (J4, K4)
        if col < 12:
            data_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Formula for Verbruik
    ws_gwe["L4"] = "=IF(AND(ISNUMBER(K4),ISNUMBER(J4)), K4-J4, 0)"
    
    # Create Named Ranges (Idempotent-ish)
    try:
        new_range = openpyxl.workbook.defined_name.DefinedName("Water_begin", attr_text="GWE_Detail!$J$4")
        wb.defined_names.add(new_range)
        new_range = openpyxl.workbook.defined_name.DefinedName("Water_eind", attr_text="GWE_Detail!$K$4")
        wb.defined_names.add(new_range)
        new_range = openpyxl.workbook.defined_name.DefinedName("Water_verbruik", attr_text="GWE_Detail!$L$4")
        wb.defined_names.add(new_range)
    except:
        pass # Ignore if exists or fails
    
    print("✅ Updated Water columns styling in 'GWE_Detail'.")

    # 4. Add BTW Column to Cost Tables
    # GWE Table (Row 11 Header, 12+ Data)
    # Match style of existing headers (Green background usually)
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Always apply style to header
    cell = ws_gwe["E11"]
    cell.value = "BTW %"
    cell.font = Font(bold=True)
    cell.fill = green_fill # Match existing header color
    cell.border = thin_border
    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    
    # Set default 21% for existing rows
    for row in range(12, 50):
        cell = ws_gwe[f"E{row}"]
        if not cell.value: # Only set if empty to preserve user data if any
            cell.value = 0.21
        cell.number_format = '0%'
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
    print("✅ Updated BTW % column in GWE table.")
        
    # Damage Table (Schade Sheet, Row 4 Header, 5+ Data)
    ws_schade = wb["Schade"]
    
    # Always apply style to header
    cell = ws_schade["E4"]
    cell.value = "BTW %"
    cell.font = Font(bold=True)
    cell.fill = green_fill # Match existing header color
    cell.border = thin_border
    cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    
    # Set default 21% for existing rows
    for row in range(5, 50):
        cell = ws_schade[f"E{row}"]
        if not cell.value:
            cell.value = 0.21
        cell.number_format = '0%'
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
    print("✅ Updated BTW % column in Schade table.")

    # Save
    wb.save(TEMPLATE_PATH)
    print(f"💾 Saved upgraded template to: {TEMPLATE_PATH}")

if __name__ == "__main__":
    upgrade_template()
