#!/usr/bin/env python3
"""
Update Excel Template with Latest Properties
Syncs property data from ryanrent_core.db to input_template.xlsx
"""

import os
import sys
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# Add Shared directory to path
current_dir = os.path.dirname(os.path.abspath(__file__))
eindafrekening_dir = os.path.dirname(current_dir)
root_dir = os.path.dirname(eindafrekening_dir)
shared_dir = os.path.join(root_dir, 'Shared')
sys.path.append(shared_dir)

from database import Database

def update_excel_template():
    """Update input_template.xlsx with latest properties from Huizeninventaris"""
    
    # Paths
    source_excel = os.path.join(root_dir, 'Ryanrent Sales - Woningen Status  V2.1 - huisinventaris update.xlsm')
    target_excel = os.path.join(eindafrekening_dir, 'input_template.xlsx')
    
    if not os.path.exists(source_excel):
        print(f"❌ Source Excel not found: {source_excel}")
        return
    
    if not os.path.exists(target_excel):
        print(f"❌ Target template not found: {target_excel}")
        return
    
    print("\n" + "="*70)
    print("🔄 Updating Excel Template with Latest Properties")
    print("="*70)
    
    # Load source workbook (Huizeninventaris)
    print(f"\n📊 Reading properties from Huizeninventaris sheet...")
    source_wb = openpyxl.load_workbook(source_excel, read_only=True, data_only=True)
    
    if 'Huisinventaris' not in source_wb.sheetnames:
        print(f"❌ Huisinventaris sheet not found")
        return
    
    source_sheet = source_wb['Huisinventaris']
    
    # Read properties (skip header row)
    # Columns: OBJECT_ID (0), Address (1), Postcode (2), Plaats (3)
    properties = []
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:  # If address exists
            properties.append({
                'object_id': row[0] or '',
                'address': str(row[1]).strip(),
                'postcode': row[2] or '',
                'city': row[3] or ''
            })
    
    source_wb.close()
    
    print(f"   ✅ Found {len(properties)} properties in Huizeninventaris")
    
    # Load target workbook
    target_wb = openpyxl.load_workbook(target_excel)
    
    # Create or get PropertyList sheet
    if 'PropertyList' in target_wb.sheetnames:
        print("   Updating existing PropertyList sheet...")
        ws = target_wb['PropertyList']
        # Clear existing data
        ws.delete_rows(2, ws.max_row)
    else:
        print("   Creating new PropertyList sheet...")
        ws = target_wb.create_sheet('PropertyList')
        # Add headers
        ws['A1'] = 'Object_ID'
        ws['B1'] = 'Address'
        ws['C1'] = 'Postcode'
        ws['D1'] = 'City'
        
        # Format headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
    
    # Write property data
    for idx, prop in enumerate(properties, start=2):
        ws[f'A{idx}'] = prop['object_id']
        ws[f'B{idx}'] = prop['address']
        ws[f'C{idx}'] = prop['postcode']
        ws[f'D{idx}'] = prop['city']
    
    # Hide the PropertyList sheet (keeps it clean)
    ws.sheet_state = 'hidden'
    
    # Get the main Input sheet
    if 'Algemeen' in target_wb.sheetnames:
        input_sheet = target_wb['Algemeen']
    elif 'Input' in target_wb.sheetnames:
        input_sheet = target_wb['Input']
    else:
        input_sheet = target_wb.worksheets[0]
    
    print(f"   ✅ Updated PropertyList with {len(properties)} properties")
    print(f"   🔧 Configuring dropdowns and formulas in '{input_sheet.title}' sheet...")
    
    # Add Data Validation dropdown to Object_adres (B10)
    dv = DataValidation(
        type="list",
        formula1="PropertyList!$B$2:$B$999",
        allow_blank=False
    )
    dv.error = 'Please select a property from the list'
    dv.errorTitle = 'Invalid Property'
    dv.prompt = 'Select a property from the dropdown'
    dv.promptTitle = 'Property Selection'
    
    input_sheet.add_data_validation(dv)
    dv.add(input_sheet['B10'])  # Object_adres cell
    
    # Clear existing value in B10
    if input_sheet['B10'].value:
        print(f"      Cleared existing value: {input_sheet['B10'].value}")
    input_sheet['B10'].value = None
    
    # Add VLOOKUP formulas for auto-fill
    # Object ID (B14) - looks up address in B10 and returns Object_ID
    input_sheet['B14'].value = '=IF(B10="","",VLOOKUP(B10,PropertyList!$B$2:$D$999,1,FALSE))'
    
    # Postcode (B12) - looks up address and returns Postcode
    input_sheet['B12'].value = '=IF(B10="","",VLOOKUP(B10,PropertyList!$B$2:$D$999,2,FALSE))'
    
    # City (B13) - looks up address and returns City  
    input_sheet['B13'].value = '=IF(B10="","",VLOOKUP(B10,PropertyList!$B$2:$D$999,3,FALSE))'
    
    # Add comment to guide user
    from openpyxl.comments import Comment
    input_sheet['B10'].comment = Comment(
        "Select a property from the dropdown.\nObject ID, Postcode, and City will auto-fill.",
        "RyanRent System"
    )
    
    print(f"      ✅ Added dropdown to cell B10 (Object_adres)")
    print(f"      ✅ Added auto-fill formulas:")
    print(f"         - B14 (Object ID)")
    print(f"         - B12 (Postcode)")
    print(f"         - B13 (City)")
    
    # Save workbook
    target_wb.save(target_excel)
    print(f"\n✅ Excel template updated successfully!")
    print(f"   File: {target_excel}")
    print("\n📝 How to Use:")
    print("   1. Open input_template.xlsx")
    print("   2. In cell B10 (Adres Object), click the dropdown arrow")
    print("   3. Select a property from the list")
    print("   4. Object ID, Postcode, and City will auto-fill automatically!")
    print("\n" + "="*70)

if __name__ == "__main__":
    try:
        update_excel_template()
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
