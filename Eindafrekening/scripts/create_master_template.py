import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
import os
import sys

# Add Shared to path for Database
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(root_dir, 'Shared'))
from database import Database
from datetime import datetime

def create_master_template(output_path=None):
    # Default to Eindafrekening Inputs folder with timestamp
    if output_path is None:
        eindafrekening_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        inputs_folder = os.path.join(eindafrekening_dir, "Eindafrekening Inputs")
        os.makedirs(inputs_folder, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y-%m-%d")
        output_path = os.path.join(inputs_folder, f"input_master_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    
    # Create Lists sheet first (for dropdowns & lookups)
    create_lists_sheet(wb)
    
    # Create Input sheet
    ws = wb.active
    ws.title = "Input"
    
    # ==================== STYLES ====================
    header_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_gold = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_calc = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # ==================== COLUMNS CONFIGURATION ====================
    columns_config = [
        # GLOBAL
        ("Adres", None), 
        ("Type", None), 
        ("Klantnaam", None), 
        ("Object ID", header_blue), # D
        ("Klant Nr (Auto)", header_blue), # E
        ("RR Inspecteur", header_blue), # F
        ("Folder Link", header_blue),   # G
        
        # BASIS (Blue)
        ("Check-in", header_blue),        # H
        ("Check-out", header_blue),       # I
        ("Borg (€)", header_blue),        # J
        ("GWE Beheer", header_blue),      # K
        ("GWE Maandbedrag (€)", header_blue), # L
        ("BTW %", header_blue),           # M
        ("Voorschot GWE (Incl)", header_blue), # N
        ("Totaal Voorschot", header_blue), # O 
        
        # P was Verrekening - REMOVED

        ("Meterbeheerder", header_blue),  # P
        ("Leverancier", header_blue),     # Q
        ("Leverancier Nr (Auto)", header_blue), # R
        ("Contractnr", header_blue),      # S
        ("Extra Voorschot (€)", header_blue), # T
        ("Extra Voor. Omschr.", header_blue), # U
        
        # GWE Readings (Orange) (Cols V-AA)
        ("Elek Begin", header_orange),    # V
        ("Elek Eind", header_orange),     # W
        ("Gas Begin", header_orange),     # X
        ("Gas Eind", header_orange),      # Y
        ("Water Begin", header_orange),   # Z
        ("Water Eind", header_orange),    # AA
        
        # SCHOONMAAK (Gold) (Cols AB-AH)
        ("Schoon Maak Pakket", header_gold), # AB
        ("Schoon Uren", header_gold),     # AC
        ("Uurtarief (€)", header_gold),   # AD
        ("Totaal Excl (€)", header_gold), # AE
        ("BTW %", header_gold),           # AF
        ("BTW Bedrag (€)", header_gold),  # AG
        ("Totaal Incl (€)", header_gold), # AH
        
        # ITEMS (Red) (Cols AI-AQ)
        ("Kosten Type", header_red),      # AI
        ("Eenheid", header_red),          # AJ
        ("Beschrijving", header_red),     # AK
        ("Aantal", header_red),           # AL
        ("Prijs/Stuk (€)", header_red),   # AM
        ("Totaal Excl (€)", header_red),  # AN
        ("BTW %", header_red),            # AO
        ("BTW Bedrag (€)", header_red),   # AP
        ("Totaal Incl (€)", header_red)   # AQ
    ]

    # Write Headers
    count = 1
    for item in columns_config:
        name, fill = item
        if not fill:
            fill = PatternFill(start_color="5B5B5B", end_color="5B5B5B", fill_type="solid")
            
        cell = ws.cell(row=1, column=count)
        cell.value = name
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(count)].width = 15
        count += 1

    # Specific widths
    ws.column_dimensions['A'].width = 25 
    ws.column_dimensions['C'].width = 25 
    ws.column_dimensions['G'].width = 30 # Folder Link
    ws.column_dimensions['L'].width = 15 # GWE Excl
    ws.column_dimensions['N'].width = 15 # GWE Incl
    ws.column_dimensions['T'].width = 15 # Extra Voorschot (was U)
    ws.column_dimensions['U'].width = 30 # Extra Voor Omschr (was V)
    ws.column_dimensions['AK'].width = 30 # Beschrijving (AI starts items, AK is desc)

    # ==================== FORMULAS & VALIDATION ====================

    # Row Type
    dv_type = DataValidation(type="list", formula1='"Basis,GWE,GWE_Item,Schoonmaak,Schade,Extra"', allow_blank=False)
    ws.add_data_validation(dv_type)
    dv_type.add('B2:B5000')

    # Addresses
    dv_adres = DataValidation(type="list", formula1='=Lists!$A$2:$A$5000', allow_blank=True)
    ws.add_data_validation(dv_adres)
    dv_adres.add('A2:A5000')

    # Clients (Col C)
    dv_client = DataValidation(type="list", formula1='=Lists!$C$2:$C$5000', allow_blank=True)
    ws.add_data_validation(dv_client)
    dv_client.add('C2:C5000')

    # GWE Beheer (Col K)
    dv_gwe_beheer = DataValidation(type="list", formula1="=Lists!$K$2:$K$3", allow_blank=True)
    ws.add_data_validation(dv_gwe_beheer)
    dv_gwe_beheer.add('K2:K5000')

    # Meterbeheerder (Col P) (was Q)
    dv_metertype = DataValidation(type="list", formula1="=Lists!$I$2:$I$5", allow_blank=True) 
    ws.add_data_validation(dv_metertype)
    dv_metertype.add('P2:P5000') 

    # Leverancier (Col Q) (was R)
    dv_lev = DataValidation(type="list", formula1="=Lists!$E$2:$E$100", allow_blank=True)
    ws.add_data_validation(dv_lev)
    dv_lev.add('Q2:Q5000') 

    # Schoonmaak Pakket (Col AB) (was AC)
    dv_pakket = DataValidation(type="list", formula1="=Lists!$M$2:$M$6", allow_blank=True)
    ws.add_data_validation(dv_pakket)
    dv_pakket.add('AB2:AB5000')

    # Kosten Type (Col AI) (Item Type) (was AJ)
    dv_items = DataValidation(type="list", formula1="=Lists!$O$2:$O$5", allow_blank=True)
    ws.add_data_validation(dv_items)
    dv_items.add('AI2:AI5000')

    # Styles
    grey_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Prefill Formulas
    for r in range(2, 202):
        # D: ObjID -> VLOOKUP Address in Lists A:B -> 2
        ws[f'D{r}'] = f'=IF(A{r}="","",VLOOKUP(A{r},Lists!$A:$B,2,FALSE))'
        ws[f'D{r}'].fill = grey_fill 
        
        # E: Klant Nr -> VLOOKUP ClientName(C) in Lists C:D -> 2
        ws[f'E{r}'] = f'=IF(C{r}="","",VLOOKUP(C{r},Lists!$C:$D,2,FALSE))'
        ws[f'E{r}'].fill = grey_fill
        
        # F & G Meta
        ws[f'F{r}'].fill = grey_fill
        ws[f'G{r}'].fill = grey_fill

        # J: Borg (€) - Pre-fill from Lists Column H (Index 8)
        ws[f'J{r}'] = f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},Lists!$A:$H,8,FALSE),0))'
        ws[f'J{r}'].number_format = '€ #,##0.00'

        # L: GWE Maandbedrag - PREFILL from Lists Col G (Index 7)
        ws[f'L{r}'] = f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},Lists!$A:$H,7,FALSE),0))'
        ws[f'L{r}'].number_format = '€ #,##0.00'

        # M: BTW % (GWE)
        ws[f'M{r}'] = 0.21
        ws[f'M{r}'].number_format = '0%'
        ws[f'M{r}'].fill = grey_fill

        # N: Voorschot GWE (Incl) = L * (1 + M)
        ws[f'N{r}'] = f'=L{r} * (1 + M{r})'
        ws[f'N{r}'].number_format = '€ #,##0.00'
        ws[f'N{r}'].fill = grey_fill

        # O: Totaal Voorschot = IF(GWE Beheer="Eigen Beheer", 0, YEARFRAC(Checkin, Checkout, 4) * 12 * MonthlyPrice)
        # K = GWE Beheer, N = Voorschot GWE (Incl), I = Checkout, H = Checkin, 4 = European 30/360
        ws[f'O{r}'] = f'=IF(K{r}="Eigen Beheer",0,IF(AND(ISNUMBER(N{r}),I{r}<>"",H{r}<>""),YEARFRAC(H{r},I{r},4)*12*N{r},0))'
        ws[f'O{r}'].number_format = '€ #,##0.00'
        ws[f'O{r}'].fill = grey_fill

        # P: Was Verrekening - Removed
        
        # R: Lev Nr (Auto) (Was S) -> VLOOKUP LevName(Q) in Lists E:F -> 2
        # Now R targets Q
        ws[f'R{r}'] = f'=IF(Q{r}="","",VLOOKUP(Q{r},Lists!$E:$F,2,FALSE))'
        ws[f'R{r}'].fill = grey_fill 
        ws[f'R{r}'].alignment = Alignment(horizontal='center') 
        
        # Schoonmaak Calculations (Cols AB-AH)
        # AB:Pakket, AC:Uren, AD:Tarief (€50 default), AE:TotEx, AF:BTW%, AG:BTW€, AH:TotInc
        # 
        # NEW PRICING LOGIC:
        # - Basis Schoonmaak: Fixed €250 incl BTW (5h included)
        # - Intensief Schoonmaak: Fixed €375 incl BTW (7h included)
        # - Extra hours: €50/hour excl BTW
        
        # AD: Default tarief to 50 if blank
        ws[f'AD{r}'] = 50
        ws[f'AD{r}'].number_format = '€ #,##0.00'
        
        # AF (BTW%) - default 21%
        ws[f'AF{r}'] = 0.21
        ws[f'AF{r}'].number_format = '0%'
        
        # AE (TotEx) - Fixed package + extra hours
        # Formula: IF Basis -> (250/1.21) + MAX(0, hours-5)*50
        #          IF Intensief -> (375/1.21) + MAX(0, hours-7)*50
        #          ELSE -> hours * tarief
        f_basis_excl = f'(250/1.21)+MAX(0,AC{r}-5)*AD{r}'
        f_intensief_excl = f'(375/1.21)+MAX(0,AC{r}-7)*AD{r}'
        f_other_excl = f'AC{r}*AD{r}'
        
        ws[f'AE{r}'] = f'=IF(AB{r}="","",IF(ISNUMBER(SEARCH("Basis",AB{r})),{f_basis_excl},IF(ISNUMBER(SEARCH("Intensief",AB{r})),{f_intensief_excl},{f_other_excl})))'
        ws[f'AE{r}'].fill = fill_calc
        ws[f'AE{r}'].number_format = '€ #,##0.00'

        # AG (BTW€) = AE * AF
        ws[f'AG{r}'] = f'=IF(AE{r}<>"",AE{r}*AF{r},"")'
        ws[f'AG{r}'].fill = fill_calc
        ws[f'AG{r}'].number_format = '€ #,##0.00'
        
        # AH (TotInc) = AE + AG
        ws[f'AH{r}'] = f'=IF(AE{r}<>"",AE{r}+AG{r},"")'
        ws[f'AH{r}'].fill = fill_calc
        ws[f'AH{r}'].number_format = '€ #,##0.00'
        
        # ITEM CALCULATIONS (Cols AI-AQ) (Was AJ-AR)
        # AI:Type, AJ:Unit, AK:Desc, AL:Aant, AM:Prijs, AN:TotEx, AO:BTW%, AP:BTW€, AQ:TotInc
        
        # AL (Aantal) - SMART LOOKUP
        # GWE Indices (Previously W-AB): ElekBegin=V(was W), ElekEind=W, GasBegin=X, GasEind=Y, WaterBegin=Z, WaterEind=AA
        # Need to shift references back 1 char.
        # W -> V, X -> W, Y -> X, Z -> Y, AA -> Z, AB -> AA
        
        f_elek = f'SUMIFS($W:$W,$A:$A,$A{r},$B:$B,"GWE")-SUMIFS($V:$V,$A:$A,$A{r},$B:$B,"GWE")'
        f_gas = f'SUMIFS($Y:$Y,$A:$A,$A{r},$B:$B,"GWE")-SUMIFS($X:$X,$A:$A,$A{r},$B:$B,"GWE")'
        f_water = f'SUMIFS($AA:$AA,$A:$A,$A{r},$B:$B,"GWE")-SUMIFS($Z:$Z,$A:$A,$A{r},$B:$B,"GWE")'
        
        f_smart = f'IF($AI{r}="Elektra",{f_elek},IF($AI{r}="Gas",{f_gas},IF($AI{r}="Water",{f_water},"")))'
        
        ws[f'AL{r}'] = f'=IF($B{r}="GWE_Item",{f_smart},"")'
        
        # AN (TotEx) = AL * AM
        ws[f'AN{r}'] = f'=IF(AND(AL{r}<>"",AM{r}<>""),AL{r}*AM{r},"")'
        ws[f'AN{r}'].fill = fill_calc
        ws[f'AN{r}'].number_format = '€ #,##0.00'
        
        # AO (BTW%)
        ws[f'AO{r}'].number_format = '0%'

        # AP (BTW€) = AN * AO
        ws[f'AP{r}'] = f'=IF(AND(AN{r}<>"",AO{r}<>""),AN{r}*AO{r},"")'
        ws[f'AP{r}'].fill = fill_calc
        ws[f'AP{r}'].number_format = '€ #,##0.00'
        
        # AQ (TotInc) = AN + AP
        ws[f'AQ{r}'] = f'=IF(AND(AN{r}<>"",AP{r}<>""),AN{r}+AP{r},"")'
        ws[f'AQ{r}'].fill = fill_calc
        ws[f'AQ{r}'].number_format = '€ #,##0.00'

    # Conditional Formatting
    # Basis: D2:U5000 (was V)
    ws.conditional_formatting.add('D2:U5000', FormulaRule(formula=['$B2<>"Basis"'], stopIfTrue=True, fill=grey_fill))
    # Readings V-AA (was W-AB)
    ws.conditional_formatting.add('V2:AA5000', FormulaRule(formula=['$B2<>"GWE"'], stopIfTrue=True, fill=grey_fill))
    # Cleaning AB-AH (was AC-AI)
    ws.conditional_formatting.add('AB2:AH5000', FormulaRule(formula=['$B2<>"Schoonmaak"'], stopIfTrue=True, fill=grey_fill))
    # Items AI-AQ (was AJ-AR)
    ws.conditional_formatting.add('AI2:AQ5000', FormulaRule(formula=['AND($B2<>"Schade",$B2<>"Extra",$B2<>"GWE_Item")'], stopIfTrue=True, fill=grey_fill))

    ws.freeze_panes = 'A2'
    
    # Dashboard
    create_dashboard_sheet(wb, ws) 

    wb.save(output_path)
    print(f"✅ Generated Universal Master Template: {output_path}")

def create_dashboard_sheet(wb, input_ws):
    ws = wb.create_sheet("Dashboard", 0) 
    ws.sheet_properties.tabColor = "00B050" 
    
    ws['A1'] = "COMPLETENESS DASHBOARD"
    ws['A1'].font = Font(size=14, bold=True)
    
    headers = ["Adres", "Basis?", "GWE?", "Schoonmaak?", "STATUS"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="364E6F", fill_type="solid")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
    ws.column_dimensions['A'].width = 30
    
    ws['A4'] = '=_xlfn.UNIQUE(_xlfn.FILTER(Input!A2:A5000, Input!A2:A5000<>""))'
    
    for i in range(4, 54): 
        ws[f'B{i}'] = f'=IF(A{i}="","",IF(COUNTIFS(Input!$A:$A, A{i}, Input!$B:$B, "Basis")>0, "OK", "MISSING"))'
        ws[f'C{i}'] = f'=IF(A{i}="","",IF(COUNTIFS(Input!$A:$A, A{i}, Input!$B:$B, "GWE")>0, "OK", "MISSING"))'
        ws[f'D{i}'] = f'=IF(A{i}="","",IF(COUNTIFS(Input!$A:$A, A{i}, Input!$B:$B, "Schoonmaak")>0, "OK", "MISSING"))'
        ws[f'E{i}'] = f'=IF(A{i}="","",IF(AND(B{i}="OK",C{i}="OK",D{i}="OK"), "✅ READY", "⚠️ INCOMPLETE"))'
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    ws.conditional_formatting.add('E4:E54', FormulaRule(formula=['E4="✅ READY"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('E4:E54', FormulaRule(formula=['E4="⚠️ INCOMPLETE"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('B4:D54', FormulaRule(formula=['B4="MISSING"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('B4:D54', FormulaRule(formula=['B4="OK"'], stopIfTrue=True, fill=green_fill))

def create_lists_sheet(wb):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = 'hidden'
    ws['A1'] = "Address"
    ws['B1'] = "Object ID"
    ws['C1'] = "Clients (Name)"
    ws['D1'] = "Clients (ID)"
    ws['E1'] = "Leveranciers (Name)"
    ws['F1'] = "Leveranciers (ID)"
    ws['G1'] = "GWE Voorschot (Default)"
    ws['H1'] = "Borg (Default)"
    
    try:
        db = Database()
        conn = db.get_connection()
        
        # Houses with GWE info and Borg
        cursor = conn.execute("""
            SELECT adres, object_id, voorschot_gwe, borg
            FROM huizen 
            WHERE status='active' AND object_id IS NOT NULL AND object_id != ''
            ORDER BY adres
        """)
        houses = cursor.fetchall()

        # Clients
        cursor = conn.execute("SELECT naam, id FROM relaties WHERE is_klant=1 ORDER BY naam")
        clients = cursor.fetchall()
        
        # Suppliers (Leveranciers)
        cursor = conn.execute("SELECT naam, id FROM leveranciers ORDER BY naam")
        suppliers = cursor.fetchall()
        
        conn.close()
        
        for i, (addr, obj_id, gwe_val, borg_val) in enumerate(houses, 2):
            ws[f'A{i}'] = addr
            ws[f'B{i}'] = obj_id if obj_id else ""
            ws[f'G{i}'] = gwe_val if gwe_val else 0
            ws[f'H{i}'] = borg_val if borg_val else 0
            
        for i, (name, c_id) in enumerate(clients, 2):
            ws[f'C{i}'] = name
            ws[f'D{i}'] = c_id
            
        for i, (name, l_id) in enumerate(suppliers, 2):
            ws[f'E{i}'] = name
            ws[f'F{i}'] = l_id
            
        # STATIC LISTS (I, K, M, O)
        # I: Meterbeheerder
        meter_ops = ["Stedin", "Liander", "Enexis", "Westland Infra"]
        ws['I1'] = "Meterbeheerder"
        ws['I1'].font = Font(bold=True)
        for idx, val in enumerate(meter_ops, 2):
            ws[f'I{idx}'] = val
            
        # K: GWE Beheer
        gwe_opts = ["Via RyanRent", "Eigen Beheer"]
        ws['K1'] = "GWE Beheer"
        ws['K1'].font = Font(bold=True)
        for idx, val in enumerate(gwe_opts, 2):
            ws[f'K{idx}'] = val
            
        # M: Schoonmaak Pakket
        clean_opts = ["Basis Schoonmaak", "Intensief Schoonmaak", "Geen Schoonmaak", "Achteraf Betaald", "Op Maat"]
        ws['M1'] = "Schoonmaak Pakket"
        ws['M1'].font = Font(bold=True)
        for idx, val in enumerate(clean_opts, 2):
            ws[f'M{idx}'] = val
            
        # O: Item Types (Kosten Type)
        item_types = ["Elektra", "Gas", "Water", "Overig"]
        ws['O1'] = "Kosten Type"
        ws['O1'].font = Font(bold=True)
        for idx, val in enumerate(item_types, 2):
            ws[f'O{idx}'] = val
            
        print(f"📊 Added {len(houses)} houses, {len(clients)} clients, {len(suppliers)} suppliers to lists")
            
    except Exception as e:
        print(f"⚠️ Could not fetch data: {e}")
        # Valid Fallback for development
        ws['A2'] = "Test Address 1"
        ws['B2'] = "OBJ-001"
        ws['G2'] = 50.00
        ws['H2'] = 500.00
        ws['C2'] = "Test Client"

if __name__ == "__main__":
    create_master_template()
