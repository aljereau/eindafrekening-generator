import openpyxl
import shutil
import os
import subprocess

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR = os.path.join(BASE_DIR, 'src')
INPUT_TEMPLATE = os.path.join(SRC_DIR, 'input_template.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

def run_scenario(name, changes):
    print(f"\n🧪 Running Scenario: {name}")
    
    # 1. Create temporary input file
    temp_input = os.path.join(SRC_DIR, f'input_{name.lower().replace(" ", "_")}.xlsx')
    shutil.copy2(INPUT_TEMPLATE, temp_input)
    
    # 2. Modify Excel
    wb = openpyxl.load_workbook(temp_input)
    
    for sheet_name, updates in changes.items():
        if sheet_name not in wb.sheetnames:
            print(f"   ⚠️  Sheet '{sheet_name}' not found!")
            continue
            
        ws = wb[sheet_name]
        for row, col, value in updates:
            ws.cell(row=row, column=col, value=value)
            print(f"   - [{sheet_name}] Set ({row}, {col}) to '{value}'")

    wb.save(temp_input)
    wb.close()
    
    # 3. Run Generator
    cmd = [
        'python3', 'generate.py',
        '--input', os.path.basename(temp_input),
        '--no-open',
        '--no-pause'
    ]
    
    try:
        subprocess.run(cmd, cwd=SRC_DIR, check=True, capture_output=True, text=True)
        print(f"   ✅ Generator finished successfully.")
        
        # 4. Rename Output
        # Find the latest generated file (it will have a timestamp)
        # We know the client name is 'DS People B.V.' from the template
        # But let's just look for the most recent .html file in output dir
        files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('_onepager.html')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(OUTPUT_DIR, x)), reverse=True)
        
        if files:
            latest = files[0]
            new_name = f"scenario_{name.lower().replace(' ', '_')}.html"
            src = os.path.join(OUTPUT_DIR, latest)
            dst = os.path.join(OUTPUT_DIR, new_name)
            shutil.move(src, dst)
            print(f"   📂 Saved report to: {new_name}")
            return dst
        else:
            print("   ❌ No output file found.")
            return None
            
    except subprocess.CalledProcessError as e:
        print(f"   ❌ Generator failed: {e.stderr}")
        return None
    finally:
        # Cleanup temp input
        if os.path.exists(temp_input):
            os.remove(temp_input)

def main():
    scenarios = [
        {
            "name": "Basis_Neutral",
            "changes": {
                # Sheet: (Row, Col, Value)
                "Algemeen": [
                    (29, 2, 181.50),  # Voorschot Schoonmaak (Matches cost)
                    (35, 2, "Basis Schoonmaak"), # Pakket
                    (39, 2, 181.50),  # Totaal Kosten (Manually calculated)
                ],
                "Schoonmaak": [
                    (4, 2, "Basis Schoonmaak"), # Pakket
                    (6, 2, 3),    # Totaal Uren Gewerkt
                    (8, 2, 50),   # Uurtarief
                ]
            }
        },
        {
            "name": "Intensief_Charge",
            "changes": {
                "Algemeen": [
                    (29, 2, 150.00),  # Voorschot (Low)
                    (35, 2, "Intensief Schoonmaak"),
                    (39, 2, 452.50),  # Totaal Kosten: (150 + (5 extra * 50)) * 1.21? 
                                      # Let's assume: Base 150 (3 hrs). Total 10 hrs. Extra 7.
                                      # Cost = 150 + (7 * 50) = 500 excl VAT?
                                      # Let's simplify: Total Cost = 452.50 incl VAT.
                                      # Advance = 150. Result = 302.50 Te betalen.
                ],
                "Schoonmaak": [
                    (4, 2, "Intensief Schoonmaak"),
                    (6, 2, 10),   # Totaal Uren
                    (8, 2, 50),
                ]
            }
        },
        {
            "name": "Refund_Green",
            "changes": {
                "Algemeen": [
                    (29, 2, 500.00),  # Voorschot (High)
                    (35, 2, "Basis Schoonmaak"),
                    (39, 2, 181.50),  # Totaal Kosten (Low)
                ],
                "Schoonmaak": [
                    (4, 2, "Basis Schoonmaak"),
                    (6, 2, 3),    # Totaal Uren
                    (8, 2, 50),
                ]
            }
        }
    ]
    
    results = []
    for s in scenarios:
        # Convert changes format for run_scenario
        # We need to modify run_scenario to handle this new format
        path = run_scenario(s['name'], s['changes'])
        if path:
            results.append(path)
            
    print("\n🏁 All scenarios completed.")
    for r in results:
        print(f" - {r}")

if __name__ == "__main__":
    main()
