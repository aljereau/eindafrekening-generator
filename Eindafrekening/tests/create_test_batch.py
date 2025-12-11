import openpyxl
from datetime import date

TEMPLATE = "Eindafrekening/src/input_master_0179_V2.xlsx"
OUTPUT = "Eindafrekening/src/input_master_batch_test.xlsx"

def create_batch_data():
    print(f"Opening template: {TEMPLATE}")
    wb = openpyxl.load_workbook(TEMPLATE)
    if 'Input' in wb.sheetnames:
        ws = wb['Input']
    else:
        # Fallback if V2 didn't have it, but we know it does
        ws = wb.active

    # Clear rows 2-100
    for row in ws.iter_rows(min_row=2, max_row=100):
        for cell in row:
            cell.value = None

    # Test Data
    # 3 Bookings
    
    # Booking 1: Test House 1 (Standard)
    # Row 2: Basis
    ws['A2'] = "Test Straat 1"
    ws['B2'] = "Basis"
    ws['C2'] = "Client Een"
    ws['H2'] = "2024-01-01" # Checkin
    ws['I2'] = "2024-01-31" # Checkout
    ws['J2'] = 1000.0 # Borg
    ws['K2'] = "Via RyanRent"
    ws['M2'] = 250.00 # GWE Auto Voorschot
    
    # Row 3: GWE (Electricity)
    ws['A3'] = "Test Straat 1"
    ws['B3'] = "GWE"
    ws['T3'] = 1000 # Elek Begin
    ws['U3'] = 1200 # Elek Eind
    
    # Row 4: Schoonmaak
    ws['A4'] = "Test Straat 1"
    ws['B4'] = "Schoonmaak"
    ws['Z4'] = "Basis Schoonmaak"
    ws['AA4'] = 3.0 # Uren (Underused test: 3h vs 5h package)
    ws['AB4'] = 50.0 # Tarief
    
    # Booking 2: Test House 2 (Damage Only)
    # Row 5: Basis
    ws['A5'] = "Test Straat 2"
    ws['B5'] = "Basis"
    ws['C5'] = "Client Twee"
    ws['H5'] = "2024-02-01"
    ws['I5'] = "2024-02-28"
    ws['J5'] = 800.0
    ws['K5'] = "Eigen Beheer" # No GWE
    
    # Row 6: Schade
    ws['A6'] = "Test Straat 2"
    ws['B6'] = "Schade"
    ws['AG6'] = "Overig"
    ws['AI6'] = "Gebroken Lamp"
    ws['AJ6'] = 1
    ws['AK6'] = 50.00
    ws['AM6'] = 0.21
    
    # Booking 3: Test House 3 (Full Option)
    # Row 7: Basis
    ws['A7'] = "Test Straat 3"
    ws['B7'] = "Basis"
    ws['C7'] = "Client Drie"
    ws['H7'] = "2024-03-01"
    ws['I7'] = "2024-03-31"
    ws['J7'] = 1500.0
    ws['K7'] = "Via RyanRent"
    ws['M7'] = 300.0 # GWE Voorschot
    
    # Row 8: cleaning extra
    ws['A8'] = "Test Straat 3"
    ws['B8'] = "Schoonmaak"
    ws['Z8'] = "Intensief Schoonmaak"
    ws['AA8'] = 10.0 # 3 hours extra (7 included)
    ws['AB8'] = 60.0
    
    ws['T7'] = 0; ws['U7'] = 0 # Prevent format error on read perhaps?
    
    wb.save(OUTPUT)
    print(f"Created batch test file: {OUTPUT}")

if __name__ == "__main__":
    create_batch_data()
