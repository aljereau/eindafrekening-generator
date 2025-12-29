#!/usr/bin/env python3
"""
Master Reader
Reads row-based Master Excel sheet and converts it to structured booking data.
"""

import openpyxl
from collections import defaultdict
from typing import List, Dict, Any, Optional
from datetime import date, datetime
import sys
import os

# Add Shared and src to path
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(root_dir, 'Shared'))

# Import entities from existing excel_reader (reuse definitions)
from entities import (
    Client, RentalProperty, Period, Deposit, ExtraVoorschot, GWEMeterReading, GWERegel, 
    GWETotalen, Cleaning, DamageRegel, DamageTotalen, GWEMeterstanden
)
# We need Database to fetch missing details (like Postcode/City)
from database import Database

class MasterReader:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.db = Database()
        
    def read_all(self) -> List[Dict[str, Any]]:
        """Read master sheet and return list of booking data dicts"""
        print(f"📖 Reading Master Input: {self.filepath}")
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        if 'Input' in wb.sheetnames:
            ws = wb['Input']
        else:
            print("⚠️ 'Input' sheet not found, using active sheet")
            ws = wb.active
        
        # Group rows by Address (or Booking Identifier if available)
        # Using Address + Checkin Date as unique key would be safer, 
        # but for now let's assume rows are grouped by Address or just process sequentially.
        # Actually, grouping by address allows scattered rows.
        
        grouped_data = defaultdict(list)
        
        # Iterating rows, skipping header
        for row in ws.iter_rows(min_row=2, values_only=True):
            adres = row[0]
            if not adres: continue
            grouped_data[adres].append(row)
            
        results = []
        
        for adres, rows in grouped_data.items():
            try:
                booking_data = self._process_booking_rows(adres, rows)
                if booking_data:
                    results.append(booking_data)
            except Exception as e:
                print(f"❌ Error processing booking for {adres}: {e}")
                # Continue with next booking
                
        return results

    def _process_booking_rows(self, adres: str, rows: List[tuple]) -> Optional[Dict[str, Any]]:
        """Process all rows for a single booking/address"""
        
        # Initialize containers
        client = None
        period = None
        deposit = None
        gwe_standen = None
        cleaning = None
        damage_regels = []
        gwe_regels = [] # Added initialization
        extra_voorschot = None
        gwe_settings = {'beheer_type': 'Via RyanRent'} # Default
        
        # Helper to parse amount strings
        def parse_float(val):
            if not val: return 0.0
            if isinstance(val, (int, float)): return float(val)
            try:
                return float(str(val).replace('€', '').replace(',', '.').strip())
            except:
                return 0.0

        # Helper to parse dates
        def parse_date(val):
            if not val: return date.today()
            if isinstance(val, datetime): return val.date()
            if isinstance(val, date): return val
            return date.today() # Fallback

        has_basis = False

        for row in rows:
            # Ensure row has enough columns (pad if needed)
            if len(row) < 26:
                row = row + (None,) * (26 - len(row))
                
            row_type = str(row[1]).strip() if row[1] else ""
            
            if row_type == "Basis":
                has_basis = True
                
            if row_type == "Basis":
                has_basis = True
                
                # Column mapping based on updated template
                # 0:Adres, 1:Type, 2:Klant, 3:ObjID, 4:KlantNr, 5:Insp, 6:Link
                # 7:Checkin, 8:Checkout, 9:Borg
                # 10:GWEBeh, 11:GWEMaand, 12:GWEAuto, 13:Verreken, 14:LevName, 15:LevID, 16:Contract, 17:ExVoor, 18:ExDesc
                
                klant_naam = row[2]
                checkin = parse_date(row[7]) # H
                checkout = parse_date(row[8]) # I
                borg_voorschot = parse_float(row[9]) # J
                
                gwe_beheer_val = str(row[10]).strip() if row[10] else "Via RyanRent" # K
                gwe_settings['beheer_type'] = gwe_beheer_val
                
                # GWE Voorschot (Auto) - Column O (Index 14)
                # This is the monthly amount * months usually, calculated in Excel.
                # INPUT IS NOW VAT-INCLUSIVE via formula in template.
                gwe_voorschot_auto = parse_float(row[14])
                
                # Extra Voorschot - Col T(19), U(20) (Was U, V)
                ex_voor_bedrag = parse_float(row[19])
                ex_voor_desc = str(row[20]) if row[20] else "Extra Voorschot"
                
                if ex_voor_bedrag > 0:
                    extra_voorschot = ExtraVoorschot(
                        voorschot=ex_voor_bedrag,
                        gebruikt=0.0,
                        terug=ex_voor_bedrag,
                        restschade=0.0,
                        omschrijving=ex_voor_desc
                    )
                
                # Client
                client = Client(name=klant_naam, contact_person="", email="", phone=None)
                
                # Period
                days = (checkout - checkin).days
                period = Period(checkin_date=checkin, checkout_date=checkout, days=days)
                
                # Deposit
                deposit = Deposit(voorschot=borg_voorschot, gebruikt=0.0, terug=0.0, restschade=0.0)

            elif row_type == "GWE":
                # Readings (Cols V-AA -> Indices 21-26)
                # V21, W22, X23, Y24, Z25, AA26
                elek_start = parse_float(row[21])
                elek_eind = parse_float(row[22])
                gas_start = parse_float(row[23])
                gas_eind = parse_float(row[24])
                water_start = parse_float(row[25])
                water_eind = parse_float(row[26])
                
                stroom = GWEMeterReading(begin=elek_start, eind=elek_eind, verbruik=elek_eind-elek_start)
                gas = GWEMeterReading(begin=gas_start, eind=gas_eind, verbruik=gas_eind-gas_start)
                water = GWEMeterReading(begin=water_start, eind=water_eind, verbruik=water_eind-water_start)
                
                gwe_standen = GWEMeterstanden(stroom=stroom, gas=gas, water=water)
                
            elif row_type == "Schoonmaak":
                # Cleaning (Cols AB-AH -> Indices 27-33)
                # AB(27):Pakket, AC:Uren, AD:Tarief, AE:TotEx, AF:BTW%, AG:BTW€, AH:TotInc
                pakket = str(row[27]).strip()
                uren = parse_float(row[28])
                tarief = parse_float(row[29])
                btw_pct = parse_float(row[31]) # AF is Index 31
                
                if not btw_pct: btw_pct = 0.21
                if btw_pct > 1: btw_pct = btw_pct / 100 
                
                # Fixed package prices (incl BTW)
                PAKKET_PRIJZEN = {
                    'basis': 250.0,      # Basis Schoonmaak
                    'intensief': 375.0   # Intensief Schoonmaak
                }
                
                # Map packet name
                pakket_code = 'basis'
                pakket_prijs_incl = 0.0
                
                if 'basis' in pakket.lower(): 
                    pakket_code = 'basis'
                    pakket_prijs_incl = PAKKET_PRIJZEN['basis']
                elif 'intensief' in pakket.lower(): 
                    pakket_code = 'intensief'
                    pakket_prijs_incl = PAKKET_PRIJZEN['intensief']
                elif 'geen' in pakket.lower(): 
                    pakket_code = 'geen'
                    pakket_prijs_incl = 0.0
                elif 'achteraf' in pakket.lower(): 
                    pakket_code = 'achteraf'
                    pakket_prijs_incl = 0.0
                
                # Get total cost from Excel - check both Excl (AE) and Incl (AH)
                total_cost_excl = parse_float(row[30])  # AE: Totaal Excl
                total_cost_incl = parse_float(row[33])  # AH: Totaal Incl
                
                # Cleaning voorschot = fixed package price (what tenant already paid)
                cleaning_voorschot = pakket_prijs_incl
                
                # Calculate final cost - prioritize excl input, then incl input
                if total_cost_excl > 0:
                    # User entered excl value - calculate incl
                    final_total_cost = total_cost_excl * (1 + btw_pct)
                elif total_cost_incl > 0:
                    # User entered incl value - use directly
                    final_total_cost = total_cost_incl
                elif uren > 0 and tarief > 0:
                    # Calculate from hours * rate
                    final_total_cost = uren * tarief * (1 + btw_pct)
                else:
                    # Use package price as minimum
                    final_total_cost = pakket_prijs_incl
                
                # Enforce minimum package price (no refunds for unused hours)
                if pakket_code in ['basis', 'intensief']:
                    final_total_cost = max(pakket_prijs_incl, final_total_cost)
                
                # Extra bedrag = what tenant has to pay on top of package
                extra_bedrag = max(0, final_total_cost - pakket_prijs_incl)
                
                cleaning = Cleaning(
                    pakket_type=pakket_code,
                    pakket_naam=pakket,
                    inbegrepen_uren=0,  # Not relevant anymore - fixed price packages
                    totaal_uren=uren,
                    extra_uren=0,  # Not relevant - we use total costs
                    uurtarief=tarief,
                    extra_bedrag=extra_bedrag / (1 + btw_pct),  # Store excl
                    voorschot=cleaning_voorschot, 
                    totaal_kosten_incl=final_total_cost,
                    btw_percentage=btw_pct,
                    btw_bedrag=final_total_cost - (final_total_cost/(1+btw_pct))
                )

            elif row_type == "GWE_Item":
                # Cols AI-AQ (Indices 34-42)
                # AI(34):Type, AJ:Unit, AK:Desc, AL:Aant, AM:Prijs, AN:TotEx, AO:BTW%, AP:BTW€, AQ:TotInc
                k_type = str(row[34]).strip() if row[34] else "Overig"
                eenheid = str(row[35]).strip() if row[35] else ""
                desc = row[36]
                aantal = parse_float(row[37])
                prijs = parse_float(row[38])
                btw = parse_float(row[40]) # AO is 40
                
                if btw > 1: btw = btw / 100
                if not btw: btw = 0.21
                
                bedrag_excl = aantal * prijs
                
                # Append to gwe_regels list
                gwe_regels.append(GWERegel(
                    type=k_type,
                    omschrijving=desc,
                    eenheid=eenheid,
                    verbruik_of_dagen=aantal,
                    tarief_excl=prijs,
                    kosten_excl=bedrag_excl,
                    btw_percentage=btw
                ))

            elif row_type in ["Schade", "Extra"]:
                # Item Cols (AI-AQ -> Indices 34-42)
                # AI(34):Type, AJ(35):Unit, AK(36):Desc, AL(37):Aant, AM(38):Prijs, AN(39):TotEx, AO(40):BTW%, AP(41):BTW€, AQ(42):TotInc
                
                desc = row[36]  # AK: Beschrijving
                qty = parse_float(row[37])   # AL: Aantal
                amt = parse_float(row[38])   # AM: Prijs/Stuk
                total_excl_cell = parse_float(row[39])  # AN: Totaal Excl (direct input)
                btw = parse_float(row[40])   # AO: BTW %
                total_incl_cell = parse_float(row[42])  # AQ: Totaal Incl (direct input)
                
                if btw > 1: btw = btw / 100
                if not btw: btw = 0.21 
                
                # Calculate total_excl - prioritize direct input over calculation
                if total_excl_cell > 0:
                    # User entered total excl directly
                    total_excl = total_excl_cell
                    tarief = total_excl / qty if qty > 0 else total_excl
                elif total_incl_cell > 0:
                    # User entered total incl - back-calculate excl
                    total_excl = total_incl_cell / (1 + btw)
                    tarief = total_excl / qty if qty > 0 else total_excl
                elif amt > 0:
                    # User entered unit price - calculate total
                    tarief = amt
                    total_excl = qty * tarief
                else:
                    # No price data at all
                    tarief = 0
                    total_excl = 0
                
                # Only add if we have a description
                if desc:
                    damage_regels.append(DamageRegel(
                        beschrijving=desc,
                        aantal=qty if qty > 0 else 1,
                        tarief_excl=tarief,
                        bedrag_excl=total_excl,
                        btw_percentage=btw
                    ))

        if not has_basis:
            print(f"⚠️  Skipping {adres}: No 'Basis' row found.")
            return None

        # Fetch Property Details from DB
        prop_details = self._fetch_property_details(adres)
        rental_property = RentalProperty(
            address=adres,
            unit=None,
            postal_code=prop_details.get('postcode'),
            city=prop_details.get('plaats'),
            object_id=prop_details.get('object_id')
        )
        
        # Defaults if missing
        if not gwe_standen:
            z = GWEMeterReading(0,0,0)
            gwe_standen = GWEMeterstanden(z,z,z)
            
        if not cleaning:
            cleaning = Cleaning('geen', 'Geen pakket', 0,0,0,0,0,0,0,0,0)
            
        # GWE Voorschot Validation
        # If gwe_voorschot_auto (calculated in Excel) is available, use it.
        # Otherwise default to 0. 
        # Note: variable gwe_voorschot_auto was typically set in 'Basis' block.
        # If 'Basis' block executed (has_basis checked), then gwe_voorschot_auto must have been set or valid scope?
        # Yes, defined in Basis block. But scope in Python loop? 
        # Actually gwe_voorschot_auto is set inside the loop. If multiple Basis rows (unlikely), last one wins.
        # But wait, python scope leaks variable after loop. But if loop didn't run basis, we returned None.
        # So gwe_voorschot_auto is safe to access here if initialized.
        # Let's initialize it before loop to be safe.
        
        # Initialize loop vars
        gwe_voorschot_auto_final = 0.0

        # Wait, I can't easily access variables from inside the previous loop block if they weren't init outside.
        # I should've initialized gwe_voorschot_auto at top of _process_booking_rows.
        # Or I can assume if has_basis is true, the variable is bound.
        # But cleaner to check. I'll pass it via a local dict or just trust logic.
        # Actually, `gwe_voorschot_auto` is defined inside `if row_type == "Basis":`.
        # Python variables DO leak to function scope. So it exists.
        
        # Initialize at top for clarity in future, but for this edit patch I will just trust it or default it calculation.
        
        # Totals mapping
        gwe_totalen = GWETotalen(0,0,0, gwe_settings['beheer_type'])
        damage_totalen = DamageTotalen(0,0,0)
        
        return {
            'client': client,
            'object': rental_property,
            'period': period,
            'deposit': deposit,
            'gwe_voorschot': locals().get('gwe_voorschot_auto', 0.0), # Use locals get to be safe
            'gwe_meterstanden': gwe_standen,
            'gwe_regels': gwe_regels, 
            'gwe_totalen': gwe_totalen,
            'cleaning': cleaning,
            'damage_regels': damage_regels,
            'damage_totalen': damage_totalen,
            'extra_voorschot': extra_voorschot
        }

    def _fetch_property_details(self, adres: str) -> Dict[str, Any]:
        """Fetch details from DB"""
        try:
             # We need to get a fresh connection - context manager usage is better but ad-hoc here
             conn = self.db.get_connection()
             cursor = conn.execute("SELECT object_id, postcode, plaats FROM huizen WHERE adres=?", (adres,))
             row = cursor.fetchone()
             conn.close()
             if row:
                 return {'object_id': row[0], 'postcode': row[1], 'plaats': row[2]}
        except Exception as e:
            print(f"⚠️ DB Error fetching property {adres}: {e}")
            
        return {}

if __name__ == "__main__":
    import sys
    file = sys.argv[1] if len(sys.argv) > 1 else "../input_master.xlsx"
    reader = MasterReader(file)
    data = reader.read_all()
    for d in data:
        print(f"Found: {d['object'].address} - {d['client'].name}")
