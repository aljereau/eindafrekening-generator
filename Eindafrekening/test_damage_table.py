import openpyxl
import shutil
import os
import subprocess

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR = os.path.join(BASE_DIR, 'src')
INPUT_TEMPLATE = os.path.join(SRC_DIR, 'input_template.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

def run_test():
    print(f"\n🧪 Running Test: Damage Table Verification")
    
    # 1. Create temporary input file
    temp_input = os.path.join(SRC_DIR, 'input_damage_test.xlsx')
    shutil.copy2(INPUT_TEMPLATE, temp_input)
    
    # 2. Modify Excel
    wb = openpyxl.load_workbook(temp_input)
    
    # Add Damage Items
    if "Schade" in wb.sheetnames:
        ws = wb["Schade"]
        # Clear existing rows 5-10 just in case
        for row in range(5, 11):
            for col in range(1, 7):
                ws.cell(row=row, column=col, value=None)
        
        # Add test data
        # Col 1: Desc, 2: Qty, 3: Rate, 4: Amount, 5: VAT%
        damage_items = [
            (5, "Gebroken lamp", 1, 15.00, 15.00, 0.21),
            (6, "Extra schoonmaak (laag tarief)", 2, 25.00, 50.00, 0.09),
            (7, "Verloren sleutel", 1, 30.00, 30.00, 0.21),
        ]
        
        for row, desc, qty, rate, amount, vat in damage_items:
            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=qty)
            ws.cell(row=row, column=3, value=rate)
            ws.cell(row=row, column=4, value=amount)
            ws.cell(row=row, column=5, value=vat)
            print(f"   - Added damage: {desc} (€{amount} excl VAT)")

    wb.save(temp_input)
    wb.close()
    
    # 3. Run Generator
    cmd = [
        'python3', 'generate.py',
        '--input', os.path.basename(temp_input),
        '--no-open',
        '--no-pause'
    ]
    
    try:
        result = subprocess.run(cmd, cwd=SRC_DIR, check=True, capture_output=True, text=True)
        print(result.stdout)
        print(f"   ✅ Generator finished successfully.")
        
        # 4. Rename Output
        files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('_detail.html')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(OUTPUT_DIR, x)), reverse=True)
        
        if files:
            latest = files[0]
            new_name = "test_damage_table.html"
            src = os.path.join(OUTPUT_DIR, latest)
            dst = os.path.join(OUTPUT_DIR, new_name)
            shutil.move(src, dst)
            print(f"   📂 Saved report to: {new_name}")
            return dst
            
    except subprocess.CalledProcessError as e:
        print(f"   ❌ Generator failed: {e.stderr}")
        return None
    finally:
        if os.path.exists(temp_input):
            os.remove(temp_input)

if __name__ == "__main__":
    run_test()
