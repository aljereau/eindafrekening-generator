import openpyxl
import shutil
import os
import subprocess

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR = os.path.join(BASE_DIR, 'src')
INPUT_TEMPLATE = os.path.join(SRC_DIR, 'input_template.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

def run_test():
    print(f"\n🧪 Running Test: Missing Voorschot for Basis Schoonmaak")
    
    # 1. Create temporary input file
    temp_input = os.path.join(SRC_DIR, 'input_missing_voorschot.xlsx')
    shutil.copy2(INPUT_TEMPLATE, temp_input)
    
    # 2. Modify Excel
    wb = openpyxl.load_workbook(temp_input)
    
    # Set Basis Schoonmaak but 0 Voorschot
    changes = {
        "Algemeen": [
            (29, 2, 0),           # Voorschot Schoonmaak = 0
            (35, 2, "Basis Schoonmaak"), # Pakket
            (39, 2, 0),           # Totaal Kosten (Excel formula might be 0)
        ],
        "Schoonmaak": [
            (4, 2, "Basis Schoonmaak"),
            (6, 2, 13),           # Totaal Uren (High usage)
            (8, 2, 50),           # Uurtarief
        ]
    }
    
    for sheet_name, updates in changes.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row, col, value in updates:
                ws.cell(row=row, column=col, value=value)
                print(f"   - [{sheet_name}] Set ({row}, {col}) to '{value}'")

    wb.save(temp_input)
    wb.close()
    
    # 3. Run Generator
    cmd = [
        'python3', 'generate.py',
        '--input', os.path.basename(temp_input),
        '--no-open',
        '--no-pause'
    ]
    
    try:
        subprocess.run(cmd, cwd=SRC_DIR, check=True, capture_output=True, text=True)
        print(f"   ✅ Generator finished successfully.")
        
        # 4. Rename Output
        files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('_onepager.html')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(OUTPUT_DIR, x)), reverse=True)
        
        if files:
            latest = files[0]
            new_name = "test_missing_voorschot.html"
            src = os.path.join(OUTPUT_DIR, latest)
            dst = os.path.join(OUTPUT_DIR, new_name)
            shutil.move(src, dst)
            print(f"   📂 Saved report to: {new_name}")
            return dst
            
    except subprocess.CalledProcessError as e:
        print(f"   ❌ Generator failed: {e.stderr}")
        return None
    finally:
        if os.path.exists(temp_input):
            os.remove(temp_input)

if __name__ == "__main__":
    run_test()
