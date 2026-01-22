#!/usr/bin/env python3
"""
Sales Dashboard Excel Generator v2
Creates a robust, DYNAMIC multi-sheet Excel workbook for the Sales team.

Key Features:
- Address-based lookups (not object IDs)
- Eigenaar lookup sheet
- Huurder lookup sheet  
- Dynamic KPIs using Excel formulas
- All data references the Data table for easy updates

Usage:
    python generate_sales_dashboard.py
"""

import sqlite3
import os
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
import pandas as pd

# Constants
DB_PATH = Path(__file__).parent.parent / "database" / "ryanrent_v2.db"
OUTPUT_DIR = Path(__file__).parent.parent / "exports"

# Styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
KPI_FONT = Font(bold=True, size=28, color="1F4E79")
KPI_LABEL_FONT = Font(bold=True, size=11, color="1F4E79")
MONEY_FONT = Font(bold=True, size=20, color="228B22")
LINK_FONT = Font(color="0563C1", underline="single")
INPUT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_sales_data() -> pd.DataFrame:
    """Fetch all data from v_sales view."""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM v_sales", conn)
    conn.close()
    return df


def auto_column_width(ws, max_width: int = 45):
    """Auto-adjust column widths."""
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 2, max_width)


def create_data_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 1: Raw data as Excel Table for dynamic references."""
    ws = wb.active
    ws.title = "Data"
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create as Excel Table - critical for dynamic formulas
    table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    table = Table(displayName="Woningen", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    ws.freeze_panes = "A2"
    auto_column_width(ws)
    
    # Create helper column for address lookup (Object + Address combined)
    # Add this as extra column after main data
    helper_col = len(df.columns) + 1
    ws.cell(row=1, column=helper_col, value="ZoekAdres")
    ws.cell(row=1, column=helper_col).fill = HEADER_FILL
    ws.cell(row=1, column=helper_col).font = HEADER_FONT
    
    for row_idx in range(2, len(df) + 2):
        # Combine object_id + adres + plaats for easy searching
        formula = f'=A{row_idx}&" - "&C{row_idx}&", "&E{row_idx}'
        ws.cell(row=row_idx, column=helper_col, value=formula)
    
    print(f"  ✓ Data sheet: {len(df)} properties as Excel Table 'Woningen'")
    return helper_col  # Return position of helper column


def create_dashboard_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 2: Dynamic Dashboard with formula-based KPIs."""
    ws = wb.create_sheet("Dashboard")
    
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "SALES DASHBOARD"
    ws["A1"].font = Font(bold=True, size=28, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Live data - update door nieuw data in 'Data' sheet te plakken"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="666666", size=10)
    
    # ========================================
    # ROW 4-5: Portfolio KPIs (DYNAMIC FORMULAS)
    # ========================================
    kpis = [
        ("B4", "Totaal Woningen", '=ROWS(Woningen[object_id])'),
        ("D4", "Actief", '=COUNTIF(Woningen[huis_actief],1)'),
        ("F4", "Beschikbaar", '=COUNTIF(Woningen[beschikbaarheid_status],"Beschikbaar")'),
        ("B6", "Verhuurd", '=COUNTIFS(Woningen[beschikbaarheid_status],"Verhu*")'),
        ("D6", "Binnenkort Vrij", '=COUNTIF(Woningen[beschikbaarheid_status],"Binnenkort*")'),
        ("F6", "Leegstand", '=COUNTIF(Woningen[beschikbaarheid_status],"Leegstand")'),
    ]
    
    for cell_ref, label, formula in kpis:
        # Label
        ws[cell_ref] = label
        ws[cell_ref].font = KPI_LABEL_FONT
        ws[cell_ref].alignment = Alignment(horizontal="center")
        
        # Value (formula) - one row below
        value_row = int(cell_ref[1:]) + 1
        value_cell = f"{cell_ref[0]}{value_row}"
        ws[value_cell] = formula
        ws[value_cell].font = KPI_FONT
        ws[value_cell].alignment = Alignment(horizontal="center")
    
    # ========================================
    # ROW 9-10: Financial KPIs (DYNAMIC)
    # ========================================
    ws["B9"] = "Totale Marge/Maand"
    ws["B9"].font = KPI_LABEL_FONT
    ws["B10"] = '=SUM(Woningen[marge_maand_excl_btw])'
    ws["B10"].font = MONEY_FONT
    ws["B10"].number_format = '€ #,##0'
    
    ws["D9"] = "Gem. Marge/Woning"
    ws["D9"].font = KPI_LABEL_FONT
    ws["D10"] = '=AVERAGE(Woningen[marge_maand_excl_btw])'
    ws["D10"].font = MONEY_FONT
    ws["D10"].number_format = '€ #,##0'
    
    ws["F9"] = "Jaarlijkse Marge"
    ws["F9"].font = KPI_LABEL_FONT
    ws["F10"] = '=SUM(Woningen[marge_maand_excl_btw])*12'
    ws["F10"].font = MONEY_FONT
    ws["F10"].number_format = '€ #,##0'
    
    # ========================================
    # ROW 12+: Quick Stats Tables
    # ========================================
    ws["A12"] = "TOP 10 HOOGSTE MARGE"
    ws["A12"].font = SECTION_FONT
    ws["A12"].fill = SECTION_FILL
    ws.merge_cells("A12:C12")
    
    # Headers for top margin
    ws["A13"] = "Adres"
    ws["B13"] = "Plaats"
    ws["C13"] = "Marge/Maand"
    for cell in ["A13", "B13", "C13"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    # Top 10 margin (static but sorted from data)
    top_margin = df.nlargest(10, 'marge_maand_excl_btw')
    for i, (_, row) in enumerate(top_margin.iterrows(), 14):
        ws[f"A{i}"] = row['adres']
        ws[f"B{i}"] = row['plaats']
        ws[f"C{i}"] = row['marge_maand_excl_btw']
        ws[f"C{i}"].number_format = '€ #,##0'
    
    # Properties ending soon
    ws["E12"] = "BINNENKORT BESCHIKBAAR"
    ws["E12"].font = SECTION_FONT
    ws["E12"].fill = SECTION_FILL
    ws.merge_cells("E12:G12")
    
    ws["E13"] = "Adres"
    ws["F13"] = "Eind Datum"
    ws["G13"] = "Dagen"
    for cell in ["E13", "F13", "G13"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    # Properties ending in next 60 days
    ending_soon = df[
        (df['dagen_tot_beschikbaar'] > 0) & 
        (df['dagen_tot_beschikbaar'] <= 60)
    ].sort_values('dagen_tot_beschikbaar').head(10)
    
    for i, (_, row) in enumerate(ending_soon.iterrows(), 14):
        ws[f"E{i}"] = row['adres']
        ws[f"F{i}"] = row['verhuur_eind']
        ws[f"G{i}"] = row['dagen_tot_beschikbaar']
    
    # City breakdown
    ws["A26"] = "WONINGEN PER STAD"
    ws["A26"].font = SECTION_FONT
    ws["A26"].fill = SECTION_FILL
    ws.merge_cells("A26:B26")
    
    ws["A27"] = "Stad"
    ws["B27"] = "Aantal"
    ws["A27"].font = HEADER_FONT
    ws["A27"].fill = HEADER_FILL
    ws["B27"].font = HEADER_FONT
    ws["B27"].fill = HEADER_FILL
    
    city_counts = df['plaats'].value_counts().head(15)
    for i, (city, count) in enumerate(city_counts.items(), 28):
        ws[f"A{i}"] = city if city else "(Onbekend)"
        ws[f"B{i}"] = count
    
    auto_column_width(ws)
    print("  ✓ Dashboard with dynamic KPI formulas")


def create_woning_lookup_sheet(wb: Workbook, df: pd.DataFrame, helper_col: int):
    """Sheet 3: Address-based property lookup."""
    ws = wb.create_sheet("Zoek Woning")
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "WONING OPZOEKEN"
    ws["A1"].font = Font(bold=True, size=20, color="1F4E79")
    
    # Instructions
    ws["A3"] = "Typ adres of kies uit lijst:"
    ws["A3"].font = Font(bold=True)
    
    # Search input cell - uses combined address field
    ws["B3"] = "(typ hier)"
    ws["B3"].fill = INPUT_FILL
    ws["B3"].font = Font(size=12)
    ws.merge_cells("B3:D3")
    
    # Create dropdown with "Object - Address, City" format
    addresses = sorted([
        f"{row['object_id']} - {row['adres']}, {row['plaats']}" 
        for _, row in df.iterrows() 
        if pd.notna(row['adres'])
    ])
    
    # Excel has 255 char limit for validation, so use named range approach
    # Write address list to hidden area
    for i, addr in enumerate(addresses, 1):
        ws.cell(row=i, column=20, value=addr)  # Column T (hidden)
    
    ws.column_dimensions['T'].hidden = True
    
    # Data validation from the list
    dv = DataValidation(
        type="list",
        formula1=f'$T$1:$T${len(addresses)}',
        allowBlank=True
    )
    dv.error = "Selecteer een geldig adres"
    dv.prompt = "Begin te typen of selecteer"
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    
    # Extract object_id from selection for lookups
    ws["F3"] = '=IFERROR(LEFT(B3,FIND(" -",B3)-1),"")'
    ws["F3"].font = Font(color="999999", size=9)
    
    # ========================================
    # Property Info Card
    # ========================================
    sections = [
        ("A5", "WONING", [
            ("A6", "Object ID:", "object_id"),
            ("A7", "Adres:", "adres"),
            ("A8", "Postcode:", "postcode"),
            ("A9", "Plaats:", "plaats"),
            ("A10", "Type:", "woning_type"),
            ("A11", "Oppervlakte m²:", "oppervlakte_m2"),
            ("A12", "Slaapkamers:", "aantal_slaapkamers"),
            ("A13", "Capaciteit:", "capaciteit_personen"),
            ("A14", "Status:", "beschikbaarheid_status"),
            ("A15", "Kluis 1:", "kluis_code_1"),
            ("A16", "Kluis 2:", "kluis_code_2"),
        ]),
        ("D5", "EIGENAAR", [
            ("D6", "Naam:", "eigenaar_naam"),
            ("D7", "Email:", "eigenaar_email"),
            ("D8", "Telefoon:", "eigenaar_telefoon"),
            ("D9", "Adres:", "eigenaar_adres"),
            ("D10", "Postcode:", "eigenaar_postcode"),
            ("D11", "Plaats:", "eigenaar_plaats"),
            ("D12", "IBAN:", "eigenaar_iban"),
        ]),
        ("D14", "HUURDER", [
            ("D15", "Naam:", "huurder_naam"),
            ("D16", "Email:", "huurder_email"),
            ("D17", "Telefoon:", "huurder_telefoon"),
            ("D18", "Type:", "huurder_type"),
        ]),
        ("A18", "INHUUR (van eigenaar)", [
            ("A19", "Start:", "inhuur_start"),
            ("A20", "Eind:", "inhuur_eind"),
            ("A21", "Bedrag/maand:", "inhuur_totaal_excl_btw"),
            ("A22", "Borg:", "inhuur_borg"),
            ("A23", "Status:", "inhuur_status"),
        ]),
        ("D20", "VERHUUR (aan huurder)", [
            ("D21", "Start:", "verhuur_start"),
            ("D22", "Eind:", "verhuur_eind"),
            ("D23", "Bedrag/maand:", "verhuur_totaal_excl_btw"),
            ("D24", "Borg:", "verhuur_borg"),
        ]),
        ("A25", "MARGE", [
            ("A26", "Marge/maand:", "marge_maand_excl_btw"),
            ("A27", "Marge/jaar:", "marge_jaar_excl_btw"),
            ("A28", "Marge %:", "marge_percentage"),
        ]),
        ("D26", "DOCUMENTEN", [
            ("D27", "Woning Folder:", "sharepoint_folder"),
            ("D28", "Inhuur Docs:", "sharepoint_inhuur"),
            ("D29", "Verhuur Docs:", "sharepoint_verhuur"),
        ]),
    ]
    
    # Build column index map
    col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}
    
    for section_cell, section_title, fields in sections:
        # Section header
        ws[section_cell] = section_title
        ws[section_cell].font = SECTION_FONT
        ws[section_cell].fill = SECTION_FILL
        
        for cell_ref, label, col_name in fields:
            # Label
            ws[cell_ref] = label
            ws[cell_ref].font = Font(bold=True)
            
            # Value with VLOOKUP using object_id extracted from address
            value_col = cell_ref[0]
            value_col_next = chr(ord(value_col) + 1)
            value_cell = f"{value_col_next}{cell_ref[1:]}"
            
            if col_name in col_indices:
                col_idx = col_indices[col_name]
                # VLOOKUP on object_id (extracted from address selection)
                formula = f'=IFERROR(VLOOKUP($F$3,Data!$A:${get_column_letter(len(df.columns))},{col_idx},FALSE),"")'
                ws[value_cell] = formula
                
                # Format money columns
                if 'excl_btw' in col_name or 'borg' in col_name or 'marge' in col_name:
                    ws[value_cell].number_format = '€ #,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 35
    
    print("  ✓ Zoek Woning sheet with address-based lookup")


def create_eigenaar_lookup_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 4: Owner/Eigenaar lookup."""
    ws = wb.create_sheet("Zoek Eigenaar")
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "EIGENAAR OPZOEKEN"
    ws["A1"].font = Font(bold=True, size=20, color="1F4E79")
    
    ws["A3"] = "Selecteer eigenaar:"
    ws["A3"].font = Font(bold=True)
    
    ws["B3"] = "(kies)"
    ws["B3"].fill = INPUT_FILL
    ws.merge_cells("B3:C3")
    
    # Get unique owners
    owners = df[['eigenaar_id', 'eigenaar_naam']].drop_duplicates().dropna(subset=['eigenaar_naam'])
    owner_names = sorted(owners['eigenaar_naam'].unique().tolist())
    
    # Write to hidden column
    for i, name in enumerate(owner_names, 1):
        ws.cell(row=i, column=15, value=name)
    ws.column_dimensions['O'].hidden = True
    
    dv = DataValidation(
        type="list",
        formula1=f'$O$1:$O${len(owner_names)}',
        allowBlank=True
    )
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    
    # Owner info section
    ws["A5"] = "EIGENAAR GEGEVENS"
    ws["A5"].font = SECTION_FONT
    ws["A5"].fill = SECTION_FILL
    ws.merge_cells("A5:C5")
    
    owner_fields = [
        ("A6", "Naam:", "eigenaar_naam"),
        ("A7", "Email:", "eigenaar_email"),
        ("A8", "Telefoon:", "eigenaar_telefoon"),
        ("A9", "Adres:", "eigenaar_adres"),
        ("A10", "Postcode:", "eigenaar_postcode"),
        ("A11", "Plaats:", "eigenaar_plaats"),
        ("A12", "IBAN:", "eigenaar_iban"),
        ("A13", "KVK:", "eigenaar_kvk"),
        ("A14", "BTW nr:", "eigenaar_btw"),
    ]
    
    col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}
    
    for cell_ref, label, col_name in owner_fields:
        ws[cell_ref] = label
        ws[cell_ref].font = Font(bold=True)
        
        value_cell = f"B{cell_ref[1:]}"
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            # Use eigenaar_naam column for lookup
            eigenaar_naam_col = col_indices.get('eigenaar_naam', 1)
            formula = f'=IFERROR(VLOOKUP($B$3,Data!${get_column_letter(eigenaar_naam_col)}:${get_column_letter(len(df.columns))},{col_idx - eigenaar_naam_col + 1},FALSE),"")'
            ws[value_cell] = formula
    
    # Properties owned by this owner
    ws["A16"] = "WONINGEN VAN DEZE EIGENAAR"
    ws["A16"].font = SECTION_FONT
    ws["A16"].fill = SECTION_FILL
    ws.merge_cells("A16:E16")
    
    ws["A17"] = "Object"
    ws["B17"] = "Adres"
    ws["C17"] = "Plaats"
    ws["D17"] = "Huurder"
    ws["E17"] = "Marge"
    for cell in ["A17", "B17", "C17", "D17", "E17"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    # Add rows that filter by selected owner (using array formula concept, but simplified)
    # In practice, users can filter the Data sheet directly
    ws["A18"] = "(Filter de 'Data' sheet op eigenaar naam voor volledige lijst)"
    ws["A18"].font = Font(italic=True, color="666666")
    
    auto_column_width(ws)
    print("  ✓ Zoek Eigenaar sheet")


def create_huurder_lookup_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 5: Tenant/Huurder lookup."""
    ws = wb.create_sheet("Zoek Huurder")
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "HUURDER OPZOEKEN"
    ws["A1"].font = Font(bold=True, size=20, color="1F4E79")
    
    ws["A3"] = "Selecteer huurder:"
    ws["A3"].font = Font(bold=True)
    
    ws["B3"] = "(kies)"
    ws["B3"].fill = INPUT_FILL
    ws.merge_cells("B3:C3")
    
    # Get unique tenants
    tenants = df[['huurder_id', 'huurder_naam']].drop_duplicates().dropna(subset=['huurder_naam'])
    tenant_names = sorted(tenants['huurder_naam'].unique().tolist())
    
    # Write to hidden column
    for i, name in enumerate(tenant_names, 1):
        ws.cell(row=i, column=15, value=name)
    ws.column_dimensions['O'].hidden = True
    
    dv = DataValidation(
        type="list",
        formula1=f'$O$1:$O${len(tenant_names)}',
        allowBlank=True
    )
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    
    # Tenant info section
    ws["A5"] = "HUURDER GEGEVENS"
    ws["A5"].font = SECTION_FONT
    ws["A5"].fill = SECTION_FILL
    ws.merge_cells("A5:C5")
    
    tenant_fields = [
        ("A6", "Naam:", "huurder_naam"),
        ("A7", "Type:", "huurder_type"),
        ("A8", "Email:", "huurder_email"),
        ("A9", "Telefoon:", "huurder_telefoon"),
        ("A10", "Adres:", "huurder_adres"),
        ("A11", "Postcode:", "huurder_postcode"),
        ("A12", "Plaats:", "huurder_plaats"),
        ("A13", "KVK:", "huurder_kvk"),
    ]
    
    col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}
    
    for cell_ref, label, col_name in tenant_fields:
        ws[cell_ref] = label
        ws[cell_ref].font = Font(bold=True)
        
        value_cell = f"B{cell_ref[1:]}"
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            huurder_naam_col = col_indices.get('huurder_naam', 1)
            formula = f'=IFERROR(VLOOKUP($B$3,Data!${get_column_letter(huurder_naam_col)}:${get_column_letter(len(df.columns))},{col_idx - huurder_naam_col + 1},FALSE),"")'
            ws[value_cell] = formula
    
    # Properties rented by this tenant
    ws["A15"] = "GEHUURDE WONINGEN"
    ws["A15"].font = SECTION_FONT
    ws["A15"].fill = SECTION_FILL
    ws.merge_cells("A15:E15")
    
    ws["A16"] = "Object"
    ws["B16"] = "Adres"
    ws["C16"] = "Plaats"
    ws["D16"] = "Start"
    ws["E16"] = "Eind"
    for cell in ["A16", "B16", "C16", "D16", "E16"]:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    ws["A17"] = "(Filter de 'Data' sheet op huurder naam voor volledige lijst)"
    ws["A17"].font = Font(italic=True, color="666666")
    
    auto_column_width(ws)
    print("  ✓ Zoek Huurder sheet")


def create_print_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 6: Printable info card linked to Zoek Woning."""
    ws = wb.create_sheet("Print Kaart")
    
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "WONING INFORMATIEKAART"
    ws["A1"].font = Font(bold=True, size=18, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Pull from lookup sheet
    ws["A3"] = "Object:"
    ws["B3"] = "='Zoek Woning'!B6"
    ws["B3"].font = Font(bold=True, size=14)
    ws["D3"] = "Status:"
    ws["E3"] = "='Zoek Woning'!B14"
    
    ws["A4"] = "Adres:"
    ws["B4"] = "='Zoek Woning'!B7"
    ws["D4"] = "Plaats:"
    ws["E4"] = "='Zoek Woning'!B9"
    
    # Divider
    ws.merge_cells("A6:F6")
    ws["A6"].fill = HEADER_FILL
    
    # Specs
    ws["A7"] = "SPECIFICATIES"
    ws["A7"].font = Font(bold=True, color="1F4E79")
    
    ws["A8"] = "Type:"
    ws["B8"] = "='Zoek Woning'!B10"
    ws["C8"] = "Opp.:"
    ws["D8"] = "='Zoek Woning'!B11"
    
    ws["A9"] = "Slaapkamers:"
    ws["B9"] = "='Zoek Woning'!B12"
    ws["C9"] = "Capaciteit:"
    ws["D9"] = "='Zoek Woning'!B13"
    
    ws["A10"] = "Kluis 1:"
    ws["B10"] = "='Zoek Woning'!B15"
    ws["C10"] = "Kluis 2:"
    ws["D10"] = "='Zoek Woning'!B16"
    
    # Owner
    ws.merge_cells("A12:F12")
    ws["A12"] = "EIGENAAR"
    ws["A12"].font = Font(bold=True, color="FFFFFF")
    ws["A12"].fill = HEADER_FILL
    
    ws["A13"] = "Naam:"
    ws["B13"] = "='Zoek Woning'!E6"
    ws["A14"] = "Email:"
    ws["B14"] = "='Zoek Woning'!E7"
    ws["A15"] = "Telefoon:"
    ws["B15"] = "='Zoek Woning'!E8"
    
    # Tenant
    ws.merge_cells("A17:F17")
    ws["A17"] = "HUURDER"
    ws["A17"].font = Font(bold=True, color="FFFFFF")
    ws["A17"].fill = HEADER_FILL
    
    ws["A18"] = "Naam:"
    ws["B18"] = "='Zoek Woning'!E15"
    ws["A19"] = "Email:"
    ws["B19"] = "='Zoek Woning'!E16"
    ws["A20"] = "Telefoon:"
    ws["B20"] = "='Zoek Woning'!E17"
    
    # Financial
    ws.merge_cells("A22:F22")
    ws["A22"] = "FINANCIEEL"
    ws["A22"].font = Font(bold=True, color="FFFFFF")
    ws["A22"].fill = HEADER_FILL
    
    ws["A23"] = "Inhuur:"
    ws["B23"] = "='Zoek Woning'!B21"
    ws["C23"] = "Verhuur:"
    ws["D23"] = "='Zoek Woning'!E23"
    
    ws["A24"] = "Marge/maand:"
    ws["B24"] = "='Zoek Woning'!B26"
    ws["C24"] = "Marge/jaar:"
    ws["D24"] = "='Zoek Woning'!B27"
    
    # Documents
    ws.merge_cells("A26:F26")
    ws["A26"] = "DOCUMENTEN"
    ws["A26"].font = Font(bold=True, color="FFFFFF")
    ws["A26"].fill = HEADER_FILL
    
    ws["A27"] = "Folder:"
    ws["B27"] = "='Zoek Woning'!E27"
    
    # Footer
    ws["A29"] = f"Gegenereerd: {datetime.now().strftime('%d-%m-%Y')}"
    ws["A29"].font = Font(italic=True, size=8, color="666666")
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.print_area = "A1:F29"
    
    print("  ✓ Print Kaart sheet")


def create_contacts_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 7: All contacts overview."""
    ws = wb.create_sheet("Contacten")
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "ALLE CONTACTEN"
    ws["A1"].font = Font(bold=True, size=18, color="1F4E79")
    
    # Eigenaren
    ws["A3"] = "EIGENAREN"
    ws["A3"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A3"].fill = HEADER_FILL
    ws.merge_cells("A3:G3")
    
    headers = ["Naam", "Email", "Telefoon", "Adres", "Plaats", "IBAN", "# Woningen"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h).font = HEADER_FONT
        ws.cell(row=4, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    owners = df[['eigenaar_naam', 'eigenaar_email', 'eigenaar_telefoon', 
                 'eigenaar_adres', 'eigenaar_plaats', 'eigenaar_iban', 'eigenaar_id']].drop_duplicates().dropna(subset=['eigenaar_naam'])
    
    row = 5
    for _, o in owners.iterrows():
        count = len(df[df['eigenaar_id'] == o['eigenaar_id']])
        ws.cell(row=row, column=1, value=o['eigenaar_naam'])
        ws.cell(row=row, column=2, value=o['eigenaar_email'])
        ws.cell(row=row, column=3, value=o['eigenaar_telefoon'])
        ws.cell(row=row, column=4, value=o['eigenaar_adres'])
        ws.cell(row=row, column=5, value=o['eigenaar_plaats'])
        ws.cell(row=row, column=6, value=o['eigenaar_iban'])
        ws.cell(row=row, column=7, value=count)
        row += 1
    
    # Huurders
    row += 2
    ws[f"A{row}"] = "HUURDERS"
    ws[f"A{row}"].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f"A{row}"].fill = HEADER_FILL
    ws.merge_cells(f"A{row}:G{row}")
    
    row += 1
    headers = ["Naam", "Type", "Email", "Telefoon", "Adres", "Plaats", "# Woningen"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = HEADER_FONT
        ws.cell(row=row, column=col).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    tenants = df[['huurder_naam', 'huurder_type', 'huurder_email', 'huurder_telefoon',
                  'huurder_adres', 'huurder_plaats', 'huurder_id']].drop_duplicates().dropna(subset=['huurder_naam'])
    
    row += 1
    for _, t in tenants.iterrows():
        count = len(df[df['huurder_id'] == t['huurder_id']])
        ws.cell(row=row, column=1, value=t['huurder_naam'])
        ws.cell(row=row, column=2, value=t['huurder_type'])
        ws.cell(row=row, column=3, value=t['huurder_email'])
        ws.cell(row=row, column=4, value=t['huurder_telefoon'])
        ws.cell(row=row, column=5, value=t['huurder_adres'])
        ws.cell(row=row, column=6, value=t['huurder_plaats'])
        ws.cell(row=row, column=7, value=count)
        row += 1
    
    ws.freeze_panes = "A5"
    auto_column_width(ws)
    print(f"  ✓ Contacten sheet: {len(owners)} eigenaren, {len(tenants)} huurders")


def create_financieel_sheet(wb: Workbook, df: pd.DataFrame):
    """Sheet 8: Financial overview with dynamic totals."""
    ws = wb.create_sheet("Financieel")
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "FINANCIEEL OVERZICHT"
    ws["A1"].font = Font(bold=True, size=18, color="1F4E79")
    
    # Dynamic totals using table references
    ws["A3"] = "TOTALEN (dynamisch)"
    ws["A3"].font = SECTION_FONT
    ws["A3"].fill = SECTION_FILL
    ws.merge_cells("A3:D3")
    
    ws["A4"] = "Totale Inhuur/maand:"
    ws["B4"] = '=SUM(Woningen[inhuur_totaal_excl_btw])'
    ws["B4"].number_format = '€ #,##0.00'
    ws["C4"] = "Jaarlijks:"
    ws["D4"] = '=B4*12'
    ws["D4"].number_format = '€ #,##0.00'
    
    ws["A5"] = "Totale Verhuur/maand:"
    ws["B5"] = '=SUM(Woningen[verhuur_totaal_excl_btw])'
    ws["B5"].number_format = '€ #,##0.00'
    ws["D5"] = '=B5*12'
    ws["D5"].number_format = '€ #,##0.00'
    
    ws["A6"] = "Totale Marge/maand:"
    ws["B6"] = '=SUM(Woningen[marge_maand_excl_btw])'
    ws["B6"].number_format = '€ #,##0.00'
    ws["B6"].font = Font(bold=True, color="228B22")
    ws["D6"] = '=B6*12'
    ws["D6"].number_format = '€ #,##0.00'
    ws["D6"].font = Font(bold=True, color="228B22")
    
    # Per-property breakdown (sorted)
    ws["A9"] = "MARGE PER WONING (gesorteerd)"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = HEADER_FILL
    ws["A9"].font = Font(bold=True, color="FFFFFF")
    ws.merge_cells("A9:G9")
    
    headers = ["Object", "Adres", "Inhuur", "Verhuur", "Marge", "Marge %", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=10, column=col, value=h).font = HEADER_FONT
        ws.cell(row=10, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    sorted_df = df.sort_values('marge_maand_excl_btw', ascending=False)
    
    row = 11
    for _, p in sorted_df.iterrows():
        ws.cell(row=row, column=1, value=p['object_id'])
        ws.cell(row=row, column=2, value=p['adres'])
        
        c = ws.cell(row=row, column=3, value=p['inhuur_totaal_excl_btw'])
        c.number_format = '€ #,##0'
        
        c = ws.cell(row=row, column=4, value=p['verhuur_totaal_excl_btw'])
        c.number_format = '€ #,##0'
        
        c = ws.cell(row=row, column=5, value=p['marge_maand_excl_btw'])
        c.number_format = '€ #,##0'
        if p['marge_maand_excl_btw'] and p['marge_maand_excl_btw'] > 0:
            c.font = Font(color="228B22")
        elif p['marge_maand_excl_btw'] and p['marge_maand_excl_btw'] < 0:
            c.font = Font(color="FF0000")
        
        c = ws.cell(row=row, column=6, value=p['marge_percentage'])
        if p['marge_percentage']:
            c.number_format = '0.0%'
        
        ws.cell(row=row, column=7, value=p['beschikbaarheid_status'])
        row += 1
    
    ws.freeze_panes = "A11"
    auto_column_width(ws)
    print("  ✓ Financieel sheet with dynamic totals")


def main():
    """Generate the Sales Dashboard Excel workbook."""
    print("\n" + "=" * 60)
    print("SALES DASHBOARD GENERATOR v2")
    print("=" * 60)
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    print("\n📊 Fetching data from v_sales...")
    df = get_sales_data()
    print(f"   Found {len(df)} properties with {len(df.columns)} columns")
    
    print("\n📝 Creating workbook...")
    wb = Workbook()
    
    helper_col = create_data_sheet(wb, df)
    create_dashboard_sheet(wb, df)
    create_woning_lookup_sheet(wb, df, helper_col)
    create_eigenaar_lookup_sheet(wb, df)
    create_huurder_lookup_sheet(wb, df)
    create_print_sheet(wb, df)
    create_contacts_sheet(wb, df)
    create_financieel_sheet(wb, df)
    
    output_file = OUTPUT_DIR / f"sales_dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print(f"✅ DONE! Created: {output_file}")
    print("=" * 60)
    print("\nSheets:")
    print("  1. Data - Paste new data here, formulas update automatically")
    print("  2. Dashboard - Live KPIs (formulas reference Data table)")
    print("  3. Zoek Woning - Search by ADDRESS")
    print("  4. Zoek Eigenaar - Search by owner name")
    print("  5. Zoek Huurder - Search by tenant name")
    print("  6. Print Kaart - Printable property card")
    print("  7. Contacten - All contacts overview")
    print("  8. Financieel - Dynamic financial analysis")
    print()
    
    return str(output_file)


if __name__ == "__main__":
    main()
