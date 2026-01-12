#!/usr/bin/env python3
"""
Schadekosten Parser
Reads schadekosten Excel files and extracts cleaning + damage costs.

Expected Excel structure:
- Row 1, Col A: Address (e.g., "Korte Singelstraat 21-2")
- Row 4+: Cost items (Omschrijving, Kosten)
- First item is typically "Schoonmaak"
- Last row: "Totaal" with sum

Usage:
    python parse_schadekosten.py [--folder PATH] [--output PATH]
"""

import openpyxl
import os
import sys
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Any, Optional

# Add Shared to path for entities
current_dir = Path(__file__).parent.parent
sys.path.append(str(current_dir / 'Shared'))


@dataclass
class SchadeItem:
    """Single damage/cost item"""
    omschrijving: str
    kosten: float


@dataclass
class SchadeData:
    """Parsed data from one schadekosten Excel"""
    adres: str
    source_file: str
    schoonmaak_kosten: float
    schade_items: List[SchadeItem]
    totaal: float


def parse_schadekosten_excel(filepath: str) -> Optional[SchadeData]:
    """
    Parse a single schadekosten Excel file.
    
    Returns:
        SchadeData object with extracted costs, or None if parsing fails
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Row 1: Address
        adres = str(ws.cell(1, 1).value or "").strip()
        if not adres:
            print(f"  ⚠️ No address found in {filepath}")
            return None
        
        # Parse rows starting from row 4
        schoonmaak_kosten = 0.0
        schade_items = []
        totaal = 0.0
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            omschrijving = str(row[0] or "").strip()
            kosten = row[1] if len(row) > 1 else 0
            
            if not omschrijving:
                continue
                
            # Parse kosten value
            if kosten is None:
                kosten = 0.0
            elif isinstance(kosten, str):
                # Handle string values like "€250" or "250,00"
                kosten = kosten.replace('€', '').replace(',', '.').strip()
                try:
                    kosten = float(kosten)
                except:
                    kosten = 0.0
            else:
                kosten = float(kosten)
            
            # Categorize
            omschrijving_lower = omschrijving.lower()
            
            if 'totaal' in omschrijving_lower:
                totaal = kosten
            elif 'schoonmaak' in omschrijving_lower:
                schoonmaak_kosten = kosten
            else:
                # It's a schade item
                schade_items.append(SchadeItem(
                    omschrijving=omschrijving,
                    kosten=kosten
                ))
        
        return SchadeData(
            adres=adres,
            source_file=filepath,
            schoonmaak_kosten=schoonmaak_kosten,
            schade_items=schade_items,
            totaal=totaal
        )
        
    except Exception as e:
        print(f"  ❌ Error parsing {filepath}: {e}")
        return None


def scan_folder(folder_path: str) -> List[SchadeData]:
    """
    Scan a folder for schadekosten Excel files and parse them all.
    
    Returns:
        List of SchadeData objects
    """
    results = []
    folder = Path(folder_path)
    
    if not folder.exists():
        print(f"❌ Folder not found: {folder_path}")
        return results
    
    # Find all Excel files (excluding temp files starting with ~$)
    excel_files = [f for f in folder.glob("*.xlsx") if not f.name.startswith("~$")]
    
    print(f"📂 Scanning {folder_path}...")
    print(f"   Found {len(excel_files)} Excel files")
    
    for filepath in sorted(excel_files):
        print(f"   📄 Parsing: {filepath.name}")
        data = parse_schadekosten_excel(str(filepath))
        if data:
            results.append(data)
            print(f"      ✅ {data.adres}: Schoonmaak €{data.schoonmaak_kosten:.2f}, {len(data.schade_items)} schade items")
    
    return results


def generate_summary(data_list: List[SchadeData]) -> Dict[str, Any]:
    """Generate a summary of all parsed data"""
    summary = {
        'total_files': len(data_list),
        'by_address': {}
    }
    
    for data in data_list:
        summary['by_address'][data.adres] = {
            'schoonmaak': data.schoonmaak_kosten,
            'schade_items': [(item.omschrijving, item.kosten) for item in data.schade_items],
            'totaal': data.totaal,
            'source': data.source_file
        }
    
    return summary


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Parse schadekosten Excel files')
    parser.add_argument('--folder', default='Shared/Sources/Schade', 
                        help='Folder containing schadekosten Excel files')
    parser.add_argument('--output', default=None,
                        help='Output JSON file for parsed data')
    args = parser.parse_args()
    
    print("🔍 Schadekosten Parser")
    print("=" * 50)
    
    # Scan folder
    data_list = scan_folder(args.folder)
    
    if not data_list:
        print("\n⚠️ No schadekosten data found.")
        return
    
    # Print summary
    print("\n📊 Summary")
    print("-" * 50)
    
    total_schoonmaak = 0
    total_schade = 0
    
    for data in data_list:
        schade_sum = sum(item.kosten for item in data.schade_items)
        total_schoonmaak += data.schoonmaak_kosten
        total_schade += schade_sum
        
        print(f"\n🏠 {data.adres}")
        print(f"   Schoonmaak: €{data.schoonmaak_kosten:.2f}")
        print(f"   Schade ({len(data.schade_items)} items):")
        for item in data.schade_items:
            print(f"      - {item.omschrijving}: €{item.kosten:.2f}")
        print(f"   Totaal: €{data.totaal:.2f}")
    
    print("\n" + "=" * 50)
    print(f"📈 Totaal alle woningen:")
    print(f"   Schoonmaak: €{total_schoonmaak:.2f}")
    print(f"   Schade: €{total_schade:.2f}")
    print(f"   Gecombineerd: €{total_schoonmaak + total_schade:.2f}")
    
    # Save to JSON if requested
    if args.output:
        import json
        summary = generate_summary(data_list)
        with open(args.output, 'w') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        print(f"\n💾 Saved to {args.output}")


if __name__ == "__main__":
    main()
