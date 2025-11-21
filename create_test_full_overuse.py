#!/usr/bin/env python3
"""
Create test_full_overuse.xlsx - Complete overuse scenario with heavy damages
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import shutil
from pathlib import Path
from datetime import datetime

def safe_set_cell(ws, cell_ref, value):
    """Safely set a cell value, handling merged cells."""
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left = ws.cell(min_row, min_col)
                top_left.value = value
                return
    else:
        ws[cell_ref].value = value

def create_full_overuse_test():
    """Create comprehensive test with all types of overuse."""
    
    # Copy clean template
    shutil.copy('input_template.xlsx', 'test_full_overuse.xlsx')
    print("✓ Created test_full_overuse.xlsx from clean template")
    
    wb = openpyxl.load_workbook('test_full_overuse.xlsx')
    
    # Fill Algemeen sheet
    ws_alg = wb['Algemeen']
    safe_set_cell(ws_alg, 'B4', 'Familie Hendriksen')
    safe_set_cell(ws_alg, 'B5', 'Sophie van RyanRent')
    safe_set_cell(ws_alg, 'B6', 't.hendriksen@email.nl')
    safe_set_cell(ws_alg, 'B7', '06-12345678')
    safe_set_cell(ws_alg, 'B10', 'Parkstraat 88')
    safe_set_cell(ws_alg, 'B12', '2000 XY')
    safe_set_cell(ws_alg, 'B13', 'Utrecht')
    
    # Dates - 6 month rental
    start_date = datetime(2024, 6, 1)
    end_date = datetime(2024, 11, 30)
    safe_set_cell(ws_alg, 'B17', start_date.strftime('%d-%m-%Y'))
    safe_set_cell(ws_alg, 'B18', end_date.strftime('%d-%m-%Y'))
    
    # Financial
    # Financial - Set Voorschotten explicitly
    safe_set_cell(ws_alg, 'B22', 4000)  # Voorschot Borg
    safe_set_cell(ws_alg, 'B23', 3900)  # Voorschot GWE
    safe_set_cell(ws_alg, 'B24', 250)   # Voorschot Schoonmaak
    
    # Selections & Rates
    safe_set_cell(ws_alg, 'B28', 'Intensief Schoonmaak')  # Schoonmaak Pakket
    safe_set_cell(ws_alg, 'B29', 7)     # Inbegrepen Uren (Manual override for test)
    safe_set_cell(ws_alg, 'B30', 40)    # Uurtarief Schoonmaak
    
    print("✓ Filled Algemeen sheet")
    
    # Fill GWE_Detail - EXTREME overuse
    ws_gwe = wb['GWE_Detail']
    safe_set_cell(ws_gwe, 'B4', 15000)  # Elektra Begin
    safe_set_cell(ws_gwe, 'B5', 19500)  # Elektra Eind (4500 kWh - HUGE)
    safe_set_cell(ws_gwe, 'B7', 8000)   # Gas Begin
    safe_set_cell(ws_gwe, 'B8', 10200)  # Gas Eind (2200 m³ - HUGE)
    safe_set_cell(ws_gwe, 'B10', 3000)  # Water Begin
    safe_set_cell(ws_gwe, 'B11', 3350)  # Water Eind (350 m³ - HUGE)
    
    # Tarieven
    safe_set_cell(ws_gwe, 'B14', 0.32)  # Elektra Tarief
    safe_set_cell(ws_gwe, 'B15', 1.55)  # Gas Tarief
    safe_set_cell(ws_gwe, 'B16', 1.70)  # Water Tarief
    
    # Low budgets to ensure heavy overage
    safe_set_cell(ws_gwe, 'B19', 2500)  # Elektra Budget
    safe_set_cell(ws_gwe, 'B20', 1200)  # Gas Budget
    safe_set_cell(ws_gwe, 'B21', 200)   # Water Budget
    
    print("✓ Filled GWE_Detail - HEAVY overuse!")
    
    # Write GWE Detail Lines (Rows 12+) manually since openpyxl won't calc formulas
    # Row 12: Elektra
    safe_set_cell(ws_gwe, 'A12', 'Elektra verbruik')
    safe_set_cell(ws_gwe, 'B12', 4500)  # Verbruik
    safe_set_cell(ws_gwe, 'C12', 0.32)  # Tarief
    safe_set_cell(ws_gwe, 'D12', 4500 * 0.32)  # Kosten
    
    # Row 13: Gas
    safe_set_cell(ws_gwe, 'A13', 'Gas verbruik')
    safe_set_cell(ws_gwe, 'B13', 2200)
    safe_set_cell(ws_gwe, 'C13', 1.55)
    safe_set_cell(ws_gwe, 'D13', 2200 * 1.55)
    
    # Row 14: Water
    safe_set_cell(ws_gwe, 'A14', 'Water verbruik')
    safe_set_cell(ws_gwe, 'B14', 350)
    safe_set_cell(ws_gwe, 'C14', 1.70)
    safe_set_cell(ws_gwe, 'D14', 350 * 1.70)
    
    # Fill Schoonmaak - LOTS of overtime
    ws_schoon = wb['Schoonmaak']
    safe_set_cell(ws_schoon, 'B6', 18.0)  # Totaal Uren (vs 7 included = 11 overtime!)
    safe_set_cell(ws_schoon, 'B10', 40)   # Uurtarief
    
    print("✓ Filled Schoonmaak - HEAVY overtime!")
    
    # Fill Schade - 15 damage items
    ws_schade = wb['Schade']
    
    damages = [
        ('Grote schade aan vloerbedekking woonkamer', 1, 450.00),
        ('Meerdere gaten in muren (herstel + schilderen)', 5, 120.00),
        ('Kapotte balkondeur inclusief slot', 1, 380.00),
        ('Beschadigde keukenkastjes (4 stuks)', 4, 95.00),
        ('Gebroken tegels badkamer + voegen', 8, 35.00),
        ('Brandschade op aanrechtblad', 1, 520.00),
        ('Kapotte wasmachine (tenant schade)', 1, 650.00),
        ('Schade aan gordijnen + rails (alle kamers)', 3, 85.00),
        ('Gebroken douchewand', 1, 425.00),
        ('Toegangspoort beschadigd', 1, 280.00),
        ('Waterlek schade plafond slaapkamer', 1, 380.00),
        ('Kapotte verlichting (lampen + armaturen)', 6, 55.00),
        ('Beschadigde radiator woonkamer', 1, 295.00),
        ('Verfschade op meerdere deuren', 4, 75.00),
        ('Kapotte vaatwasser (misbruik)', 1, 580.00),
    ]
    
    start_row = 6
    for i, (beschrijving, aantal, tarief) in enumerate(damages):
        row = start_row + i
        safe_set_cell(ws_schade, f'A{row}', beschrijving)
        safe_set_cell(ws_schade, f'B{row}', aantal)
        safe_set_cell(ws_schade, f'C{row}', tarief)
        safe_set_cell(ws_schade, f'D{row}', aantal * tarief)  # Calculated cost
    
    print(f"✓ Added {len(damages)} serious damage items")
    
    # Save
    wb.save('test_full_overuse.xlsx')
    
    print(f"\n{'='*60}")
    print(f"✅ TEST FILE CREATED: test_full_overuse.xlsx")
    print(f"{'='*60}")
    print(f"\n📊 SCENARIO SUMMARY:")
    print(f"  Client: Familie Hendriksen")
    print(f"  Period: 6 months (Jun 2024 - Nov 2024)")
    print(f"  Rent: €2,000/month | Deposit: €4,000")
    print(f"\n⚡ GWE OVERUSE:")
    print(f"  • Elektra: 4500 kWh used vs 2500 budget = 80% OVER!")
    print(f"  • Gas: 2200 m³ used vs 1200 budget = 83% OVER!")
    print(f"  • Water: 350 m³ used vs 200 budget = 75% OVER!")
    print(f"\n🧹 CLEANING OVERUSE:")
    print(f"  • 18 hours worked vs 7 included = 11 OVERTIME HOURS!")
    print(f"  • Overtime cost: €440 (11 hrs × €40)")
    print(f"\n💥 DAMAGES:")
    print(f"  • {len(damages)} serious damage items")
    print(f"  • Major items: wasmachine, vaatwasser, balkondeur")
    print(f"  • Estimated total: €5,000+ in damages")
    print(f"\n💰 EXPECTED RESULT:")
    print(f"  Client will need to pay significantly!")
    print(f"  All budgets exceeded heavily.")
    print(f"{'='*60}\n")

if __name__ == '__main__':
    create_full_overuse_test()
