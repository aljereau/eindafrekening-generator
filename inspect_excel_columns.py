import openpyxl
import sys
import os

EXCEL_FILES = [
    "Shared/Sources/Houses + Informationrelated.xlsx",
    "Shared/Sources/Klant+Leverancier.xlsx",
    "Shared/Sources/leverancier.xlsx",
    "Shared/Sources/huizen+ archief huizen.xlsx",
]

for file in EXCEL_FILES:
    if not os.path.exists(file):
        print(f"\n❌ File not found: {file}")
        continue
        
    print(f"\n{'='*60}")
    print(f"📄 {file}")
    print('='*60)
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        for sheet_name in wb.sheetnames[:3]:  # Limit to first 3 sheets
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            print(f"\n📋 Sheet: '{sheet_name}'")
            print(f"   Columns ({len(headers)}):")
            for idx, h in enumerate(headers):
                if h:  # Only print non-empty headers
                    print(f"      {idx}: {h}")
    except Exception as e:
        print(f"   ⚠️ Error reading: {e}")
