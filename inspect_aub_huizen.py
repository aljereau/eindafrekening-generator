import openpyxl

file = "Shared/Sources/aub_huizenlijst.xlsx"
wb = openpyxl.load_workbook(file, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n📋 Sheet: '{sheet_name}'")
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"   Columns ({len([h for h in headers if h])}):")
    for idx, h in enumerate(headers):
        if h:
            print(f"      {idx}: {h}")
    
    # Sample first 2 data rows
    print("\n   Sample Data (first 2 rows):")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=3, values_only=True), start=2):
        print(f"      Row {row_idx}: {row[:10]}...")  # First 10 cols
