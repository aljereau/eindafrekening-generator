import sys
import os
from datetime import date

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

from HuizenManager.src.bot import RyanRentBot

def test_batch_generation():
    print("🤖 Initializing Bot...")
    # Initialize bot (mocking keys where possible, but we need DB access)
    bot = RyanRentBot(provider="ollama", model_name="qwen3-coder:30b", api_key="dummy")
    
    print("\n🧪 Testing 'generate_excel_template' for Settlement Inputs...")
    
    # Simulate Tool Call
    args = {
        "template_type": "settlement_inputs",
        "filters": {
            "start_date": "2024-12-01",
            "end_date": "2025-02-01"
        }
    }
    
    try:
        result = bot._execute_tool("generate_excel_template", args)
        print("\n✅ Result:")
        print(result)
        
        if result.get('success'):
            # Inspect one file
            import glob
            import openpyxl
            
            output_dir = result['message'].split(": ")[1].split(". Please")[0]
            files = glob.glob(os.path.join(output_dir, "*.xlsx"))
            
            if files:
                f = files[0]
                print(f"\n🔎 Inspecting: {f}")
                wb = openpyxl.load_workbook(f)
                ws = wb['Algemeen']
                print(f"Adres (B10): {ws['B10'].value}")
                print(f"Postcode (B13): {ws['B13'].value} (Formula: {ws['B13'].value})")
                print(f"Plaats (B14): {ws['B14'].value} (Formula: {ws['B14'].value})")
                print(f"GWE Maand (B24): {ws['B24'].value}")
                print(f"Voorschot GWE (B25): {ws['B25'].value} (Formula: {ws['B25'].value})")
                
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_batch_generation()
