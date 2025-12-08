import os
import pandas as pd
from datetime import datetime
import shutil
from typing import List, Dict, Optional
import sys

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

from HuizenManager.src.intelligence_api import IntelligenceAPI

class FileExchangeHandler:
    """
    Handles the generation of Excel templates and processing of filled input files.
    Acts as a bridge between the AI Bot and the Database/Scripts.
    """

    def __init__(self, base_path: str = None):
        if base_path:
            self.base_path = base_path
        else:
            # Default to 'Shared/Exchange' in the project root
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))
            self.base_path = os.path.join(project_root, 'Shared', 'Exchange')

        self.templates_dir = os.path.join(self.base_path, 'Templates')
        self.input_dir = os.path.join(self.base_path, 'Input')
        self.processed_dir = os.path.join(self.base_path, 'Processed')

        # Ensure directories exist
        os.makedirs(self.templates_dir, exist_ok=True)
        os.makedirs(self.input_dir, exist_ok=True)
        os.makedirs(self.processed_dir, exist_ok=True)

        self.api = IntelligenceAPI()

    def generate_template(self, request_type: str, data: List[Dict] = None, filename: str = None) -> str:
        """
        Generates an Excel template based on the request type.
        
        Args:
            request_type: 'checkout_update', 'settlement_batch', etc.
            data: Optional list of dicts to pre-fill the template.
            filename: Optional custom filename.
            
        Returns:
            Absolute path to the generated file.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if request_type == 'checkout_update':
            df = self._create_checkout_update_template(data)
            default_name = f"update_checkouts_{timestamp}.xlsx"
            
        elif request_type == 'settlement_batch':
            df = self._create_settlement_batch_template(data)
            default_name = f"generate_settlements_{timestamp}.xlsx"
            
        else:
            raise ValueError(f"Unknown request_type: {request_type}")

        if not filename:
            filename = default_name
            
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        file_path = os.path.join(self.templates_dir, filename)
        
        # Save to Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Input')
            
            # Auto-adjust column widths (basic approximation)
            worksheet = writer.sheets['Input']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)

        return file_path

    def process_file(self, filename: str, request_type: str) -> Dict:
        """
        Reads a file from the Input directory and processes it.
        
        Args:
            filename: Name of the file in Exchange/Input.
            request_type: Type of processing to apply.
            
        Returns:
            Dictionary with results (success count, errors, etc.)
        """
        input_path = os.path.join(self.input_dir, filename)
        
        if not os.path.exists(input_path):
            return {"success": False, "error": f"File not found: {input_path}"}

        try:
            df = pd.read_excel(input_path)
            
            results = {
                "total_rows": len(df),
                "success_count": 0,
                "errors": []
            }

            if request_type == 'checkout_update':
                results = self._process_checkout_updates(df, results)
            elif request_type == 'settlement_batch':
                results = self._process_settlement_batch(df, results)
            else:
                return {"success": False, "error": f"Unknown request_type: {request_type}"}

            # Move to Processed if successful (or partially successful)
            if results['success_count'] > 0:
                archive_name = f"PROCESSED_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                shutil.move(input_path, os.path.join(self.processed_dir, archive_name))
                results['archived_as'] = archive_name

            return results

        except Exception as e:
            return {"success": False, "error": str(e)}

    # ==========================================
    # Template Generators
    # ==========================================

    def generate_settlement_inputs(self, bookings: List[Dict]) -> str:
        """
        Generates individual pre-filled input templates for a list of bookings.
        
        Args:
            bookings: List of booking dictionaries (must contain id, client, object, dates, financials).
            
        Returns:
            Path to the directory containing the generated files.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        batch_dir = os.path.join(self.base_path, 'Output', f"Batch_Settlements_{timestamp}")
        os.makedirs(batch_dir, exist_ok=True)
        
        # Path to master template
        # Try Shared/Templates first, then Eindafrekening/src
        master_template = os.path.join(self.templates_dir, "RyanRent_Input_Template.xlsx")
        if not os.path.exists(master_template):
             # Fallback to source
             master_template = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'Eindafrekening', 'src', 'input_template.xlsx'))
             
        if not os.path.exists(master_template):
            raise FileNotFoundError("Master input template not found.")

        import openpyxl
        
        generated_count = 0
        for booking in bookings:
            try:
                # Load template
                wb = openpyxl.load_workbook(master_template)
                
                # Fill 'Algemeen' sheet
                if 'Algemeen' in wb.sheetnames:
                    ws = wb['Algemeen']
                    
                    # Helper to set value by named range
                    def set_value(name, value):
                        if name in wb.defined_names:
                            dests = wb.defined_names[name].destinations
                            for sheet, cell in dests:
                                if sheet == 'Algemeen':
                                    ws[cell.replace('$', '')] = value
                                    
                    # Fill Data
                    set_value('Klantnaam', booking.get('client_name', ''))
                    set_value('Object_adres', booking.get('adres') or booking.get('property_address', ''))
                    # Postcode/Plaats are now VLOOKUPs, do not overwrite
                    # set_value('Postcode', booking.get('postcode', ''))
                    # set_value('Plaats', booking.get('plaats', ''))
                    set_value('Incheck_datum', booking.get('checkin_datum', ''))
                    set_value('Uitcheck_datum', booking.get('checkout_datum', ''))
                    
                    # Financials
                    if 'deposit' in booking: set_value('Voorschot_borg', booking['deposit'])
                    if 'rent' in booking: set_value('Kale_huur', booking['rent'])
                    
                    # GWE: Write monthly amount to GWE_Maandbedrag
                    # The formula in Voorschot_GWE will calculate the total
                    if 'voorschot_gwe' in booking: set_value('GWE_Maandbedrag', booking['voorschot_gwe'])
                    
                    # Save
                    client_safe = "".join([c for c in booking.get('client_name', 'Client') if c.isalnum() or c in (' ', '_')]).strip()
                    filename = f"Settlement_{client_safe}_{booking.get('id')}.xlsx"
                    wb.save(os.path.join(batch_dir, filename))
                    generated_count += 1
                    
            except Exception as e:
                print(f"Failed to generate for booking {booking.get('id')}: {e}")
                
        return batch_dir

    def _create_checkout_update_template(self, data: List[Dict]) -> pd.DataFrame:
        """Columns: BookingID, Address, Client, CurrentCheckout, NewCheckout"""
        if not data:
            data = []
            
        rows = []
        for item in data:
            rows.append({
                "BookingID": item.get('booking_id'),
                "Address": item.get('adres') or item.get('property_address'),
                "Client": item.get('client') or item.get('client_name'),
                "CurrentCheckout": item.get('checkout_datum') or item.get('date'),
                "NewCheckout": "" # User fills this
            })
            
        if not rows:
            # Create empty structure
            return pd.DataFrame(columns=["BookingID", "Address", "Client", "CurrentCheckout", "NewCheckout"])
            
        return pd.DataFrame(rows)

    def _create_settlement_batch_template(self, data: List[Dict]) -> pd.DataFrame:
        """Columns: BookingID, Address, Client, CheckoutDate, Generate(Y/N)"""
        if not data:
            data = []
            
        rows = []
        for item in data:
            rows.append({
                "BookingID": item.get('booking_id'),
                "Address": item.get('adres') or item.get('property_address'),
                "Client": item.get('client') or item.get('client_name'),
                "CheckoutDate": item.get('checkout_datum') or item.get('date'),
                "Generate": "Y"
            })
            
        if not rows:
            return pd.DataFrame(columns=["BookingID", "Address", "Client", "CheckoutDate", "Generate"])
            
        return pd.DataFrame(rows)

    # ==========================================
    # Processors
    # ==========================================

    def _process_checkout_updates(self, df: pd.DataFrame, results: Dict) -> Dict:
        """Update checkout dates in DB."""
        required_cols = ['BookingID', 'NewCheckout']
        if not all(col in df.columns for col in required_cols):
            results['errors'].append(f"Missing columns. Required: {required_cols}")
            return results

        for index, row in df.iterrows():
            booking_id = row['BookingID']
            new_date = row['NewCheckout']
            
            # Skip if empty
            if pd.isna(new_date) or str(new_date).strip() == '':
                continue
                
            # Format date
            try:
                if isinstance(new_date, datetime):
                    date_str = new_date.strftime("%Y-%m-%d")
                else:
                    # Assume string YYYY-MM-DD
                    date_str = str(new_date).strip()
                    # Basic validation
                    datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                results['errors'].append(f"Row {index+2}: Invalid date format '{new_date}' for Booking {booking_id}")
                continue

            # Execute Update
            response = self.api.update_booking(booking_id=int(booking_id), checkout_date=date_str)
            
            if response.get('success'):
                results['success_count'] += 1
            else:
                results['errors'].append(f"Row {index+2}: {response.get('error')}")
                
        return results

    def _process_settlement_batch(self, df: pd.DataFrame, results: Dict) -> Dict:
        """Trigger settlement generation."""
        required_cols = ['BookingID', 'Generate']
        if not all(col in df.columns for col in required_cols):
            results['errors'].append(f"Missing columns. Required: {required_cols}")
            return results

        for index, row in df.iterrows():
            if str(row['Generate']).upper() != 'Y':
                continue
                
            booking_id = row['BookingID']
            
            # For now, we just verify we can fetch the booking. 
            # In a real scenario, this would call generate.py logic.
            # Since generate.py is a script, we might need to invoke it or import its logic.
            # For this MVP, let's use the API's generate_settlement if available, or just log success.
            
            try:
                # Fetch booking details to simulate "Ready to generate"
                booking = self.api.get_booking_for_settlement(int(booking_id))
                
                # Here we would call the actual generation logic
                # For now, we'll assume success if the booking exists and data is valid
                results['success_count'] += 1
                
            except Exception as e:
                results['errors'].append(f"Row {index+2}: Failed to prep settlement for Booking {booking_id}: {e}")
                
        return results
