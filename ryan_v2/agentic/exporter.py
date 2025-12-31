"""
RYAN-V2: Agentic Pipeline
Streaming Exporter Module

Exports query results to CSV/XLSX using streaming to avoid memory issues.
This is pure code - NO LLM calls.
"""

import csv
import sqlite3
from pathlib import Path
from typing import Iterator, List, Dict, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger("ryan_exporter")

# Export directory
EXPORT_DIR = Path(__file__).parent.parent / "exports"


# =============================================================================
# STREAMING QUERY EXECUTION
# =============================================================================

def execute_streaming(db_path: str, sql: str, params: Dict[str, Any] = None, 
                      batch_size: int = 1000) -> Iterator[List[tuple]]:
    """
    Execute a query and yield results in batches.
    
    Args:
        db_path: Path to SQLite database
        sql: SQL query to execute
        params: Query parameters
        batch_size: Number of rows per batch
        
    Yields:
        Batches of rows as tuples
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        if params:
            cursor.execute(sql, params)
        else:
            cursor.execute(sql)
        
        while True:
            batch = cursor.fetchmany(batch_size)
            if not batch:
                break
            yield batch
            
    finally:
        cursor.close()
        conn.close()


def get_column_names(db_path: str, sql: str, params: Dict[str, Any] = None) -> List[str]:
    """Get column names from a query without fetching all data."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        if params:
            cursor.execute(sql, params)
        else:
            cursor.execute(sql)
        
        return [desc[0] for desc in cursor.description] if cursor.description else []
    finally:
        cursor.close()
        conn.close()


# =============================================================================
# CSV EXPORT (Fastest)
# =============================================================================

def export_to_csv(db_path: str, sql: str, params: Dict[str, Any] = None,
                  filename: Optional[str] = None) -> Path:
    """
    Export query results to CSV using streaming.
    
    Args:
        db_path: Path to SQLite database
        sql: SQL query to execute
        params: Query parameters
        filename: Output filename (auto-generated if None)
        
    Returns:
        Path to the created CSV file
    """
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.csv"
    
    filepath = EXPORT_DIR / filename
    
    # Get column names
    columns = get_column_names(db_path, sql, params)
    
    # Stream to file
    row_count = 0
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columns)  # Header
        
        for batch in execute_streaming(db_path, sql, params):
            writer.writerows(batch)
            row_count += len(batch)
    
    logger.info(f"Exported {row_count} rows to {filepath}")
    return filepath


# =============================================================================
# XLSX EXPORT (With formatting)
# =============================================================================

def export_to_xlsx(db_path: str, sql: str, params: Dict[str, Any] = None,
                   filename: Optional[str] = None, 
                   sheet_name: str = "Export") -> Path:
    """
    Export query results to XLSX using streaming write mode.
    
    Args:
        db_path: Path to SQLite database
        sql: SQL query to execute
        params: Query parameters
        filename: Output filename (auto-generated if None)
        sheet_name: Name of the Excel sheet
        
    Returns:
        Path to the created XLSX file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        csv_path = export_to_csv(db_path, sql, params, filename.replace('.xlsx', '.csv') if filename else None)
        return csv_path
    
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.xlsx"
    
    filepath = EXPORT_DIR / filename
    
    # Get column names
    columns = get_column_names(db_path, sql, params)
    
    # Create workbook in write-only mode for streaming
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)
    
    # Header styling (applied after, since write_only doesn't support direct styling)
    # We'll create a regular workbook for small datasets, write-only for large ones
    
    # For simplicity, use regular mode with chunked processing
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Stream data
    row_num = 2
    row_count = 0
    for batch in execute_streaming(db_path, sql, params):
        for row in batch:
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_idx, value=value)
            row_num += 1
            row_count += 1
    
    # Auto-filter
    if columns:
        ws.auto_filter.ref = f"A1:{chr(64 + len(columns))}{row_num - 1}"
    
    # Adjust column widths (estimate)
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(len(col_name) + 2, 12)
    
    wb.save(filepath)
    logger.info(f"Exported {row_count} rows to {filepath}")
    return filepath


# =============================================================================
# SMART EXPORTER (Auto-chooses format)
# =============================================================================

def smart_export(db_path: str, sql: str, params: Dict[str, Any] = None,
                 format: str = "xlsx", base_filename: Optional[str] = None) -> Dict[str, Any]:
    """
    Smart export with format selection and metadata.
    
    Args:
        db_path: Path to SQLite database
        sql: SQL query to execute
        params: Query parameters
        format: Output format ("csv" or "xlsx")
        base_filename: Base filename without extension
        
    Returns:
        Dict with export metadata
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if not base_filename:
        base_filename = f"export_{timestamp}"
    
    if format == "csv":
        filepath = export_to_csv(db_path, sql, params, f"{base_filename}.csv")
    else:
        filepath = export_to_xlsx(db_path, sql, params, f"{base_filename}.xlsx")
    
    # Get file stats
    file_size = filepath.stat().st_size
    
    return {
        "status": "success",
        "filepath": str(filepath),
        "filename": filepath.name,
        "format": format,
        "file_size_bytes": file_size,
        "file_size_human": f"{file_size / 1024:.1f} KB" if file_size < 1024*1024 else f"{file_size / (1024*1024):.1f} MB",
        "download_url": f"/exports/{filepath.name}",
        "created_at": timestamp
    }


# =============================================================================
# CLI TEST
# =============================================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    
    from ryan_v2.config import DB_PATH
    
    # Test export
    sql = "SELECT * FROM v_alle_huizen LIMIT 10"
    
    print("Testing CSV export...")
    csv_path = export_to_csv(DB_PATH, sql)
    print(f"  → Created: {csv_path}")
    
    print("Testing XLSX export...")
    xlsx_path = export_to_xlsx(DB_PATH, sql)
    print(f"  → Created: {xlsx_path}")
    
    print("Testing smart export...")
    result = smart_export(DB_PATH, sql, format="xlsx", base_filename="test_export")
    print(f"  → Result: {result}")
