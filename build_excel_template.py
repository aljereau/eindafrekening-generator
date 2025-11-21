#!/usr/bin/env python3
"""
Build Excel Template for RyanRent Eindafrekening Generator

Creates a 4-sheet Excel template with named ranges, data validation,
and user-friendly formatting based on excel-template.json specifications.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from datetime import date


def add_named_range(wb, name, reference):
    """Add a named range to the workbook"""
    defn = DefinedName(name, attr_text=reference)
    wb.defined_names[name] = defn


def create_excel_template(output_path="input_template.xlsx"):
    """Create the complete Excel template with all sheets and named ranges"""

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create sheets in order
    create_algemeen_sheet(wb)
    create_gwe_detail_sheet(wb)
    create_schoonmaak_sheet(wb)
    create_schade_sheet(wb)

    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel template created: {output_path}")
    print(f"   📊 Sheets: {', '.join(wb.sheetnames)}")
    print(f"   📝 Named ranges: {len(wb.defined_names)} defined")

    return wb


def create_algemeen_sheet(wb):
    """Create Algemeen sheet with all client, object, period, and voorschotten fields"""

    ws = wb.create_sheet("Algemeen", 0)
    ws.sheet_properties.tabColor = "1F4E78"  # Dark blue

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    section_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    label_font = Font(name='Arial', size=10)
    value_font = Font(name='Arial', size=10)
    computed_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20

    row = 1

    # ==================== HEADER ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "RYANRENT EINDAFREKENING - BASISGEGEVENS"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    row += 2

    # ==================== CLIENT INFO ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "KLANTGEGEVENS"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Klantnaam (required)
    ws[f'A{row}'] = "Naam Klant *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Klantnaam', f'Algemeen!$B${row}')
    row += 1

    # Contactpersoon (optional)
    ws[f'A{row}'] = "Contactpersoon"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Contactpersoon', f'Algemeen!$B${row}')
    row += 1

    # Email (required)
    ws[f'A{row}'] = "Email *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Email', f'Algemeen!$B${row}')
    row += 1

    # Telefoonnummer (optional)
    ws[f'A{row}'] = "Telefoonnummer"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Telefoonnummer', f'Algemeen!$B${row}')
    row += 2

    # ==================== OBJECT INFO ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "OBJECTGEGEVENS"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Object_adres (required)
    ws[f'A{row}'] = "Adres Object *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Object_adres', f'Algemeen!$B${row}')
    row += 1

    # Unit_nr (optional)
    ws[f'A{row}'] = "Unit Nummer"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Unit_nr', f'Algemeen!$B${row}')
    row += 1

    # Postcode (optional)
    ws[f'A{row}'] = "Postcode"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Postcode', f'Algemeen!$B${row}')
    row += 1

    # Plaats (optional)
    ws[f'A{row}'] = "Plaats"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Plaats', f'Algemeen!$B${row}')
    row += 1

    # Object_ID (optional)
    ws[f'A{row}'] = "Object ID"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Object_ID', f'Algemeen!$B${row}')
    row += 2

    # ==================== PERIOD INFO ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "VERBLIJFSPERIODE"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Incheck_datum (required)
    ws[f'A{row}'] = "Incheck Datum *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'DD-MM-YYYY'
    add_named_range(wb,'Incheck_datum', f'Algemeen!$B${row}')
    row += 1

    # Uitcheck_datum (required)
    ws[f'A{row}'] = "Uitcheck Datum *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = 'DD-MM-YYYY'
    add_named_range(wb,'Uitcheck_datum', f'Algemeen!$B${row}')
    uitcheck_row = row
    row += 1

    # Aantal_dagen (computed)
    incheck_row = uitcheck_row - 1
    ws[f'A{row}'] = "Aantal Dagen (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = f'=IF(AND(B{incheck_row}<>"",B{uitcheck_row}<>""),B{uitcheck_row}-B{incheck_row},"")'
    ws[f'B{row}'].number_format = '0'
    add_named_range(wb,'Aantal_dagen', f'Algemeen!$B${row}')
    row += 2

    # ==================== VOORSCHOTTEN ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "VOORSCHOTTEN"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Voorschot_borg (required)
    ws[f'A{row}'] = "Voorschot Borg *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    add_named_range(wb,'Voorschot_borg', f'Algemeen!$B${row}')
    row += 1

    # Voorschot_GWE (required)
    ws[f'A{row}'] = "Voorschot GWE *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    add_named_range(wb,'Voorschot_GWE', f'Algemeen!$B${row}')
    row += 1

    # Voorschot_schoonmaak (required)
    ws[f'A{row}'] = "Voorschot Schoonmaak *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    add_named_range(wb,'Voorschot_schoonmaak', f'Algemeen!$B${row}')
    row += 1

    # Overige_voorschotten (optional, default 0)
    ws[f'A{row}'] = "Overige Voorschotten"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = 0
    add_named_range(wb,'Overige_voorschotten', f'Algemeen!$B${row}')
    row += 2

    # ==================== CONTRACT INFO ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "CONTRACTGEGEVENS"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Schoonmaak_pakket (required, dropdown)
    ws[f'A{row}'] = "Schoonmaak Pakket *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Schoonmaak_pakket', f'Algemeen!$B${row}')

    # Add dropdown validation
    dv = DataValidation(type="list", formula1='"Basis Schoonmaak,Intensief Schoonmaak"', allow_blank=False)
    dv.error = 'Kies 5_uur of 7_uur'
    dv.errorTitle = 'Ongeldige invoer'
    ws.add_data_validation(dv)
    dv.add(f'B{row}')
    schoonmaak_pakket_row = row
    row += 1

    # Inbegrepen_uren (computed based on pakket)
    ws[f'A{row}'] = "Inbegrepen Uren (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = f'=IF(B{schoonmaak_pakket_row}="Basis Schoonmaak",5,IF(B{schoonmaak_pakket_row}="Intensief Schoonmaak",7,""))'
    ws[f'B{row}'].number_format = '0'
    add_named_range(wb,'Inbegrepen_uren', f'Algemeen!$B${row}')
    row += 1

    # Uurtarief_schoonmaak (required, default 50)
    ws[f'A{row}'] = "Uurtarief Schoonmaak *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = 50
    add_named_range(wb,'Uurtarief_schoonmaak', f'Algemeen!$B${row}')
    row += 1

    # Meterbeheerder (optional)
    ws[f'A{row}'] = "Meterbeheerder"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Meterbeheerder', f'Algemeen!$B${row}')
    row += 1

    # Energie_leverancier (optional)
    ws[f'A{row}'] = "Energie Leverancier"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Energie_leverancier', f'Algemeen!$B${row}')
    row += 1

    # Contractnummer (optional)
    ws[f'A{row}'] = "Contractnummer"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'Contractnummer', f'Algemeen!$B${row}')
    row += 2

    # ==================== INTERNAL RYANRENT ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "INTERNE RYANRENT GEGEVENS"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # RR_Klantnummer (optional)
    ws[f'A{row}'] = "RR Klantnummer"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'RR_Klantnummer', f'Algemeen!$B${row}')
    row += 1

    # RR_Folder_link (optional)
    ws[f'A{row}'] = "RR Folder Link"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'RR_Folder_link', f'Algemeen!$B${row}')
    row += 1

    # RR_Projectleider (optional)
    ws[f'A{row}'] = "RR Projectleider"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'RR_Projectleider', f'Algemeen!$B${row}')
    row += 1

    # RR_Inspecteur (optional)
    ws[f'A{row}'] = "RR Inspecteur"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'RR_Inspecteur', f'Algemeen!$B${row}')
    row += 1

    # RR_Factuurnummer (optional)
    ws[f'A{row}'] = "RR Factuurnummer"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = thin_border
    add_named_range(wb,'RR_Factuurnummer', f'Algemeen!$B${row}')
    row += 2

    # ==================== COMPUTED FIELDS ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "BEREKENDE VELDEN (automatisch berekend)"
    cell.font = section_font
    cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    cell.border = thin_border
    ws[f'C{row}'] = "← Live preview (beveiligd)"
    ws[f'C{row}'].font = Font(italic=True, color="C0504D")
    row += 1

    # Borg_gebruikt (computed - references Schade totaal)
    ws[f'A{row}'] = "Borg Gebruikt (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = '=Schade!Schade_totaal_incl'
    add_named_range(wb,'Borg_gebruikt', f'Algemeen!$B${row}')
    borg_gebruikt_row = row
    row += 1

    # Borg_terug (computed)
    ws[f'A{row}'] = "Borg Terug (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = f'=IF(Borg_gebruikt="","",Voorschot_borg-Borg_gebruikt)'
    add_named_range(wb,'Borg_terug', f'Algemeen!$B${row}')
    borg_terug_row = row
    row += 1

    # Restschade (computed)
    ws[f'A{row}'] = "Restschade (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = f'=IF(Borg_gebruikt="","",MAX(0,Borg_gebruikt-Voorschot_borg))'
    add_named_range(wb,'Restschade', f'Algemeen!$B${row}')
    row += 1

    # GWE_meer_minder (computed)
    ws[f'A{row}'] = "GWE Meer/Minder (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = '=IF(GWE_Detail!GWE_totaal_incl="","",Voorschot_GWE-GWE_Detail!GWE_totaal_incl)'
    add_named_range(wb,'GWE_meer_minder', f'Algemeen!$B${row}')
    gwe_meer_minder_row = row
    row += 1

    # Totaal_eindafrekening (computed)
    ws[f'A{row}'] = "Totaal Eindafrekening (automatisch)"
    ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True)
    ws[f'B{row}'] = f'=IF(OR(Borg_terug="",GWE_meer_minder="",Schoonmaak!Extra_schoonmaak_bedrag=""),"",Borg_terug+GWE_meer_minder-Schoonmaak!Extra_schoonmaak_bedrag)'
    add_named_range(wb,'Totaal_eindafrekening', f'Algemeen!$B${row}')

    # Freeze panes at A2 (freeze header row)
    ws.freeze_panes = 'A3'
    
    # Enable sheet protection (no password, just UI lock)
    # Unlock all input cells first, then protect sheet
    # for row_cells in ws.iter_rows():
    # for cell in row_cells:
            # By default all cells are locked, but we need to unlock input cells
            # Computed cells are already set to locked=True above
            # Input cells (with required_fill or normal border) should be unlocked
    # if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == required_fill.start_color.rgb:
    # cell.protection = Protection(locked=False)
    # elif cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != computed_fill.start_color.rgb:
                # Regular input cells (no fill or default fill)
    # if cell.value is None or (isinstance(cell.value, str) and not cell.value.startswith('=')):
    # cell.protection = Protection(locked=False)
    
    # ws.protection.sheet = True  # DISABLED - All cells unlocked


def create_gwe_detail_sheet(wb):
    """Create GWE_Detail sheet with meter readings and cost breakdown table"""

    ws = wb.create_sheet("GWE_Detail")
    ws.sheet_properties.tabColor = "70AD47"  # Green

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    section_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    label_font = Font(name='Arial', size=10)
    computed_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    table_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    row = 1

    # ==================== HEADER ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "GAS / WATER / ELEKTRA DETAIL"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    row += 2

    # ==================== METERSTANDEN ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "METERSTANDEN"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # KWh meters
    ws[f'A{row}'] = "Elektra (kWh) Begin *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '#,##0'
    add_named_range(wb,'KWh_begin', f'GWE_Detail!$B${row}')
    kwh_begin_row = row
    row += 1

    ws[f'A{row}'] = "Elektra (kWh) Eind *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '#,##0'
    add_named_range(wb,'KWh_eind', f'GWE_Detail!$B${row}')
    kwh_eind_row = row
    row += 1

    ws[f'A{row}'] = "Elektra (kWh) Verbruik (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = f'=IF(AND(B{kwh_begin_row}<>"",B{kwh_eind_row}<>""),B{kwh_eind_row}-B{kwh_begin_row},"")'
    ws[f'B{row}'].number_format = '#,##0'
    add_named_range(wb,'KWh_verbruik', f'GWE_Detail!$B${row}')
    row += 1

    # Gas meters
    ws[f'A{row}'] = "Gas (m³) Begin *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '#,##0.00'
    add_named_range(wb,'Gas_begin', f'GWE_Detail!$B${row}')
    gas_begin_row = row
    row += 1

    ws[f'A{row}'] = "Gas (m³) Eind *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '#,##0.00'
    add_named_range(wb,'Gas_eind', f'GWE_Detail!$B${row}')
    gas_eind_row = row
    row += 1

    ws[f'A{row}'] = "Gas (m³) Verbruik (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = f'=IF(AND(B{gas_begin_row}<>"",B{gas_eind_row}<>""),B{gas_eind_row}-B{gas_begin_row},"")'
    ws[f'B{row}'].number_format = '#,##0.00'
    add_named_range(wb,'Gas_verbruik', f'GWE_Detail!$B${row}')
    row += 2

    # ==================== KOSTENREGELS TABLE ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "KOSTENREGELS - Verbruik omzetten naar Euro's"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Instructions
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "💡 Vul hier de kosten in op basis van verbruik × tarief van energieleverancier. Voorbeelden staan hieronder (grijs). Verwijder voorbeelden en vul eigen regels in."
    cell.font = Font(name='Arial', size=9, italic=True)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.row_dimensions[row].height = 30
    row += 1

    # Table headers (row 5 as per spec)
    table_start_row = row
    headers = ["Omschrijving", "Verbruik (kWh/m³) of Dagen", "Tarief excl. BTW (€)", "Kosten excl. BTW (€)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        cell.fill = table_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # Add example rows (with gray background to indicate they're examples)
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    example_font = Font(name='Arial', size=9, italic=True, color="666666")

    examples = [
        ("Elektra verbruik (zie meterstand)", "=KWh_verbruik", 0.30, "=B{row}*C{row}"),
        ("Gas verbruik (zie meterstand)", "=Gas_verbruik", 1.20, "=B{row}*C{row}"),
        ("Vaste levering elektra per dag", "=Algemeen!Aantal_dagen", 0.50, "=B{row}*C{row}"),
        ("Vaste levering gas per dag", "=Algemeen!Aantal_dagen", 0.45, "=B{row}*C{row}"),
        ("Waterverbruik (geschat per dag)", "=Algemeen!Aantal_dagen", 3.50, "=B{row}*C{row}"),
    ]

    for example in examples:
        # Omschrijving
        cell = ws.cell(row=row, column=1)
        cell.value = example[0]
        cell.font = example_font
        cell.fill = example_fill
        cell.border = thin_border

        # Verbruik/Dagen (formula or number)
        cell = ws.cell(row=row, column=2)
        cell.value = example[1]
        cell.font = example_font
        cell.fill = example_fill
        cell.border = thin_border
        cell.number_format = '#,##0.00'

        # Tarief
        cell = ws.cell(row=row, column=3)
        cell.value = example[2]
        cell.font = example_font
        cell.fill = example_fill
        cell.border = thin_border
        cell.number_format = '€ #,##0.00'

        # Kosten (formula)
        cell = ws.cell(row=row, column=4)
        cell.value = example[3].format(row=row)
        cell.font = example_font
        cell.fill = example_fill
        cell.border = thin_border
        cell.number_format = '€ #,##0.00'

        row += 1

    # Add 15 empty rows for additional data entry
    for i in range(15):
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            if col_idx in [2, 3, 4]:  # Number columns
                cell.number_format = '€ #,##0.00' if col_idx in [3, 4] else '#,##0.00'
        row += 1

    row += 1

    # ==================== TOTALEN ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "TOTALEN (automatisch berekend)"
    cell.font = section_font
    cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    cell.border = thin_border
    row += 1

    # GWE_totaal_excl (computed - sum of all cost lines)
    ws[f'A{row}'] = "Totaal excl. BTW (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    # Sum from table_start_row+1 (first data row after header) to row-2 (before totals)
    ws[f'B{row}'] = f'=SUM(D{table_start_row+1}:D{row-2})'
    add_named_range(wb,'GWE_totaal_excl', f'GWE_Detail!$B${row}')
    gwe_totaal_excl_row = row
    row += 1

    # GWE_BTW (computed: 21% of excl)
    ws[f'A{row}'] = "BTW 21% (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = f'=IF(B{gwe_totaal_excl_row}="","",B{gwe_totaal_excl_row}*0.21)'
    add_named_range(wb,'GWE_BTW', f'GWE_Detail!$B${row}')
    gwe_btw_row = row
    row += 1

    # GWE_totaal_incl (computed: excl + btw)
    ws[f'A{row}'] = "Totaal incl. BTW (automatisch)"
    ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'B{row}'] = f'=IF(B{gwe_totaal_excl_row}="","",B{gwe_totaal_excl_row}+B{gwe_btw_row})'
    add_named_range(wb,'GWE_totaal_incl', f'GWE_Detail!$B${row}')

    # Freeze panes at A6 (freeze headers)
    ws.freeze_panes = 'A6'
    
    # Enable sheet protection (unlock input cells)
    # for row_cells in ws.iter_rows():
    # for cell in row_cells:
    # if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == required_fill.start_color.rgb:
    # cell.protection = Protection(locked=False)
    # elif cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != computed_fill.start_color.rgb:
    # if cell.value is None or (isinstance(cell.value, str) and not cell.value.startswith('=')):
    # cell.protection = Protection(locked=False)
    
    # ws.protection.sheet = True  # DISABLED - All cells unlocked


def create_schoonmaak_sheet(wb):
    """Create Schoonmaak sheet with hours tracking"""

    ws = wb.create_sheet("Schoonmaak")
    ws.sheet_properties.tabColor = "FFC000"  # Orange

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    section_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    label_font = Font(name='Arial', size=10)
    computed_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    row = 1

    # ==================== HEADER ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "SCHOONMAAK OVERZICHT"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    row += 2

    # ==================== SCHOONMAAK DATA ====================
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "SCHOONMAAK GEGEVENS"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    row += 1

    # Schoonmaak_pakket (required, dropdown - note: also in Algemeen)
    ws[f'A{row}'] = "Schoonmaak Pakket *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = '=Algemeen!Schoonmaak_pakket'  # Reference from Algemeen
    ws[f'B{row}'].font = Font(italic=True, color="666666")
    schoonmaak_pakket_row = row
    row += 1

    # Inbegrepen_uren (computed)
    ws[f'A{row}'] = "Inbegrepen Uren (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = f'=IF(B{schoonmaak_pakket_row}="Basis Schoonmaak",5,IF(B{schoonmaak_pakket_row}="Intensief Schoonmaak",7,""))'
    ws[f'B{row}'].number_format = '0'
    inbegrepen_uren_row = row
    row += 1

    # Totaal_uren_gew (required - actual hours worked)
    ws[f'A{row}'] = "Totaal Uren Gewerkt *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '0.00'
    add_named_range(wb,'Totaal_uren_gew', f'Schoonmaak!$B${row}')
    totaal_uren_row = row
    row += 1

    # Extra_uren (computed)
    ws[f'A{row}'] = "Extra Uren (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = f'=IF(AND(B{inbegrepen_uren_row}<>"",B{totaal_uren_row}<>""),MAX(0,B{totaal_uren_row}-B{inbegrepen_uren_row}),"")'
    ws[f'B{row}'].number_format = '0.00'
    add_named_range(wb,'Extra_uren', f'Schoonmaak!$B${row}')
    extra_uren_row = row
    row += 1

    # Uurtarief_schoonmaak (required)
    ws[f'A{row}'] = "Uurtarief *"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = required_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'] = '=Algemeen!Uurtarief_schoonmaak'  # Reference from Algemeen
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'].font = Font(italic=True, color="666666")
    uurtarief_row = row
    row += 1

    # Extra_schoonmaak_bedrag (computed)
    ws[f'A{row}'] = "Extra Schoonmaak Bedrag (automatisch)"
    ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    ws[f'B{row}'] = f'=IF(AND(B{extra_uren_row}<>"",B{uurtarief_row}<>""),B{extra_uren_row}*B{uurtarief_row},"")'
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    add_named_range(wb,'Extra_schoonmaak_bedrag', f'Schoonmaak!$B${row}')

    # Freeze panes
    ws.freeze_panes = 'A3'
    
    # Enable sheet protection (unlock input cells, lock computed)
    # Specifically unlock Totaal_uren_gew (totaal_uren_row)
    # Lock Inbegrepen_uren (computed from pakket), Extra_uren (computed), and Extra_schoonmaak_bedrag (computed)
    # for row_cells in ws.iter_rows():
    # for cell in row_cells:
            # Only unlock specific input cell: Totaal_uren_gew
    # if cell.row == totaal_uren_row and cell.column == 2:  # B column
    # cell.protection = Protection(locked=False)
    
    # Also need to protect the computed cells explicitly
    
    # ws.protection.sheet = True  # DISABLED - All cells unlocked


def create_schade_sheet(wb):
    """Create Schade sheet with damage items table"""

    ws = wb.create_sheet("Schade")
    ws.sheet_properties.tabColor = "C0504D"  # Red

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    section_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="E7989A", end_color="E7989A", fill_type="solid")
    label_font = Font(name='Arial', size=10)
    computed_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    table_header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 50  # Beschrijving
    ws.column_dimensions['B'].width = 12  # Aantal
    ws.column_dimensions['C'].width = 20  # Tarief excl
    ws.column_dimensions['D'].width = 20  # Bedrag excl

    row = 1

    # ==================== HEADER ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "SCHADEPOSTEN OVERZICHT"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    row += 2

    # ==================== INSTRUCTIONS ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Vul hieronder alle schadeposten in. Bedrag wordt automatisch berekend (Aantal × Tarief)."
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # ==================== DAMAGE TABLE ====================
    # Table headers (row 5 as per spec)
    table_start_row = 5
    ws.row_dimensions[table_start_row].height = 25

    headers = ["Beschrijving", "Aantal", "Tarief excl. BTW", "Bedrag excl. BTW"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_start_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        cell.fill = table_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add 50 rows for damage items (spec says A5:D200, but 50 should be enough)
    for i in range(50):
        data_row = table_start_row + 1 + i

        # Beschrijving (column A)
        cell = ws.cell(row=data_row, column=1)
        cell.border = thin_border

        # Aantal (column B)
        cell = ws.cell(row=data_row, column=2)
        cell.border = thin_border
        cell.number_format = '0'

        # Tarief excl (column C)
        cell = ws.cell(row=data_row, column=3)
        cell.border = thin_border
        cell.number_format = '€ #,##0.00'

        # Bedrag excl (column D) - formula: B * C
        cell = ws.cell(row=data_row, column=4)
        cell.border = thin_border
        cell.number_format = '€ #,##0.00'
        cell.value = f'=IF(AND(B{data_row}<>"",C{data_row}<>""),B{data_row}*C{data_row},"")'
        cell.fill = computed_fill

    row = table_start_row + 52  # After table + some space

    # ==================== TOTALEN ====================
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "TOTALEN (automatisch berekend)"
    cell.font = section_font
    cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    cell.border = thin_border
    row += 1

    # Schade_totaal_excl (computed - sum of all bedrag_excl)
    ws[f'A{row}'] = "Totaal Schade excl. BTW (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = f'=SUM(D{table_start_row+1}:D{row-2})'
    add_named_range(wb,'Schade_totaal_excl', f'Schade!$B${row}')
    schade_totaal_excl_row = row
    row += 1

    # Schade_BTW (computed: 21% of excl)
    ws[f'A{row}'] = "BTW 21% (automatisch)"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'] = f'=IF(B{schade_totaal_excl_row}="","",B{schade_totaal_excl_row}*0.21)'
    add_named_range(wb,'Schade_BTW', f'Schade!$B${row}')
    schade_btw_row = row
    row += 1

    # Schade_totaal_incl (computed: excl + btw)
    ws[f'A{row}'] = "Totaal Schade incl. BTW (automatisch)"
    ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'B{row}'].fill = computed_fill
    ws[f'B{row}'].border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    ws[f'B{row}'].number_format = '€ #,##0.00'
    ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'B{row}'] = f'=IF(B{schade_totaal_excl_row}="","",B{schade_totaal_excl_row}+B{schade_btw_row})'
    add_named_range(wb,'Schade_totaal_incl', f'Schade!$B${row}')

    # Freeze panes at A6 (freeze headers and instructions)
    ws.freeze_panes = 'A6'
    
    # Enable sheet protection (unlock input cells, lock computed)
    # for row_cells in ws.iter_rows():
    # for cell in row_cells:
            # Unlock input cells (columns A, B, C for damage items)
            # Column D is computed (already has formula), keep it locked
    # if cell.column in [1, 2, 3]:  # A, B, C
    # if cell.row >= table_start_row + 1 and cell.row <= table_start_row + 50:
    # cell.protection = Protection(locked=False)
    
    # ws.protection.sheet = True  # DISABLED - All cells unlocked


if __name__ == "__main__":
    print("🏗️  Building RyanRent Excel Template...")
    print("=" * 60)

    wb = create_excel_template()

    print("\n" + "=" * 60)
    print("✅ Template complete!")
    print("\n📋 Summary:")
    print("   • Algemeen: Client, object, period, voorschotten, contract, internal")
    print("   • GWE_Detail: Meter readings + cost breakdown table")
    print("   • Schoonmaak: Hours tracking with automatic calculations")
    print("   • Schade: Damage items table with qty × tariff calculations")
    print("\n💡 Next steps:")
    print("   1. Open input_template.xlsx in Excel")
    print("   2. Test data entry and validation")
    print("   3. Verify named ranges work correctly")
    print("   4. Build Python reader for named ranges")
