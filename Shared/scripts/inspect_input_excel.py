import openpyxl
import sys

file_path = '/Users/aljereaumarten/Library/CloudStorage/OneDrive-SharedLibraries-RyanRentB.V/RyanRent - General/01_RyanRent&Co/Aljereau/Eindafrekening Generator/Eindafrekening/src/input_template_0179-04122025.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Check named range 'Schoonmaak_totaal_kosten'
    if 'Schoonmaak_totaal_kosten' in wb.defined_names:
        dests = wb.defined_names['Schoonmaak_totaal_kosten'].destinations
        for sheet_name, coord in dests:
            ws = wb[sheet_name]
            val = ws[coord.replace('$', '')].value
            print(f"Schoonmaak_totaal_kosten ({sheet_name}!{coord}): {val}")
    else:
        print("Named range 'Schoonmaak_totaal_kosten' NOT found.")

    # Also check 'Totaal_uren_gew' in Schoonmaak sheet
    if 'Totaal_uren_gew' in wb.defined_names:
        dests = wb.defined_names['Totaal_uren_gew'].destinations
        for sheet_name, coord in dests:
            ws = wb[sheet_name]
            val = ws[coord.replace('$', '')].value
            print(f"Totaal_uren_gew ({sheet_name}!{coord}): {val}")
    else:
        print("Named range 'Totaal_uren_gew' NOT found.")

except Exception as e:
    print(f"Error: {e}")
