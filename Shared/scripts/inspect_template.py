import openpyxl
import sys
import os

# Path to template
template_path = "Shared/Templates/RyanRent_Input_Template.xlsx"
if not os.path.exists(template_path):
    template_path = "Eindafrekening/src/input_template.xlsx"

print(f"Analyzing: {template_path}")

wb = openpyxl.load_workbook(template_path)
ws = wb['Algemeen']

print("\n--- Named Ranges ---")
for name, defn in wb.defined_names.items():
    # Filter for relevant names
    if any(x in name for x in ['Adres', 'Postcode', 'Plaats', 'Voorschot', 'GWE']):
        dests = list(defn.destinations)
        for sheet, cell in dests:
            if sheet == 'Algemeen':
                print(f"{name}: {cell} (Value: {ws[cell.replace('$', '')].value})")

print("\n--- Cell Values (First 20 rows) ---")
for row in ws.iter_rows(min_row=1, max_row=20, max_col=5):
    row_data = []
    for cell in row:
        val = str(cell.value).strip() if cell.value else ""
        if val:
            row_data.append(f"{cell.coordinate}:{val}")
    if row_data:
        print(" | ".join(row_data))
