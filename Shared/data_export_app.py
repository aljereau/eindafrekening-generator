#!/usr/bin/env python3
"""
RyanRent Data Export Tool
Simple web app to export database tables/views to Dutch-formatted XLSX
"""

from flask import Flask, render_template_string, send_file, request, jsonify
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from pathlib import Path

app = Flask(__name__)

# Database path
DB_PATH = Path(__file__).parent.parent / "database" / "ryanrent_v2.db"
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


def get_tables_and_views():
    """Get all tables and views from the database"""
    conn = sqlite3.connect(DB_PATH)
    
    # Get tables
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' AND name NOT LIKE 'schema_%' ORDER BY name"
    ).fetchall()
    
    # Get views
    views = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='view' ORDER BY name"
    ).fetchall()
    
    conn.close()
    
    return {
        'tables': [t[0] for t in tables],
        'views': [v[0] for v in views]
    }


def to_dutch_number(value):
    """Convert number to Dutch format (comma decimal, dot thousands)"""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            formatted = f"{value:,.2f}"
        else:
            formatted = f"{value:,}"
        # Swap separators
        return formatted.replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")
    return str(value)


def is_id_column(column_name):
    """Check if column is an ID/code column that shouldn't be formatted as number"""
    name = column_name.lower()
    patterns = ['_id', 'object_id', 'house_id', 'klant_id', 'relatie_id', 
                'contract_id', 'boeking_id', 'postcode', 'nr', 'nummer', 'code']
    return any(p in name or name.endswith(p) for p in patterns)


def export_to_xlsx(table_name, query=None):
    """Export table/view to formatted XLSX"""
    conn = sqlite3.connect(DB_PATH)
    
    if query:
        cursor = conn.execute(query)
    else:
        cursor = conn.execute(f"SELECT * FROM {table_name}")
    
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    conn.close()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:30]  # Sheet name max 31 chars
    
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (col_name, value) in enumerate(zip(columns, row), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Format value
            if value is None:
                cell.value = ""
            elif isinstance(value, (int, float)) and not is_id_column(col_name):
                cell.value = value
                if isinstance(value, float):
                    cell.number_format = '#.##0,00'  # Dutch number format
                else:
                    cell.number_format = '#.##0'
            else:
                cell.value = str(value) if value else ""
            
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alt_fill
    
    # Auto-fit columns
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(col_name)
        for row in ws.iter_rows(min_row=2, max_row=min(100, len(rows)+1), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(50, max_length + 2)
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{table_name}_{timestamp}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    
    return filepath


# HTML Template
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="nl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>RyanRent Data Export</title>
    <style>
        :root {
            --primary: #1F4E79;
            --primary-light: #2E6B9E;
            --bg: #F8FAFC;
            --surface: #FFFFFF;
            --text: #1E293B;
            --text-muted: #64748B;
            --success: #059669;
            --border: #E2E8F0;
            --selected-bg: #DCFCE7;
            --selected-border: #059669;
        }
        
        * { box-sizing: border-box; margin: 0; padding: 0; }
        
        body {
            font-family: 'Segoe UI', -apple-system, sans-serif;
            background: var(--bg);
            color: var(--text);
            min-height: 100vh;
            padding: 2rem;
        }
        
        .container {
            max-width: 1000px;
            margin: 0 auto;
        }
        
        h1 {
            font-size: 1.5rem;
            margin-bottom: 0.25rem;
            color: var(--text);
            font-weight: 600;
        }
        
        .subtitle {
            color: var(--text-muted);
            margin-bottom: 1.5rem;
            font-size: 0.9rem;
        }
        
        .section {
            background: var(--surface);
            border-radius: 8px;
            padding: 1.25rem;
            margin-bottom: 1rem;
            border: 1px solid var(--border);
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }
        
        .section h2 {
            font-size: 0.8rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-muted);
            margin-bottom: 0.75rem;
            font-weight: 600;
        }
        
        .items {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(180px, 1fr));
            gap: 0.5rem;
        }
        
        .item {
            background: var(--bg);
            border: 1px solid var(--border);
            border-radius: 6px;
            padding: 0.5rem 0.75rem;
            cursor: pointer;
            transition: all 0.1s ease;
            display: flex;
            align-items: center;
            gap: 0.5rem;
            font-size: 0.85rem;
        }
        
        .item:hover {
            border-color: var(--primary);
            background: #F1F5F9;
        }
        
        .item.selected {
            border-color: var(--selected-border);
            background: var(--selected-bg);
        }
        
        .item input[type="checkbox"] {
            width: 16px;
            height: 16px;
            accent-color: var(--success);
            cursor: pointer;
        }
        
        .btn {
            background: var(--primary);
            color: white;
            border: none;
            padding: 0.75rem 1.5rem;
            border-radius: 6px;
            font-size: 0.9rem;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.15s ease;
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
        }
        
        .btn:hover { background: var(--primary-light); }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        
        .status {
            margin-top: 1rem;
            padding: 0.75rem 1rem;
            border-radius: 6px;
            display: none;
            font-size: 0.9rem;
        }
        
        .status.success {
            display: block;
            background: var(--selected-bg);
            border: 1px solid var(--success);
            color: var(--success);
        }
        
        .selection-count {
            margin-left: 1rem;
            font-size: 0.85rem;
            color: var(--text-muted);
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 RyanRent Data Export</h1>
        <p class="subtitle">Selecteer tabellen of views om te exporteren naar Excel</p>
        
        <form id="exportForm" action="/export" method="POST">
            <div class="section">
                <h2>Views ({{ views|length }})</h2>
                <div class="items">
                    {% for view in views %}
                    <label class="item">
                        <input type="checkbox" name="items" value="{{ view }}" onchange="updateSelection(this)">
                        <span>{{ view }}</span>
                    </label>
                    {% endfor %}
                </div>
            </div>
            
            <div class="section">
                <h2>Tables ({{ tables|length }})</h2>
                <div class="items">
                    {% for table in tables %}
                    <label class="item">
                        <input type="checkbox" name="items" value="{{ table }}" onchange="updateSelection(this)">
                        <span>{{ table }}</span>
                    </label>
                    {% endfor %}
                </div>
            </div>
            
            <button type="submit" class="btn">
                📥 Export naar Excel
            </button>
            <span class="selection-count" id="selectionCount"></span>
        </form>
        
        <div id="status" class="status"></div>
    </div>
    
    <script>
        function updateSelection(checkbox) {
            const item = checkbox.closest('.item');
            if (checkbox.checked) {
                item.classList.add('selected');
            } else {
                item.classList.remove('selected');
            }
            updateCount();
        }
        
        function updateCount() {
            const count = document.querySelectorAll('input[name="items"]:checked').length;
            document.getElementById('selectionCount').textContent = count > 0 ? count + ' geselecteerd' : '';
        }
        
        document.getElementById('exportForm').addEventListener('submit', async (e) => {
            e.preventDefault();
            const checked = document.querySelectorAll('input[name="items"]:checked');
            const items = Array.from(checked).map(el => el.value);
            
            if (items.length === 0) {
                alert('Selecteer minimaal 1 tabel of view');
                return;
            }
            
            const btn = e.target.querySelector('button');
            btn.disabled = true;
            btn.textContent = 'Bezig met exporteren...';
            
            try {
                const response = await fetch('/export', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({items})
                });
                
                if (response.ok) {
                    const blob = await response.blob();
                    const url = URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    const filename = response.headers.get('Content-Disposition')?.split('filename=')[1] || 'export.xlsx';
                    a.href = url;
                    a.download = filename.replace(/"/g, '');
                    a.click();
                    
                    document.getElementById('status').className = 'status success';
                    document.getElementById('status').innerHTML = '✅ Export succesvol!';
                }
            } catch (err) {
                alert('Export mislukt: ' + err.message);
            }
            
            btn.disabled = false;
            btn.textContent = '📥 Export naar Excel';
        });
    </script>
</body>
</html>
"""



@app.route('/')
def index():
    data = get_tables_and_views()
    return render_template_string(HTML_TEMPLATE, **data)


@app.route('/export', methods=['POST'])
def export():
    data = request.get_json()
    items = data.get('items', [])
    
    if not items:
        return jsonify({'error': 'No items selected'}), 400
    
    if len(items) == 1:
        # Single export
        filepath = export_to_xlsx(items[0])
        return send_file(filepath, as_attachment=True, download_name=filepath.name)
    else:
        # Multiple exports - create zip or multi-sheet workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        
        # Create multi-sheet workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        conn = sqlite3.connect(DB_PATH)
        
        for item in items:
            cursor = conn.execute(f"SELECT * FROM {item}")
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()
            
            ws = wb.create_sheet(title=item[:30])
            
            # Styles
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Headers
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
            
            ws.freeze_panes = 'A2'
            
            # Data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, (col_name, value) in enumerate(zip(columns, row), 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if value is None:
                        cell.value = ""
                    elif isinstance(value, float) and not is_id_column(col_name):
                        cell.value = value
                        cell.number_format = '#.##0,00'
                    else:
                        cell.value = str(value) if value else ""
        
        conn.close()
        
        filename = f"ryanrent_export_{timestamp}.xlsx"
        filepath = OUTPUT_DIR / filename
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    print(f"📊 RyanRent Data Export Tool")
    print(f"   Database: {DB_PATH}")
    print(f"   Output: {OUTPUT_DIR}")
    print(f"   Open: http://127.0.0.1:5055")
    app.run(host='127.0.0.1', port=5055, debug=True)
