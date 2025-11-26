#!/usr/bin/env python3
"""
Excel Database Module - Team-viewable summary workbook

Creates and maintains an Excel workbook with summary of all eindafrekeningen
for easy viewing by non-technical team members.
"""

import os
from datetime import date, datetime
from typing import Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelDatabase:
    """Manages Excel workbook for eindafrekening summaries"""

    DEFAULT_PATH = "database/eindafrekeningen_database.xlsx"
    SHEET_NAME = "Database"

    # Column definitions
    COLUMNS = [
        ("ID", 8),
        ("Client", 25),
        ("Check-in", 12),
        ("Check-out", 12),
        ("Dagen", 8),
        ("Versie", 8),
        ("Borg Terug", 12),
        ("GWE Totaal", 12),
        ("Netto", 12),
        ("Reden", 30),
        ("Datum", 18),
        ("Bestand", 50)
    ]

    def __init__(self, excel_path: str = None):
        """
        Initialize Excel database

        Args:
            excel_path: Path to Excel workbook (default: database/eindafrekeningen_database.xlsx)
        """
        self.excel_path = excel_path or self.DEFAULT_PATH

        # Ensure database directory exists
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)

        # Initialize workbook if it doesn't exist
        if not os.path.exists(self.excel_path):
            self._create_workbook()

    def _create_workbook(self):
        """Create new Excel workbook with formatted headers"""
        print(f"📊 Creating Excel database: {self.excel_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = self.SHEET_NAME

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, (col_name, col_width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

            # Set column width
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = col_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save workbook
        wb.save(self.excel_path)
        print(f"   ✓ Excel database created")

    def append_record(
        self,
        record_id: int,
        client_name: str,
        checkin_date: date,
        checkout_date: date,
        period_days: int,
        version: int,
        borg_terug: float,
        gwe_totaal_incl: float,
        totaal_eindafrekening: float,
        version_reason: str,
        file_path: str = ""
    ):
        """
        Append new record to Excel database

        Args:
            record_id: Database ID from SQLite
            client_name: Client name
            checkin_date: Check-in date
            checkout_date: Check-out date
            period_days: Number of days
            version: Version number
            borg_terug: Deposit refund amount
            gwe_totaal_incl: GWE total including VAT
            totaal_eindafrekening: Final settlement (positive = refund, negative = charge)
            version_reason: Reason for this version
            file_path: Path to generated file
        """
        # Load workbook
        wb = load_workbook(self.excel_path)
        ws = wb[self.SHEET_NAME]

        # Find next empty row
        next_row = ws.max_row + 1

        # Cell styles
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Prepare data
        row_data = [
            record_id,
            client_name,
            checkin_date.strftime('%d-%m-%Y') if isinstance(checkin_date, date) else checkin_date,
            checkout_date.strftime('%d-%m-%Y') if isinstance(checkout_date, date) else checkout_date,
            period_days,
            f"v{version}",
            borg_terug,
            gwe_totaal_incl,
            totaal_eindafrekening,
            version_reason,
            datetime.now().strftime('%d-%m-%Y %H:%M'),
            os.path.basename(file_path) if file_path else ""
        ]

        # Write row
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx)
            cell.value = value
            cell.border = border

            # Format currency columns
            if col_idx in [7, 8, 9]:  # Borg Terug, GWE Totaal, Netto
                cell.number_format = '€#,##0.00'

            # Highlight revisions (version > 1)
            if version > 1:
                cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

        # Auto-filter
        ws.auto_filter.ref = f"A1:L{next_row}"

        # Save workbook
        wb.save(self.excel_path)
        print(f"   ✓ Appended to Excel database: Row {next_row}")

    def get_record_count(self) -> int:
        """Get total number of records in Excel database"""
        if not os.path.exists(self.excel_path):
            return 0

        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb[self.SHEET_NAME]
        count = ws.max_row - 1  # Subtract header row
        wb.close()
        return count


# Convenience function
def append_to_excel_database(
    record_id: int,
    client_name: str,
    checkin_date: date,
    checkout_date: date,
    period_days: int,
    version: int,
    borg_terug: float,
    gwe_totaal_incl: float,
    totaal_eindafrekening: float,
    version_reason: str,
    file_path: str = "",
    excel_path: str = None
):
    """
    Append record to Excel database (convenience function)

    Args:
        Same as ExcelDatabase.append_record()
        excel_path: Optional custom Excel database path
    """
    excel_db = ExcelDatabase(excel_path)
    excel_db.append_record(
        record_id=record_id,
        client_name=client_name,
        checkin_date=checkin_date,
        checkout_date=checkout_date,
        period_days=period_days,
        version=version,
        borg_terug=borg_terug,
        gwe_totaal_incl=gwe_totaal_incl,
        totaal_eindafrekening=totaal_eindafrekening,
        version_reason=version_reason,
        file_path=file_path
    )


if __name__ == "__main__":
    """Test Excel database operations"""
    import sys
    from datetime import date

    print("🧪 Testing Excel Database Module")
    print("=" * 60)

    # Use test Excel database
    test_excel_path = "database/test_eindafrekeningen_database.xlsx"

    # Clean up test file if exists
    if os.path.exists(test_excel_path):
        os.remove(test_excel_path)
        print(f"   Cleaned up existing test Excel database")

    # Initialize Excel database
    excel_db = ExcelDatabase(test_excel_path)
    print(f"\n✓ Excel database initialized: {test_excel_path}")

    # Test 1: Append record (v1)
    print("\n1. Testing append_record (v1)...")
    excel_db.append_record(
        record_id=1,
        client_name="Familie Jansen",
        checkin_date=date(2024, 7, 1),
        checkout_date=date(2024, 7, 15),
        period_days=14,
        version=1,
        borg_terug=200.00,
        gwe_totaal_incl=308.55,
        totaal_eindafrekening=150.45,
        version_reason="Eerste verzending",
        file_path="output/familie_jansen_onepager.pdf"
    )
    print(f"   ✓ Record v1 appended")

    # Test 2: Append revision (v2)
    print("\n2. Testing append_record (v2 - revision)...")
    excel_db.append_record(
        record_id=2,
        client_name="Familie Jansen",
        checkin_date=date(2024, 7, 1),
        checkout_date=date(2024, 7, 15),
        period_days=14,
        version=2,
        borg_terug=200.00,
        gwe_totaal_incl=350.00,
        totaal_eindafrekening=108.55,
        version_reason="Correctie GWE - extra elektra verbruik",
        file_path="output/familie_jansen_onepager_v2.pdf"
    )
    print(f"   ✓ Record v2 appended (should be highlighted)")

    # Test 3: Get record count
    print("\n3. Testing get_record_count...")
    count = excel_db.get_record_count()
    print(f"   ✓ Total records: {count}")
    if count != 2:
        print(f"   ❌ Expected 2 records, got {count}")
        sys.exit(1)

    # Test 4: Another client
    print("\n4. Testing append_record (different client)...")
    excel_db.append_record(
        record_id=3,
        client_name="Familie De Vries",
        checkin_date=date(2024, 8, 1),
        checkout_date=date(2024, 8, 8),
        period_days=7,
        version=1,
        borg_terug=150.00,
        gwe_totaal_incl=180.00,
        totaal_eindafrekening=75.50,
        version_reason="Eerste verzending",
        file_path="output/familie_de_vries_onepager.pdf"
    )
    print(f"   ✓ Record appended")

    print("\n✅ All Excel database tests passed!")
    print(f"\n📍 Test Excel database location: {os.path.abspath(test_excel_path)}")
    print(f"   Open in Excel to view formatted results")
