import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from typing import List, Dict, Any, Optional, Tuple
import datetime

class ExcelExporter:
    """
    Generic tool to export data to Excel with standard formatting.
    """
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], output_path: str, sheet_name: str = "Sheet1"):
        """
        Export a list of dictionaries to an Excel file.
        Keys of the first dictionary are used as headers.
        """
        if not data:
            print("⚠️ No data to export.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = length + 2

        wb.save(output_path)
        print(f"✅ Exported {len(data)} rows to: {output_path}")


class ExcelImporter:
    """
    Generic tool to read data from Excel.
    """

    @staticmethod
    def read_from_excel(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        Read an Excel file and return a list of dictionaries.
        Assumes the first row contains headers.
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            data = []
            headers = []

            # Read headers
            for cell in ws[1]:
                headers.append(cell.value)

            if not headers:
                return []

            # Read rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not any(row):
                    continue
                
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        header = headers[idx]
                        if header: # Only map if header exists
                            row_dict[header] = value
                data.append(row_dict)

            print(f"✅ Read {len(data)} rows from: {file_path}")
            return data

        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            raise
