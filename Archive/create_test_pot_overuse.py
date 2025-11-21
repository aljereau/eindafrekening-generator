#!/usr/bin/env python3
"""
Create test_pot_overuse.xlsx - A test scenario with budget overages
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import shutil
from pathlib import Path
from datetime import datetime, timedelta

def safe_set_cell(ws, cell_ref, value):
    """Safely set a cell value, handling merged cells."""
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        # Find the merged range and set value on the top-left cell
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                # Get top-left cell of merged range
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left = ws.cell(min_row, min_col)
                top_left.value = value
                return
    else:
        ws[cell_ref].value = value

def create_pot_overuse_test():
    """Create a test template with pot overuse (budget overages)."""
    
    # Copy clean template
    shutil.copy('input_template.xlsx', 'test_pot_overuse.xlsx')
    print("✓ Created test_pot_overuse.xlsx from clean template")
    
    # Load workbook
    wb = openpyxl.load_workbook('test_pot_overuse.xlsx')
    
    # Fill Algemeen sheet
    ws_alg = wb['Algemeen']
    safe_set_cell(ws_alg, 'B4', 'Familie de Jong')
    safe_set_cell(ws_alg, 'B5', 'Mark van RyanRent')
    safe_set_cell(ws_alg, 'B6', 'p.dejong@email.nl')
    safe_set_cell(ws_alg, 'B7', '06-55443322')
    safe_set_cell(ws_alg, 'B10', 'Hoofdstraat 123')
    safe_set_cell(ws_alg, 'B11', '5678 CD')
    safe_set_cell(ws_alg, 'B12', 'Rotterdam')
    
    # Dates
    start_date = datetime(2024, 3, 1)
    end_date = datetime(2024, 11, 30)
    safe_set_cell(ws_alg, 'B15', start_date.strftime('%d-%m-%Y'))
    safe_set_cell(ws_alg, 'B16', end_date.strftime('%d-%m-%Y'))
    
    # Financial
    safe_set_cell(ws_alg, 'B19', 1800)  # Huur per maand
    safe_set_cell(ws_alg, 'B20', 9)     # Aantal maanden
    safe_set_cell(ws_alg, 'B21', 2)     # Borg (aantal maanden huur)
    
    # Selections
    safe_set_cell(ws_alg, 'B24', 'Ja, inclusief GWE detail')
    safe_set_cell(ws_alg, 'B25', 'Intensief Schoonmaak')
    
    print("✓ Filled Algemeen sheet")
    
    # Fill GWE_Detail - HEAVY usage (overages in all categories)
    ws_gwe = wb['GWE_Detail']
    safe_set_cell(ws_gwe, 'B4', 10000)  # Elektra Begin
    safe_set_cell(ws_gwe, 'B5', 13500)  # Elektra Eind (3500 kWh - way over!)
    safe_set_cell(ws_gwe, 'B7', 5000)   # Gas Begin
    safe_set_cell(ws_gwe, 'B8', 6800)   # Gas Eind (1800 m³ - way over!)
    safe_set_cell(ws_gwe, 'B10', 2000)  # Water Begin
    safe_set_cell(ws_gwe, 'B11', 2250)  # Water Eind (250 m³ - way over!)
    
    # Tarieven
    safe_set_cell(ws_gwe, 'B14', 0.30)  # Elektra Tarief
    safe_set_cell(ws_gwe, 'B15', 1.50)  # Gas Tarief
    safe_set_cell(ws_gwe, 'B16', 1.65)  # Water Tarief
    
    # Budgets (set low to ensure overage)
    safe_set_cell(ws_gwe, 'B19', 2000)  # Elektra Budget (usage will exceed)
    safe_set_cell(ws_gwe, 'B20', 1000)  # Gas Budget (usage will exceed)
    safe_set_cell(ws_gwe, 'B21', 150)   # Water Budget (usage will exceed)
    
    print("✓ Filled GWE_Detail sheet (heavy overuse in all categories!)")
    
    # Fill Schoonmaak - Lots of extra hours
    ws_schoon = wb['Schoonmaak']
    safe_set_cell(ws_schoon, 'B6', 15.0)  # Totaal Uren Gewerkt (way more than 7 included)
    safe_set_cell(ws_schoon, 'B10', 38)   # Uurtarief
    
    print("✓ Filled Schoonmaak sheet (heavy overtime)")
    
    # Fill Schade sheet with moderate damages
    ws_schade = wb['Schade']
    
    damages = [
        ('Beschadigde vloer', 1, 200.00),
        ('Schade behang', 1, 120.00),
        ('Kapotte lamp', 1, 65.00),
        ('Gebroken deurklink', 2, 40.00),
        ('Vlekken tapijt (reiniging)', 1, 150.00),
    ]
    
    start_row = 6
    for i, (beschrijving, aantal, tarief) in enumerate(damages):
        row = start_row + i
        safe_set_cell(ws_schade, f'A{row}', beschrijving)
        safe_set_cell(ws_schade, f'B{row}', aantal)
        safe_set_cell(ws_schade, f'C{row}', tarief)
    
    print(f"✓ Added {len(damages)} damage items to Schade sheet")
    
    # Save
    wb.save('test_pot_overuse.xlsx')
    print(f"\n✓ Created test_pot_overuse.xlsx successfully!")
    print(f"  - HEAVY GWE overuse (all budgets exceeded!)")
    print(f"  - Elektra: 3500 kWh vs 2000 budget")
    print(f"  - Gas: 1800 m³ vs 1000 budget")
    print(f"  - Water: 250 m³ vs 150 budget")
    print(f"  - Heavy cleaning overtime (15 hrs vs 7 included)")
    print(f"  - {len(damages)} moderate damage items")

if __name__ == '__main__':
    create_pot_overuse_test()
