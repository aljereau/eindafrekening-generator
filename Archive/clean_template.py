#!/usr/bin/env python3
"""
Script to clean input_template.xlsx by removing sample data while keeping headers and formulas.
"""
import openpyxl
from openpyxl.styles import Protection
import shutil
from pathlib import Path

def clean_template():
    """Clean the template by removing sample data from input fields."""
    
    # Backup original
    input_file = Path('input_template.xlsx')
    backup_file = Path('input_template_backup.xlsx')
    
    if input_file.exists():
        shutil.copy(input_file, backup_file)
        print(f"✓ Backup created: {backup_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    print(f"✓ Loaded workbook with sheets: {wb.sheetnames}")
    
    # Define cells to clear (based on the structure - cells marked with * are required inputs)
    # We keep headers/labels but clear the input values
    cells_to_clear = {
        'Algemeen': [
            'B4',  # Naam Klant
            'B5',  # Contactpersoon
            'B6',  # Email
            'B7',  # Telefoonnummer
            'B10', # Straat + Huisnummer
            'B11', # Postcode
            'B12', # Stad
            'B15', # Verhuurperiode Van
            'B16', # Verhuurperiode Tot
            'B19', # Huur per maand (incl. BTW)
            'B20', # Aantal maanden
            'B21', # Borg (aantal maanden huur)
            'B24', # GWE_detail_gebruik
            'B25', # Schoonmaak_pakket_type
        ],
        'GWE_Detail': [
            'B4',  # Elektra Begin
            'B5',  # Elektra Eind
            'B7',  # Gas Begin
            'B8',  # Gas Eind
            'B10', # Water Begin
            'B11', # Water Eind
            'B14', # Elektra Tarief
            'B15', # Gas Tarief
            'B16', # Water Tarief
            'B19', # Elektra Budget
            'B20', # Gas Budget
            'B21', # Water Budget
        ],
        'Schoonmaak': [
            'B6',  # Totaal Uren Gewerkt
            'B10', # Uurtarief Schoonmaak
        ],
        'Schade': [
            # Clear rows 6-20 for damage items (columns A, B, C)
            # Don't clear column D as it has formulas
        ]
    }
    
    # Clear specified cells
    for sheet_name, cells in cells_to_clear.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cleared = 0
            for cell in cells:
                try:
                    # Check if it's a merged cell
                    if not isinstance(ws[cell], openpyxl.cell.cell.MergedCell):
                        ws[cell].value = None
                        cleared += 1
                except Exception as e:
                    print(f"  Warning: Could not clear {cell}: {e}")
            print(f"✓ Cleared {cleared} cells in sheet '{sheet_name}'")
    
    # Clear damage items in Schade sheet (rows 6-20, columns A, B, C only)
    if 'Schade' in wb.sheetnames:
        ws = wb['Schade']
        cleared_count = 0
        for row in range(6, 21):  # rows 6-20
            for col in ['A', 'B', 'C']:  # Don't clear D (has formulas)
                cell = f"{col}{row}"
                if ws[cell].value and not str(ws[cell].value).startswith('='):
                    ws[cell].value = None
                    cleared_count += 1
        print(f"✓ Cleared {cleared_count} damage item cells in sheet 'Schade'")
    
    # Save cleaned template
    wb.save(input_file)
    print(f"\n✓ Template cleaned successfully: {input_file}")
    print(f"✓ Backup saved as: {backup_file}")

if __name__ == '__main__':
    clean_template()
