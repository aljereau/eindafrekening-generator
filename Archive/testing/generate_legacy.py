#!/usr/bin/env python3
"""
RyanRent Eindafrekening Generator
Genereert visuele eindafrekening PDF vanuit Excel input
"""

import openpyxl
from jinja2 import Template, Environment, FileSystemLoader
# from weasyprint import HTML  <-- Moved to inside function to avoid import error on Mac if libs missing
from datetime import datetime
import os
import sys

import argparse

def read_excel_data(filepath='input.xlsx'):
    """
    Lees data uit Excel
    Returns: dict met alle benodigde data
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel bestand '{filepath}' niet gevonden.")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        raise ValueError(f"Kon Excel bestand niet openen: {e}")
    
    if 'Hoofdgegevens' not in wb.sheetnames:
        raise ValueError("Excel bestand mist tabblad 'Hoofdgegevens'. Controleer het sjabloon.")
        
    sheet = wb['Hoofdgegevens']
    
    # Helper to safely get float value
    def get_float(cell):
        val = cell.value
        if val is None: return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            print(f"⚠️  Waarschuwing: Kon waarde '{val}' in cel {cell.coordinate} niet lezen als getal. Gebruik 0.0")
            return 0.0

    # Helper to safely get string value
    def get_str(cell):
        val = cell.value
        return str(val).strip() if val is not None else ""

    data = {
        # Basis info
        'gast_naam': get_str(sheet['B1']),
        'property_naam': get_str(sheet['B2']),
        'property_adres': get_str(sheet['B3']),
        'periode_start': get_str(sheet['B4']),
        'periode_eind': get_str(sheet['B5']),
        'aantal_dagen': get_str(sheet['B6']),
        
        # Borg
        'borg_voorschot': get_float(sheet['B9']),
        'borg_gebruikt': get_float(sheet['B10']),
        
        # GWE
        'gwe_voorschot': get_float(sheet['B13']),
        'gwe_verbruik': get_float(sheet['B14']),
        'gwe_gas': get_float(sheet['B15']),
        'gwe_water': get_float(sheet['B16']),
        'gwe_electra_euro': get_float(sheet['B17']),
        'gwe_electra_kwh': get_float(sheet['B18']),
        
        # Schoonmaak
        'clean_voorschot': get_float(sheet['B21']),
        'clean_gebruikt': get_float(sheet['B22']),
        'clean_extra': get_float(sheet['B23']),
    }
    
    # Probeer schades sheet te lezen
    data['schades'] = []
    if 'Schades Detail' in wb.sheetnames:
        schades_sheet = wb['Schades Detail']
        for row in schades_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                try:
                    bedrag = float(row[1])
                    data['schades'].append({
                        'omschrijving': row[0],
                        'bedrag': bedrag
                    })
                except (ValueError, TypeError):
                    continue
    
    wb.close()
    return data

def calculate_values(data):
    """
    Bereken afgeleide waardes
    """
    # Borg berekeningen
    data['borg_terug'] = data['borg_voorschot'] - data['borg_gebruikt']
    if data['borg_voorschot'] > 0:
        data['borg_gebruikt_pct'] = min(100, (data['borg_gebruikt'] / data['borg_voorschot']) * 100)
    else:
        data['borg_gebruikt_pct'] = 0
    data['borg_terug_pct'] = max(0, 100 - data['borg_gebruikt_pct'])
    
    # GWE berekeningen
    data['gwe_extra'] = data['gwe_verbruik'] - data['gwe_voorschot']
    if data['gwe_extra'] > 0:
        data['gwe_is_overfilled'] = True
        if data['gwe_voorschot'] > 0:
            # Extra percentage is relative to the budget width (100%)
            data['gwe_extra_pct'] = (data['gwe_extra'] / data['gwe_voorschot']) * 100
        else:
            data['gwe_extra_pct'] = 100 # Fallback if no budget but usage exists
    else:
        data['gwe_is_overfilled'] = False
        if data['gwe_voorschot'] > 0:
            data['gwe_gebruikt_pct'] = (data['gwe_verbruik'] / data['gwe_voorschot']) * 100
        else:
            data['gwe_gebruikt_pct'] = 0
        data['gwe_terug_pct'] = max(0, 100 - data['gwe_gebruikt_pct'])
        data['gwe_terug'] = data['gwe_voorschot'] - data['gwe_verbruik']
    
    # Schoonmaak berekeningen
    data['clean_extra_calc'] = data['clean_gebruikt'] - data['clean_voorschot']
    if data['clean_extra_calc'] > 0:
        data['clean_is_overfilled'] = True
        if data['clean_voorschot'] > 0:
            data['clean_extra_pct'] = (data['clean_extra_calc'] / data['clean_voorschot']) * 100
        else:
            data['clean_extra_pct'] = 100
    else:
        data['clean_is_overfilled'] = False
        if data['clean_voorschot'] > 0:
            data['clean_gebruikt_pct'] = (data['clean_gebruikt'] / data['clean_voorschot']) * 100
        else:
            data['clean_gebruikt_pct'] = 0
        data['clean_terug_pct'] = max(0, 100 - data['clean_gebruikt_pct'])
        data['clean_terug'] = data['clean_voorschot'] - data['clean_gebruikt']
    
    # Netto berekening
    netto = 0
    
    # Borg
    netto += data['borg_terug']

    # GWE
    if data.get('gwe_extra', 0) > 0:
        netto -= data['gwe_extra']
    elif data.get('gwe_terug', 0) > 0:
        netto += data['gwe_terug']
        
    # Cleaning
    if data.get('clean_extra_calc', 0) > 0:
        netto -= data['clean_extra_calc']
    elif data.get('clean_terug', 0) > 0:
        netto += data['clean_terug']
    
    data['netto'] = netto
    data['netto_is_positive'] = netto >= 0
    
    # Datum gegenereerd
    data['generated_date'] = datetime.now().strftime('%d-%m-%Y %H:%M')
    
    # Logo encoding
    import base64
    logo_path = os.path.join('assets', 'ryanrent_co.jpg')
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            data['logo_b64'] = f"data:image/jpeg;base64,{encoded_string}"
    else:
        data['logo_b64'] = None
        print(f"⚠️  Waarschuwing: Logo niet gevonden op {logo_path}")
    
    return data

def generate_pdf(data, output_filename=None):
    """
    Genereer PDF vanuit template
    """
    # Setup Jinja2 environment
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template('template.html')
    
    # Add helper functions to data
    data['abs'] = abs
    
    html_filled = template.render(**data)
    
    # Genereer output filename als niet gegeven
    if not output_filename:
        safe_name = str(data['gast_naam']).replace(' ', '_').replace('.', '')
        output_filename = f"eindafrekening_{safe_name}_{data['periode_start']}-{data['periode_eind']}"
    
    # Ensure output directory exists
    if not os.path.exists('output'):
        os.makedirs('output')

    # Save HTML first
    html_path = os.path.join('output', output_filename + '.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_filled)
    
    pdf_path = os.path.join('output', output_filename + '.pdf')
    
    # Genereer PDF met WeasyPrint
    try:
        from weasyprint import HTML
        HTML(string=html_filled, base_url='.').write_pdf(pdf_path)
        return pdf_path
    except (OSError, ImportError) as e:
        print(f"\n⚠️  Waarschuwing: Kon PDF niet genereren (systeembibliotheek ontbreekt: {e})")
        print(f"   De HTML versie is wel bewaard: {html_path}")
        return html_path
    except Exception as e:
        print(f"\n⚠️  Waarschuwing: Kon PDF niet genereren ({e})")
        print(f"   De HTML versie is wel bewaard: {html_path}")
        return html_path

def main():
    """
    Main functie - run complete flow
    """
    parser = argparse.ArgumentParser(description='RyanRent Eindafrekening Generator')
    parser.add_argument('--input', default='input.xlsx', help='Pad naar Excel bestand')
    parser.add_argument('--no-pause', action='store_true', help='Voorkom pauze aan het einde (voor scripts)')
    args = parser.parse_args()

    print("🏠 RyanRent Eindafrekening Generator")
    print("=" * 50)
    
    try:
        # Stap 1: Lees Excel
        print(f"\n📊 Excel data inlezen van: {args.input}")
        data = read_excel_data(args.input)
        print(f"   ✓ Gast: {data['gast_naam']}")
        print(f"   ✓ Property: {data['property_naam']}")
        print(f"   ✓ Periode: {data['periode_start']} - {data['periode_eind']}")
        
        # Stap 2: Bereken waardes
        print("\n🔢 Waardes berekenen...")
        data = calculate_values(data)
        print(f"   ✓ Borg: €{data['borg_terug']:.2f} terug")
        
        gwe_msg = f"€{abs(data.get('gwe_extra', 0)):.2f} extra" if data.get('gwe_is_overfilled') else f"€{data.get('gwe_terug', 0):.2f} terug"
        print(f"   ✓ GWE: {gwe_msg}")
        
        clean_msg = f"€{abs(data.get('clean_extra_calc', 0)):.2f} extra" if data.get('clean_is_overfilled') else f"€{data.get('clean_terug', 0):.2f} terug"
        print(f"   ✓ Schoonmaak: {clean_msg}")
        
        netto_msg = "terug" if data['netto_is_positive'] else "bijbetalen"
        print(f"   ✓ Netto: €{abs(data['netto']):.2f} {netto_msg}")
        
        # Stap 3: Genereer PDF
        print("\n📄 PDF genereren...")
        output_path = generate_pdf(data)
        
        if output_path.endswith('.pdf'):
            print(f"   ✓ PDF opgeslagen: {output_path}")
            print("\n✅ Klaar! PDF is gegenereerd.")
        else:
            print(f"   ✓ HTML opgeslagen: {output_path}")
            print("\n✅ Klaar! HTML is gegenereerd (PDF mislukt).")
            
        print(f"\n📂 Locatie: {os.path.abspath(output_path)}")
        
    except FileNotFoundError as e:
        print(f"\n❌ Fout: {e}")
    except ValueError as e:
        print(f"\n❌ Fout in data: {e}")
    except Exception as e:
        print(f"\n❌ Onverwachte fout: {e}")
        import traceback
        traceback.print_exc()
    
    if not args.no_pause:
        input("\nDruk op Enter om af te sluiten...")

if __name__ == "__main__":
    main()
