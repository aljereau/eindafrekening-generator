#!/usr/bin/env python3
"""
Complete cleanup script for input_template.xlsx - removes ALL sample data
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import shutil
from pathlib import Path

def is_merged_cell(ws, cell_ref):
    """Check if a cell is part of a merged range."""
    cell = ws[cell_ref]
    return isinstance(cell, openpyxl.cell.cell.MergedCell)

def safe_clear_cell(ws, cell_ref):
    """Safely clear a cell value, handling merged cells."""
    cell = ws[cell_ref]
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        # Find the merged range and clear the top-left cell
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left = ws.cell(min_row, min_col)
                # Only clear if it's not a formula
                if not (isinstance(top_left.value, str) and str(top_left.value).startswith('=')):
                    top_left.value = None
                return
    else:
        # Only clear if it's not a formula
        if cell.value is not None and not (isinstance(cell.value, str) and str(cell.value).startswith('=')):
            cell.value = None

def clean_template_complete():
    """Complete cleanup - clear all non-formula data in column B."""
    
    # Backup original
    input_file = Path('input_template.xlsx')
    backup_file = Path('input_template_with_data.xlsx')
    
    if input_file.exists():
        shutil.copy(input_file, backup_file)
        print(f"✓ Backup created: {backup_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    print(f"✓ Loaded workbook with sheets: {wb.sheetnames}")
    
    total_cleared = 0
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cleared_in_sheet = 0
        
        # Clear all data cells in column B (rows 4-100)
        # This is where user inputs typically go
        for row in range(4, 101):
            cell_ref = f'B{row}'
            cell = ws[cell_ref]
            
            # Skip if it's a merged cell
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            
            # Clear if it has a value and it's not a formula
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Keep formulas
                    continue
                else:
                    # Clear data
                    cell.value = None
                    cleared_in_sheet += 1
        
        # Also clear data in columns A, B, C for Schade sheet (damage items)
        if sheet_name == 'Schade':
            for row in range(6, 31):  # rows 6-30
                for col in ['A', 'B', 'C']:
                    cell_ref = f'{col}{row}'
                    cell = ws[cell_ref]
                    
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    
                    if cell.value is not None:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            continue
                        else:
                            cell.value = None
                            cleared_in_sheet += 1
        
        total_cleared += cleared_in_sheet
        print(f"✓ Cleared {cleared_in_sheet} cells in '{sheet_name}'")
    
    # Save cleaned template
    wb.save(input_file)
    print(f"\n✓ Template cleaned successfully: {input_file}")
    print(f"✓ Total cells cleared: {total_cleared}")
    print(f"✓ Backup saved as: {backup_file}")

if __name__ == '__main__':
    clean_template_complete()
