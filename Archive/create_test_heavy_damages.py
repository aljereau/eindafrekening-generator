#!/usr/bin/env python3
"""
Create test_heavy_damages.xlsx - A test scenario with many damage items
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import shutil
from pathlib import Path
from datetime import datetime, timedelta

def safe_set_cell(ws, cell_ref, value):
    """Safely set a cell value, handling merged cells."""
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        # Find the merged range and set value on the top-left cell
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                # Get top-left cell of merged range
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left = ws.cell(min_row, min_col)
                top_left.value = value
                return
    else:
        ws[cell_ref].value = value


def create_heavy_damages_test():
    """Create a test template with heavy damages."""
    
    # Copy clean template
    shutil.copy('input_template.xlsx', 'test_heavy_damages.xlsx')
    print("✓ Created test_heavy_damages.xlsx from clean template")
    
    # Load workbook
    wb = openpyxl.load_workbook('test_heavy_damages.xlsx')
    
    # Fill Algemeen sheet
    ws_alg = wb['Algemeen']
    safe_set_cell(ws_alg, 'B4', 'Familie Vermeulen')
    safe_set_cell(ws_alg, 'B5', 'Lisa van RyanRent')
    safe_set_cell(ws_alg, 'B6', 'j.vermeulen@email.nl')
    safe_set_cell(ws_alg, 'B7', '06-98765432')
    safe_set_cell(ws_alg, 'B10', 'Kerkstraat 45')
    safe_set_cell(ws_alg, 'B11', '1234 AB')
    safe_set_cell(ws_alg, 'B12', 'Amsterdam')
    
    # Dates
    start_date = datetime(2024, 1, 1)
    end_date = datetime(2024, 12, 31)
    safe_set_cell(ws_alg, 'B15', start_date.strftime('%d-%m-%Y'))
    safe_set_cell(ws_alg, 'B16', end_date.strftime('%d-%m-%Y'))
    
    # Financial
    safe_set_cell(ws_alg, 'B19', 1500)  # Huur per maand
    safe_set_cell(ws_alg, 'B20', 12)    # Aantal maanden
    safe_set_cell(ws_alg, 'B21', 2)     # Borg (aantal maanden huur)
    
    # Selections
    safe_set_cell(ws_alg, 'B24', 'Ja, inclusief GWE detail')
    safe_set_cell(ws_alg, 'B25', 'Basis Schoonmaak')
    
    print("✓ Filled Algemeen sheet")
    
    # Fill GWE_Detail - Normal usage (no overage)
    ws_gwe = wb['GWE_Detail']
    safe_set_cell(ws_gwe, 'B4', 10000)  # Elektra Begin
    safe_set_cell(ws_gwe, 'B5', 11500)  # Elektra Eind (1500 kWh)
    safe_set_cell(ws_gwe, 'B7', 5000)   # Gas Begin
    safe_set_cell(ws_gwe, 'B8', 5600)   # Gas Eind (600 m³)
    safe_set_cell(ws_gwe, 'B10', 2000)  # Water Begin
    safe_set_cell(ws_gwe, 'B11', 2080)  # Water Eind (80 m³)
    
    # Tarieven
    safe_set_cell(ws_gwe, 'B14', 0.28)  # Elektra Tarief
    safe_set_cell(ws_gwe, 'B15', 1.45)  # Gas Tarief
    safe_set_cell(ws_gwe, 'B16', 1.60)  # Water Tarief
    
    # Budgets (within limits)
    safe_set_cell(ws_gwe, 'B19', 1800)  # Elektra Budget
    safe_set_cell(ws_gwe, 'B20', 800)   # Gas Budget
    safe_set_cell(ws_gwe, 'B21', 100)   # Water Budget
    
    print("✓ Filled GWE_Detail sheet (normal usage)")
    
    # Fill Schoonmaak - Some extra hours
    ws_schoon = wb['Schoonmaak']
    safe_set_cell(ws_schoon, 'B6', 8.5)   # Totaal Uren Gewerkt
    safe_set_cell(ws_schoon, 'B10', 35)   # Uurtarief
    
    print("✓ Filled Schoonmaak sheet")
    
    # Fill Schade sheet with MANY damages
    ws_schade = wb['Schade']
    
    damages = [
        ('Kapotte deurklink slaapkamer', 1, 45.00),
        ('Beschadigde vloer woonkamer', 3, 125.00),
        ('Gebroken spiegel badkamer', 1, 89.00),
        ('Vlekken op bank (reiniging)', 1, 180.00),
        ('Gat in muur (reparatie + schilderen)', 2, 95.00),
        ('Kapotte gordijnrail', 1, 55.00),
        ('Beschadigde keukenkast deur', 1, 110.00),
        ('Gebroken tegel badkamer', 4, 25.00),
        ('Schade aan laminaat gang', 1, 250.00),
        ('Kapotte lamp plafond', 2, 45.00),
        ('Beschadigde vensterbank', 1, 85.00),
        ('Schade aan behang slaapkamer', 1, 150.00),
        ('Gebroken radiatorknop', 1, 35.00),
        ('Beschadigde douchekop', 1, 75.00),
        ('Kapotte keukenkraan', 1, 135.00),
    ]
    
    start_row = 6
    for i, (beschrijving, aantal, tarief) in enumerate(damages):
        row = start_row + i
        safe_set_cell(ws_schade, f'A{row}', beschrijving)
        safe_set_cell(ws_schade, f'B{row}', aantal)
        safe_set_cell(ws_schade, f'C{row}', tarief)
    
    print(f"✓ Added {len(damages)} damage items to Schade sheet")
    
    # Save
    wb.save('test_heavy_damages.xlsx')
    print(f"\n✓ Created test_heavy_damages.xlsx successfully!")
    print(f"  - Normal GWE usage (within budgets)")
    print(f"  - {len(damages)} damage items")
    print(f"  - Some extra cleaning hours")

if __name__ == '__main__':
    create_heavy_damages_test()
