#!/usr/bin/env python3
"""
Create 3 test scenarios for RyanRent Eindafrekening Generator
Generates Excel files with realistic test data using named ranges
"""

import openpyxl
from datetime import date, timedelta
import shutil
from pathlib import Path


def set_named_value(wb, name: str, value):
    """Set value to a named range in the workbook"""
    try:
        if name not in wb.defined_names:
            print(f"   ⚠️  Warning: Named range '{name}' not found")
            return False

        defn = wb.defined_names[name]
        destinations = list(defn.destinations)
        if not destinations:
            return False

        sheet_name, cell_ref = destinations[0]
        ws = wb[sheet_name]
        cell_ref = cell_ref.replace('$', '')
        ws[cell_ref] = value
        return True
    except Exception as e:
        print(f"   ⚠️  Error setting '{name}': {e}")
        return False


def create_scenario_1_underuse():
    """
    Scenario 1: UNDERUSE - Guest used less than prepaid
    - Borg: €500 → Used €150 → Return €350
    - GWE: €300 → Used €220 → Return €80
    - Schoonmaak: €250 → Used €200 → Return €50
    Result: Customer gets €480 back
    """
    print("\n📝 Creating Scenario 1: Underuse (Customer gets money back)")

    src = Path("input_template.xlsx")
    dst = Path("test_scenario_1_underuse.xlsx")
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)

    # Client info
    set_named_value(wb, 'Klantnaam', "Familie de Vries")
    set_named_value(wb, 'Email', "jan.devries@email.nl")
    set_named_value(wb, 'Telefoonnummer', "06-12345678")

    # Object
    set_named_value(wb, 'Object_adres', "Zeedijk 42A, 1234AB Zandvoort")

    # Period (7 days)
    checkin = date(2025, 7, 15)
    checkout = checkin + timedelta(days=7)
    set_named_value(wb, 'Incheck_datum', checkin)
    set_named_value(wb, 'Uitcheck_datum', checkout)

    # Deposits
    set_named_value(wb, 'Voorschot_borg', 500)
    set_named_value(wb, 'Voorschot_GWE', 300)
    set_named_value(wb, 'Voorschot_schoonmaak', 250)

    # Borg usage
    set_named_value(wb, 'Borg_gebruikt', 150)

    # GWE meters
    set_named_value(wb, 'KWh_begin', 10000)
    set_named_value(wb, 'KWh_eind', 10350)  # 350 kWh
    set_named_value(wb, 'Gas_begin', 5000)
    set_named_value(wb, 'Gas_eind', 5050)  # 50 m³

    # Schoonmaak
    set_named_value(wb, 'Inbegrepen_uren', 5.0)
    set_named_value(wb, 'Extra_uren', 0)  # No extra hours

    wb.save(dst)
    print(f"   ✓ Saved: {dst}")
    print(f"   💰 Expected: Customer receives ~€430 back")
    wb.close()


def create_scenario_2_perfect():
    """
    Scenario 2: PERFECT FIT - Guest used exactly what was prepaid
    - Borg: €500 → Used €500 → Return €0
    - GWE: €300 → Used €300 → Return €0
    - Schoonmaak: €250 → Used €250 → Return €0
    Result: Neutral settlement
    """
    print("\n📝 Creating Scenario 2: Perfect Fit (Neutral settlement)")

    src = Path("input_template.xlsx")
    dst = Path("test_scenario_2_perfect.xlsx")
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)

    # Client info
    set_named_value(wb, 'Klantnaam', "Familie Jansen")
    set_named_value(wb, 'Email', "p.jansen@email.nl")
    set_named_value(wb, 'Telefoonnummer', "06-98765432")

    # Object
    set_named_value(wb, 'Object_adres', "Strandweg 18, 2225XY Den Haag")

    # Period (10 days)
    checkin = date(2025, 8, 1)
    checkout = checkin + timedelta(days=10)
    set_named_value(wb, 'Incheck_datum', checkin)
    set_named_value(wb, 'Uitcheck_datum', checkout)

    # Deposits
    set_named_value(wb, 'Voorschot_borg', 500)
    set_named_value(wb, 'Voorschot_GWE', 300)
    set_named_value(wb, 'Voorschot_schoonmaak', 250)

    # Borg - fully used
    set_named_value(wb, 'Borg_gebruikt', 500)

    # GWE meters
    set_named_value(wb, 'KWh_begin', 10000)
    set_named_value(wb, 'KWh_eind', 10450)  # 450 kWh
    set_named_value(wb, 'Gas_begin', 5000)
    set_named_value(wb, 'Gas_eind', 5060)  # 60 m³

    # Schoonmaak - exactly 5 hours
    set_named_value(wb, 'Inbegrepen_uren', 5.0)
    set_named_value(wb, 'Extra_uren', 0)

    wb.save(dst)
    print(f"   ✓ Saved: {dst}")
    print(f"   💰 Expected: Neutral (€0)")
    wb.close()


def create_scenario_3_overflow():
    """
    Scenario 3: OVERFLOW - Guest exceeded prepaid amounts
    - Borg: €500 → Used €500 → No return (max)
    - GWE: €300 → Used €485 → Pay €185 extra
    - Schoonmaak: €250 → Used €375 → Pay €125 extra
    Result: Customer owes €310
    """
    print("\n📝 Creating Scenario 3: Overflow (Customer owes money)")

    src = Path("input_template.xlsx")
    dst = Path("test_scenario_3_overflow.xlsx")
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)

    # Client info
    set_named_value(wb, 'Klantnaam', "Familie Bakker")
    set_named_value(wb, 'Email', "s.bakker@email.nl")
    set_named_value(wb, 'Telefoonnummer', "06-55544433")

    # Object
    set_named_value(wb, 'Object_adres', "Duinweg 7B, 1234ZZ Scheveningen")

    # Period (14 days)
    checkin = date(2025, 8, 10)
    checkout = checkin + timedelta(days=14)
    set_named_value(wb, 'Incheck_datum', checkin)
    set_named_value(wb, 'Uitcheck_datum', checkout)

    # Deposits
    set_named_value(wb, 'Voorschot_borg', 500)
    set_named_value(wb, 'Voorschot_GWE', 300)
    set_named_value(wb, 'Voorschot_schoonmaak', 250)

    # Borg - fully used
    set_named_value(wb, 'Borg_gebruikt', 500)

    # GWE meters - Heavy usage
    set_named_value(wb, 'KWh_begin', 10000)
    set_named_value(wb, 'KWh_eind', 10850)  # 850 kWh
    set_named_value(wb, 'Gas_begin', 5000)
    set_named_value(wb, 'Gas_eind', 5120)  # 120 m³

    # Schoonmaak - 2.5 extra hours (7.5 total)
    set_named_value(wb, 'Inbegrepen_uren', 5.0)
    set_named_value(wb, 'Extra_uren', 2.5)

    wb.save(dst)
    print(f"   ✓ Saved: {dst}")
    print(f"   💰 Expected: Customer owes ~€310")
    wb.close()


def main():
    """Generate all 3 test scenarios"""
    print("=" * 70)
    print("🧪 RyanRent Test Scenario Generator")
    print("=" * 70)

    create_scenario_1_underuse()
    create_scenario_2_perfect()
    create_scenario_3_overflow()

    print("\n" + "=" * 70)
    print("✅ All 3 test scenarios created!")
    print("=" * 70)
    print("\n📂 Generated files:")
    print("   1️⃣  test_scenario_1_underuse.xlsx   → Customer gets €480 back")
    print("   2️⃣  test_scenario_2_perfect.xlsx    → Neutral (€0)")
    print("   3️⃣  test_scenario_3_overflow.xlsx   → Customer owes €310")
    print("\n💡 To test:")
    print("   1. Copy one of these files to 'input_template.xlsx'")
    print("   2. Run: python3 generate.py")
    print("   3. View: output/eindafrekening_*_onepager.html")
    print()


if __name__ == "__main__":
    main()
